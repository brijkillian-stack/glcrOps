"""
schedule_parser.py — Parse the GLCR weekly schedule Excel into per-date availability pools.

Mirrors the parsing logic in fill_engine.py but returns display names matched
against the Supabase `entities` list instead of roster keys, and returns plain
Python dicts that Reflex state can store and serialize.

Pools returned:
  grave  — TMs on the overnight grave shift (available full 11pm–7am)
  pm_ol  — TMs on a swing shift ending at 1am (available 11pm–1am)
  am_ol  — TMs on a day shift starting at 5am (available 5am–7am on the PRIOR date)
"""

from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl

GLCR_BASE = Path(__file__).resolve().parent / "engine"
SCHEDULE_DIR = GLCR_BASE / "Inputs" / "Weekly Schedules"

_DAY_NAME_ROW = 6
_DATE_ROW     = 7
_DATA_ROW     = 8


# ── Cell value helpers (same rules as fill_engine.py) ────────────────────────

def _working(v) -> bool:
    """True if the cell represents an actual shift (not OFF/PTO/etc.)."""
    if v is None:
        return False
    s = str(v).strip().upper()
    return bool(s) and s not in ("OFF", "PTO HOURLY", "PTO", "MDL", "PTO HOURLY MDL") and "PTO" not in s


def _is_5am(v) -> bool:
    return v is not None and " 5:" in str(v)


def _is_1am(v) -> bool:
    return v is not None and "1:00A" in str(v)


# ── Schedule file discovery ───────────────────────────────────────────────────

def get_latest_schedule_path() -> Optional[Path]:
    """Return the most recently modified .xlsx in Weekly Schedules/.

    On Render the local Inputs/ folder is ephemeral, so if it's empty we
    sync from Supabase Storage first. This keeps `Schedule loaded` indicator
    accurate after a container restart.
    """
    SCHEDULE_DIR.mkdir(parents=True, exist_ok=True)

    candidates = sorted(
        SCHEDULE_DIR.glob("*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    # Fall back to Storage if the local folder is empty (typical after
    # a Render redeploy).
    if not candidates:
        try:
            from shared import storage
            storage.sync_schedules_to(SCHEDULE_DIR)
            candidates = sorted(
                SCHEDULE_DIR.glob("*.xlsx"),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )
        except Exception:
            # Network hiccup or bucket missing — fall through and return None.
            pass

    return candidates[0] if candidates else None


def list_schedule_files() -> list[str]:
    """Return schedule filenames sorted newest-first."""
    if not SCHEDULE_DIR.exists():
        return []
    return [
        p.name for p in sorted(
            SCHEDULE_DIR.glob("*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
    ]


# ── Name matching ─────────────────────────────────────────────────────────────

def _build_name_lookup(entities: list[dict]) -> dict[str, list[str]]:
    """
    {first_name_lower: [display_name, ...]}
    Built from the Supabase entities list so name resolution is consistent
    with the rest of the web app.
    """
    lookup: dict[str, list[str]] = {}
    for e in entities:
        dn = e.get("display_name", "").strip()
        if not dn:
            continue
        first = dn.split()[0].lower()
        lookup.setdefault(first, []).append(dn)
    return lookup


def _match_name(first: str, last: str, lookup: dict[str, list[str]]) -> Optional[str]:
    """
    Match a schedule row's (first, last) to a display_name.
    Uses last-name initial when first name is ambiguous.
    Returns None if no match is found.
    """
    if not first or first.lower() in ("first name", "none", ""):
        return None
    fn = first.strip().lower()
    candidates = lookup.get(fn, [])
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    # Multiple display names share this first name — use last initial to pick
    ln = (last or "").strip()
    if ln:
        ln_initial = ln[0].upper()
        for dn in candidates:
            parts = dn.split()
            # display_name format is either "First" or "First L"
            if len(parts) > 1 and parts[-1].upper().startswith(ln_initial):
                return dn
    return candidates[0]   # fallback: first match


# ── Per-shift roster extractor (Phase L.2 enhancement) ───────────────────────
# Reads ALL three sheets — Days, Swings, Graves — and returns who's working,
# on PTO, on medical leave, etc. per shift per date. Powers the recap's
# Team Updates auto-populate so Days/Swings/Graves call-outs and PTO are
# all surfaced from one schedule file.

# Off-cell reason taxonomy. We categorize so the recap can format the
# right phrase ("on PTO", "on MDL", "scheduled off") rather than dumping
# raw cell values.
_PTO_VARIANTS = (
    "PTO HOURLY",
    "INCENTIVE PTO",
    "FUTURE PTO APPROVED",
    "PTO",
)
_MDL_VARIANTS = ("MDL", "MEDICAL LEAVE")
_PLAIN_OFF = ("OFF",)


def _classify_off_value(s: str) -> str:
    """Bucket a non-time-range cell value into a leave category.

    Returns one of: "scheduled_off" | "pto" | "mdl" | "other" | ""
    "" means the cell is junk (a stray number, header echo, etc.) and
    should be ignored entirely.
    """
    if not s:
        return ""
    upper = s.strip().upper()
    if not upper:
        return ""
    if any(v in upper for v in _PTO_VARIANTS):
        return "pto"
    if any(v in upper for v in _MDL_VARIANTS):
        return "mdl"
    if upper in _PLAIN_OFF:
        return "scheduled_off"
    # Day name echoes from row 6 / 7 — ignore.
    if upper in ("FRIDAY", "SATURDAY", "SUNDAY", "MONDAY",
                 "TUESDAY", "WEDNESDAY", "THURSDAY"):
        return ""
    # Pure numbers / fragments — ignore.
    digits = upper.replace(".", "").replace(",", "").strip()
    if digits.isdigit():
        return ""
    return "other"


def _detect_shift_label(ws) -> str:
    """Read row 6 of the sheet to detect shift label ('day' / 'swing' / 'grave').
    Returns "" if the sheet doesn't follow the expected GLCR template."""
    try:
        row6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    except Exception:
        return ""
    for cell in row6:
        if cell is None:
            continue
        s = str(cell).strip().upper()
        if "DAY SHIFT" in s:
            return "days"
        if "SWING SHIFT" in s:
            return "swings"
        if "GRAVE SHIFT" in s:
            return "graves"
    return ""


def peek_shift_rosters_for_date(source, target_date: str) -> dict:
    """Return per-shift roster status for a single date.

    Args:
        source:      Path to a local xlsx file or raw bytes.
        target_date: ISO date string ("2026-05-04") to look up.

    Returns:
        {
          "days":   {"working": [...], "pto": [...], "mdl": [...], "other": [...]},
          "swings": {...},
          "graves": {...},
        }
        Names are display-name format ("Brian K" / "Cookie") — caller is
        responsible for matching to entities if it needs tm_ids.
        Returns empty buckets for shifts whose sheet is missing or unreadable.
    """
    import io
    out = {
        "days":   {"working": [], "pto": [], "mdl": [], "other": []},
        "swings": {"working": [], "pto": [], "mdl": [], "other": []},
        "graves": {"working": [], "pto": [], "mdl": [], "other": []},
    }
    try:
        if isinstance(source, (bytes, bytearray)):
            wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True, read_only=True)
        else:
            wb = openpyxl.load_workbook(str(source), data_only=True, read_only=True)
    except Exception:
        return out

    target = target_date.strip()
    for sn in wb.sheetnames:
        ws = wb[sn]
        shift = _detect_shift_label(ws)
        if not shift:
            continue
        # Find the column index whose row-7 cell parses to target_date.
        try:
            date_row = list(ws.iter_rows(min_row=7, max_row=7, values_only=True))[0]
        except Exception:
            continue
        target_col = None
        for ci, dval in enumerate(date_row):
            if dval is None:
                continue
            try:
                if isinstance(dval, (date, datetime)):
                    d = dval.date() if isinstance(dval, datetime) else dval
                else:
                    s = str(dval).strip()
                    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
                    if not m:
                        continue
                    d = date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
                if d.isoformat() == target:
                    target_col = ci
                    break
            except Exception:
                continue
        if target_col is None:
            continue

        # Walk data rows, classify each TM's cell for this date
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
            if not row or len(row) <= max(3, target_col):
                continue
            first = str(row[2] or "").strip()
            last = str(row[3] or "").strip()
            if not first or first.lower() in ("first name", "none", ""):
                continue
            display = f"{first} {last[0]}".strip() if last else first
            cell = row[target_col]
            if cell is None:
                continue
            s = str(cell).strip()
            # Time range = working
            if " - " in s and ":" in s:
                out[shift]["working"].append(display)
                continue
            cat = _classify_off_value(s)
            if cat == "pto":
                out[shift]["pto"].append(display)
            elif cat == "mdl":
                out[shift]["mdl"].append(display)
            elif cat == "scheduled_off":
                # Plain "OFF" = pre-scheduled day off, NOT a call-out;
                # we exclude these from the recap entirely.
                continue
            elif cat == "other":
                out[shift]["other"].append(f"{display} ({s})")
    wb.close()
    return out


# ── Lightweight date extractor ────────────────────────────────────────────────
# Used by Phase H "create week from schedule" flow on the /zds/ index page.
# Reads ONLY the Sheet3 header dates — skips pool parsing for speed.

def peek_schedule_dates(source) -> Optional[dict]:
    """Extract the seven dates (and derived week_ending) from a schedule xlsx.

    Args:
        source: Either a Path to a local xlsx file or raw bytes.

    Returns:
        {
          "dates": ["2026-05-01", "2026-05-02", ..., "2026-05-07"],  # sorted asc
          "week_ending": "2026-05-07",                               # last date
        }
        or None if the file isn't parseable / is missing Sheet3.
    """
    import io
    try:
        if isinstance(source, (bytes, bytearray)):
            wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True, read_only=True)
        else:
            wb = openpyxl.load_workbook(str(source), data_only=True, read_only=True)
    except Exception:
        return None

    if "Sheet3" not in wb.sheetnames:
        wb.close()
        return None

    ws3 = wb["Sheet3"]
    hrows = list(ws3.iter_rows(
        min_row=_DAY_NAME_ROW, max_row=_DATE_ROW, values_only=True
    ))
    wb.close()
    if len(hrows) < 2:
        return None

    date_row = hrows[1]
    parsed: list[date] = []
    for dval in date_row:
        if not dval:
            continue
        try:
            if isinstance(dval, (date, datetime)):
                d = dval.date() if isinstance(dval, datetime) else dval
            else:
                s = str(dval).strip()
                m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
                if not m:
                    continue
                d = date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
            parsed.append(d)
        except Exception:
            continue

    if not parsed:
        return None

    parsed.sort()
    return {
        "dates":       [d.isoformat() for d in parsed],
        "week_ending": parsed[-1].isoformat(),
    }


# ── Main parser ───────────────────────────────────────────────────────────────

def parse_daily_pools(
    entities: list[dict],
    schedule_path: Optional[Path] = None,
) -> dict[str, dict[str, list[str]]]:
    """
    Parse the weekly schedule Excel and return per-date TM availability pools.

    Args:
        entities:      List of entity dicts from the Supabase `entities` table
                       (must have at least `display_name`).
        schedule_path: Optional explicit path to a schedule xlsx; defaults to
                       the most recently modified file in Weekly Schedules/.

    Returns:
        {
          "2026-04-25": {
            "grave":  ["Sheri O", "Joy", ...],
            "pm_ol":  ["Maria", ...],
            "am_ol":  ["James", ...],
          },
          ...
        }
        Empty dict if no schedule file exists or parsing fails.
    """
    path = schedule_path or get_latest_schedule_path()
    if not path or not path.exists():
        return {}

    lookup = _build_name_lookup(entities)

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    except Exception:
        return {}

    # ── Parse date columns from Sheet3 header rows ────────────────────────
    if "Sheet3" not in wb.sheetnames:
        wb.close()
        return {}

    ws3   = wb["Sheet3"]
    hrows = list(ws3.iter_rows(
        min_row=_DAY_NAME_ROW, max_row=_DATE_ROW, values_only=True
    ))
    if len(hrows) < 2:
        wb.close()
        return {}

    _date_row = hrows[1]
    date_cols: dict[str, int] = {}   # ISO date string → column index

    for ci, dval in enumerate(_date_row):
        if not dval:
            continue
        try:
            if isinstance(dval, (date, datetime)):
                d = dval.date() if isinstance(dval, datetime) else dval
            else:
                s = str(dval).strip()
                m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
                if not m:
                    continue
                d = date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
            date_cols[d.isoformat()] = ci
        except Exception:
            continue

    if not date_cols:
        wb.close()
        return {}

    # Initialise pool structure for every parsed date
    pools: dict[str, dict[str, list[str]]] = {
        ds: {"grave": [], "pm_ol": [], "am_ol": []} for ds in date_cols
    }

    # ── Sheet parser ─────────────────────────────────────────────────────
    def _parse(ws, pool_type: str):
        for row in ws.iter_rows(min_row=_DATA_ROW, values_only=True):
            if not row or len(row) < 4:
                continue
            first = str(row[2]).strip() if row[2] else ""
            last  = str(row[3]).strip() if row[3] else ""
            if not first or first in ("First Name", "None"):
                continue
            if "headcount" in first.lower():
                continue
            dn = _match_name(first, last, lookup)
            if not dn:
                continue
            for ds, ci in date_cols.items():
                cv = row[ci] if ci < len(row) else None
                if not _working(cv):
                    continue
                if pool_type == "grave":
                    if dn not in pools[ds]["grave"]:
                        pools[ds]["grave"].append(dn)
                elif pool_type == "pm_ol":
                    if _is_1am(cv) and dn not in pools[ds]["pm_ol"]:
                        pools[ds]["pm_ol"].append(dn)
                elif pool_type == "am_ol":
                    if _is_5am(cv):
                        # AM OL workers shift onto the PREVIOUS calendar date
                        try:
                            prev = (
                                date.fromisoformat(ds) - timedelta(days=1)
                            ).isoformat()
                        except Exception:
                            continue
                        if prev in pools and dn not in pools[prev]["am_ol"]:
                            pools[prev]["am_ol"].append(dn)

    _parse(ws3, "grave")
    if "Sheet2" in wb.sheetnames:
        _parse(wb["Sheet2"], "pm_ol")
    if "Sheet1" in wb.sheetnames:
        _parse(wb["Sheet1"], "am_ol")

    wb.close()
    return pools
