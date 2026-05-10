#!/usr/bin/env python3
"""
GLCR Zone Deployment Book — HTML/PDF Renderer (v2.3)
=====================================================

Reads:
    Week Overview - Filled - YYYY-MM-DD.xlsx

Writes:
    Zone Deployment Book - YYYY-MM-DD.html   (14 pages: 7 days × 2 pages each)
    Zone Deployment Book - YYYY-MM-DD.pdf    (matching print PDF)

Run:
    python3 render_deployment_book.py <input.xlsx> [output.html]

v2.3 (4/29/26 — redesign matches Brian's "Option A v2.3" mock):
  • Two pages per day — a Daily Deployment page followed by a Break Sheet.
  • Daily masthead: giant 82px tabular day number, day-name in --day-color,
    "Day N of 7", per-group break-bar dots, shift label + group key.
  • Cards: 3px top stripe in zone family color, small caps zone label,
    26px name, icon-prefixed task list. Empty zones with no canonical tasks
    render as italic "Support as needed".
  • Restrooms: split men's/women's with per-side group pill, hairline-divided
    task list at the bottom; cover-tags inline in the header.
  • Auxiliary: compact 6-card strip with 2px stripe and tight task list.
  • Overlaps: time-window left, 6 mini-cards right (no per-row accent).
  • Page-foot slug strip: colored swatch + "GLCR · Grave",
    centered date, page N / 14 indicator.
  • Coverage alerts (compute_alerts) now render as inline cover-tags
    instead of full-width crimson banners.
  • Utility Porters strip dropped from the daily page (still loaded; can
    be re-surfaced if/when the design needs it back).
  • Break Sheet page (page 2 per day): three columns by Group 1/2/3 break
    wave. Each row is name + assignment summary + slot badge, colored to
    match the slot's zone/aux/RR family.
  • Engine-side rotation, sweeper auto-assignment, skip priority,
    compute_alerts(), MANUAL_ALERTS, SWEEPER_REASSIGNMENTS, OVERLAP_OVERRIDES
    are unchanged — they continue to drive the data the renderer reads.

Page model:
    For each weekday Friday→Thursday, page (2N - 1) is the Daily Deployment
    and page 2N is the Break Sheet. Total: 14 pages.

Operational color taxonomy (zone families) intentionally overrides marketing
brand chrome — this is an internal print artifact, not a customer-facing
surface. Barlow display font and other brand-aligned typography choices
are preserved.
"""

import sys
import re
import base64
import json
import datetime as dt
from pathlib import Path
from openpyxl import load_workbook

# 5/6/26: DB helpers — replace Rules/*.json file reads (Phase F closure).
# shared/ lives at repo root (4 parents above this file).
_REPO_ROOT = Path(__file__).resolve().parent.parent.parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
from shared.db import (
    get_overlap_tasks_for_engine,
    get_training_schedule_from_db,
    get_engine_roster_from_db,
    get_engine_profiles_from_db,
    get_zone_tasks_for_engine,   # Phase 4i compat — still used by database.py
    list_tasks,                  # Phase 4k.1 — primary source for renderer
)

# --------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------

ZONE_COLOR = {
    1: "yellow", 2: "yellow", 3: "red", 4: "red", 5: "red",
    6: "pink", 7: "blue", 8: "brown", 9: "red", 10: "green",
}
RR_COLOR = {1: "yellow", 6: "pink", 7: "blue", 8: "brown", 10: "green"}

# v6 template day colors — used for both the day-circle and the week-strip
# is-current highlight. These are the deeper variants Brian baked into v6.
DAY_COLOR = {
    "Friday":    "#C13A14",   # deep orange-red
    "Saturday":  "#0065bf",   # deep blue
    "Sunday":    "#6A1B85",   # deep purple
    "Monday":    "#2E7D32",   # deep green
    "Tuesday":   "#B89708",   # deep gold/yellow
    "Wednesday": "#B7679A",   # rose/pink
    "Thursday":  "#313845",   # charcoal
}

# Group/break-rotation positional defaults
BG_ZONE = {1:1, 2:2, 3:3, 4:1, 5:2, 6:3, 7:1, 8:2, 9:3, 10:2}
BG_RR_M = {1:2, 6:2, 7:3, 8:1, 10:3}
BG_RR_W = {1:3, 6:1, 7:2, 8:3, 10:1}
BG_AUX  = {"z9_sr":2, "z9_sr_buddy":1, "admin":2, "trash_1_5":2, "trash_6_10":1,
           "support_1":1, "support_2":3, "support_3":2}

# DEPRECATED — Phase 4k.1: these hardcoded dicts are no longer the source of
# truth. TASKS_ZONE and TASKS_RR are rebuilt at render time from the zone_tasks
# DB table via _load_tasks_from_db(). Kept as emergency fallback only.
TASKS_ZONE: dict[int, list[str]] = {
    1:  ["Outdoor Smoking Area", "Elevators & Stairwells", "Family Restroom"],
    2:  ["Lobby Trash Pull", "Lobby Restrooms"],
    3:  [],
    4:  ["Poker Room Drink Trays"],
    5:  ["High Limit Table Games", "Indoor TM Smoking Room"],
    6:  ["Outdoor Smoking Area"],
    7:  ["Smoking Room", "Pit 1 & 2", "South Door Glass"],
    8:  ["Restrooms", "Pit 3"],
    9:  ["Assist with Smoking Room", "Social Bar Tables"],
    10: ["High Limit Slots", "East Door Glass", "Outdoor Smoking Area", "Pit 4"],
}
TASKS_RR: dict[int, list[str]] = {
    1:  ["Buffet RR", "Family RR"],
    6:  ["131 Restroom"],
    7:  ["Assist with Smoking Room"],
    8:  ["Family Restroom", "TDR Restroom", "TMBR Locker Room"],
    10: ["CBK Kitchen"],
}
# TASKS_AUX retains its tuple (label, description) structure — aux slots use
# a compound display format the zone_tasks table does not yet model natively.
TASKS_AUX: dict[str, tuple[str, str]] = {
    "trash_1_5":    ("Trash 1",        "Zones 1–5, plus Annex after 5am"),
    "trash_6_10":   ("Trash 2",        "Zones 6–10"),
    "admin":        ("Admin",          ""),
    "z9_sr":        ("Z9 SR",          "Smoking Room"),
    "support_1":    ("Support 1",      ""),
    "support_2":    ("Support 2",      ""),
    "support_3":    ("Support 3",      "Overflow"),
    "z9_sr_buddy":  ("Z9 SR Buddy",    "Smoking Room (paired)"),
}

# Overlap task assignments — loaded from Supabase (overlap_tasks table).
# Defaults stay as a safety net if the DB is unavailable, but in normal
# operation the canonical source is the engine's overlap_tasks table.
_OVERLAP_DEFAULT_PM = [
    "Vacuum, Bottles & Glass",
    "Glass & Counters, Trash",
    "Tables & Restroom, Bottles & Glass",
    "Trash", "Trash", "Trash",
]
_OVERLAP_DEFAULT_AM = [
    "CBK / Shkodé",
    "CBK / Shkodé Restrooms",
    "Hotel Offices",
    "Sandhill / Lobby Bar",
    "131 / Group Room / CBK Office",
    "Trash",
]

def _load_tasks_from_db() -> None:
    """Phase 4k.1 — Rebuild module-level TASKS_ZONE and TASKS_RR from the
    zone_tasks table via list_tasks(). Tasks are sorted by display_order ASC
    then name ASC (handled by list_tasks). Hardcoded fallbacks remain when the
    DB is unavailable.

    TASKS_AUX is intentionally left as hardcoded — its (label, description)
    tuple structure does not map directly to zone_tasks rows.
    """
    try:
        rows = list_tasks(active_only=True, include_overlap=False)
    except Exception:
        return  # silently keep hardcoded defaults

    # Partition rows by category and default_zone slot key
    zone_buckets: dict[int, list[str]] = {}
    rr_buckets: dict[int, list[str]] = {}

    for row in rows:
        dz = row.get("default_zone", "")
        name = row.get("name", "")
        cat = row.get("category", "")
        if cat == "zone" and dz.startswith("zone_"):
            try:
                z = int(dz.split("_", 1)[1])
                zone_buckets.setdefault(z, []).append(name)
            except (ValueError, IndexError):
                pass
        elif cat == "rr" and dz.startswith("rr_"):
            try:
                z = int(dz.split("_", 1)[1])
                rr_buckets.setdefault(z, []).append(name)
            except (ValueError, IndexError):
                pass
        # aux tasks: TASKS_AUX stays hardcoded; skip here

    # Write back to module-level dicts (fallback kept for empty buckets)
    for z in range(1, 11):
        if z in zone_buckets:
            TASKS_ZONE[z] = zone_buckets[z]
    for z in (1, 6, 7, 8, 10):
        if z in rr_buckets:
            TASKS_RR[z] = rr_buckets[z]


def load_overlap_tasks() -> tuple[list[str], list[str], dict]:
    """Read PM/AM canonical overlap tasks from Supabase via
    get_overlap_tasks_for_engine(). Per-day overrides live in
    overlap_task_overrides and are picked up at render time per-date
    (not loaded once globally as the old _per_day_overrides JSON did).

    Returns (pm_list_of_6, am_list_of_6, empty_overrides_dict).
    The third tuple element is kept for API compatibility; per-day
    overrides are now resolved on-the-fly via tasks_for_date() below.
    """
    try:
        cfg = get_overlap_tasks_for_engine()  # {"PM": {slot:task}, "AM": {slot:task}}
        pm_map = cfg.get("PM", {}) or {}
        am_map = cfg.get("AM", {}) or {}
        pm = [pm_map.get(f"PMOL{i}", _OVERLAP_DEFAULT_PM[i-1]) for i in range(1, 7)]
        am = [am_map.get(f"AMOL{i}", _OVERLAP_DEFAULT_AM[i-1]) for i in range(1, 7)]
        return pm, am, {}
    except Exception as e:
        print(f"  [warn] overlap_tasks DB read failed ({e}) — using defaults")
        return list(_OVERLAP_DEFAULT_PM), list(_OVERLAP_DEFAULT_AM), {}

TASKS_PM_OL: list[str] = list(_OVERLAP_DEFAULT_PM)
TASKS_AM_OL: list[str] = list(_OVERLAP_DEFAULT_AM)
OVERLAP_OVERRIDES: dict = {}

# Per-day sweeper task overrides. Format: { "ISO date": { "sweeper_5_8_hl": "rr_8" } }
# Brian (5/1/26): Chris does Sweeper 5/8/HL on Friday — pin it to MRR 6 (his slot)
SWEEPER_REASSIGNMENTS: dict = {
    "2026-05-01": {"sweeper_5_8_hl": "rr_6"},
}

# Per-day manual alerts — orange banner overlays on specific cards.
# Used for one-off events that don't come from auto-coverage logic:
#   - Training milestones (e.g. "ADP Training" on a trainee's MP card)
#   - Special projects ("Deep Clean Annex 5am-7am")
#   - Coverage callouts the supervisor wants to highlight
# Slot keys: zone_1..zone_10, rr_1/6/7/8/10, aux_admin, aux_z9_sr,
# aux_z9_sr_buddy, aux_trash_1_5, aux_trash_6_10, aux_support_1,
# aux_support_2, aux_support_3.
MANUAL_ALERTS: dict = {
    "2026-04-28": { "aux_support_1": "ADP Training" },   # Seth's last training day
}

# Indices into the rr_mens / rr_womens arrays (5 cards: [1+2, 6, 7, 8, 10])
_RR_IDX = {1: 0, 6: 1, 7: 2, 8: 3, 10: 4}

def compute_alerts(day: dict) -> dict:
    """Return a dict mapping position-key → orange alert banner text.
    Auto-derives coverage callouts from staffing.

    Coverage relationships (Brian's spec):
      Zone 6/7/8/10 → covered by same-numbered RR
      Zone 2 → covered by Zone 4
      Zone 3 → covered by Zone 5
      Zone 9 → covered by Zone 10 (only when Z9 SR is staffed)
                or by Z9 SR (when both Z9 and Z10 are empty)
      Z9 SR  → covered by Zone 9 (when SR is empty + Z9 staffed)
      RR 1+2 short → Zones 1 & 2 cover
    """
    alerts = {}
    zones = day["zones"]
    rr_m, rr_w = day["rr_mens"], day["rr_womens"]
    rr_empty = lambda i: not rr_m[i] and not rr_w[i]
    z9_sr_filled = bool(day.get("aux", {}).get("z9_sr"))

    # --- Zone 1+2 ↔ Restroom 1+2 ---
    z1_empty, z2_empty = not zones[0], not zones[1]
    rr1_idx = _RR_IDX[1]
    rr1_short = not rr_m[rr1_idx] or not rr_w[rr1_idx]
    rr1_has_anyone = rr_m[rr1_idx] or rr_w[rr1_idx]

    if rr1_has_anyone:
        if z1_empty and z2_empty:
            alerts["rr_1"] = "And Zones 1 + 2"
        elif z1_empty:
            alerts["rr_1"] = "And Zone 1"
        elif z2_empty:
            alerts["rr_1"] = "And Zone 2"

    if rr1_short:
        if not z1_empty: alerts["zone_1"] = "Zone 1 + 2 Restrooms"
        if not z2_empty: alerts["zone_2"] = "Zone 1 + 2 Restrooms"

    # --- Zones 6, 7, 8, 10 ↔ same-numbered Restroom ---
    for n in (6, 7, 8, 10):
        zi, ri = n - 1, _RR_IDX[n]
        if not zones[zi] and (rr_m[ri] or rr_w[ri]):
            alerts[f"rr_{n}"] = f"And Zone {n}"
        if rr_empty(ri) and zones[zi]:
            alerts[f"zone_{n}"] = f"And Restroom {n}"

    # --- Zone 2 covered by Zone 1 + Zone 4 (when RR 1+2 isn't already covering) ---
    # Brian (5/1/26): mirror the Z9 chain — both adjacent staffed cards
    # carry the "And Zone 2" callout independently so the floor sees coverage
    # from either direction.
    if z2_empty and "rr_1" not in alerts:
        if zones[0]:
            alerts["zone_1"] = "And Zone 2"
        if zones[3]:
            alerts["zone_4"] = "And Zone 2"

    # --- Zone 3 covered by Zone 5 ---
    z3_empty = not zones[2]
    if z3_empty and zones[4]:
        alerts["zone_5"] = "And Zone 3"

    # --- Z9 area chain ---
    z9_empty  = not zones[8]
    z10_empty = not zones[9]
    # Z9 staffed but Z9 SR empty → Z9 covers SR
    if not z9_sr_filled and not z9_empty:
        alerts["zone_9"] = "And Z9 Smoking Room"
    # Z9 empty → both Z9 SR and Z10 (if staffed) carry "And Zone 9".
    # Per Brian (4/29/26): each card carries the "Zone 9" callout
    # independently — the banner stays singular ("And Zone 9"), never
    # "+ Z10". Z10 coverage stays in its own chain via RR 10.
    if z9_empty:
        if z9_sr_filled:
            alerts["aux_z9_sr"] = "And Zone 9"
        if not z10_empty:
            alerts["zone_10"] = "And Zone 9"

    return alerts

# --------------------------------------------------------------------------
# v2.3 redesign helpers — icon mapping, cover-tag, break-group builder
# --------------------------------------------------------------------------
# The v6 full-width crimson `alert-banner` is replaced in the v2.3 design by
# inline `cover-tag` pills next to the relevant card label. compute_alerts()
# still produces the source text — we just transform and present it differently.

# Task-string → SVG glyph ID. First substring match wins.
ICON_RULES = [
    ("smoking",     "g-smoke"),
    ("smoke",       "g-smoke"),
    ("elevator",    "g-elevator"),
    ("stairwell",   "g-elevator"),
    ("131",         "g-toilet"),
    ("buffet",      "g-toilet"),
    ("family",      "g-toilet"),
    ("tdr",         "g-toilet"),
    ("tmbr",        "g-key"),
    ("locker",      "g-key"),
    ("restroom",    "g-toilet"),
    (" rr",         "g-toilet"),
    ("trash",       "g-trash"),
    ("vacuum",      "g-broom"),
    ("sweeper",     "g-broom"),
    ("sweep",       "g-broom"),
    ("broom",       "g-broom"),
    ("bottle",      "g-glass"),
    ("glass",       "g-glass"),
    ("counter",     "g-glass"),
    ("table",       "g-table"),
    ("drink",       "g-table"),
    ("cbk",         "g-table"),
    ("shkode",      "g-table"),
    ("shkodé",      "g-table"),
    ("kitchen",     "g-table"),
    ("hl ",         "g-pit"),
    ("high limit",  "g-pit"),
    ("pit",         "g-pit"),
]

def icon_for_task(task: str) -> str:
    if not task: return ""
    low = " " + task.lower()
    for needle, gid in ICON_RULES:
        if needle in low:
            return f'<svg><use href="#{gid}"/></svg>'
    return ""

def _render_task_li(task: str) -> str:
    """Render a single task list item. Sweeper tasks get a distinct
    'SWEEPER' pill + faint orange tint so they stand out against the regular
    task lines — sweeper duty is operationally distinct (it's a route, not a
    static task) and Brian wanted it more noticeable on the floor."""
    if not task: return "<li></li>"
    if task.lower().startswith("sweeper"):
        # Case-safe split: works for "Sweeper xyz" or "sweeper xyz" or just "Sweeper"
        rest = task.split(" ", 1)[1] if " " in task else ""
        return (f'<li class="sweeper-task">'
                f'<span class="sweeper-pill">SWEEPER</span>'
                f'<span class="sweeper-route">{esc(rest)}</span>'
                f'</li>')
    return f'<li>{icon_for_task(task)}{esc(task)}</li>'

def alert_to_short(text: str) -> str:
    """Compress compute_alerts() text into a short cover-tag label.

    'And Zone 6'              -> 'covers Z6'
    'And Restroom 6'          -> 'covers RR 6'
    'And Z9 Smoking Room'     -> 'covers Z9 SR'
    'Zone 1 + 2 Restrooms'    -> 'covers RR 1+2'
    'And Zones 9 + 10'        -> 'covers Z9 + 10'
    """
    if not text: return ""
    t = text.strip()
    if t.lower().startswith("and "):
        t = t[4:]
    repl = [
        ("Z9 Smoking Room", "Z9 SR"),
        ("Smoking Room",    "SR"),
        ("Restrooms",       "RR"),
        ("Restroom",        "RR"),
        ("Zones",           "Z"),
        ("Zone",            "Z"),
        (" 1 + 2 RR",       " RR 1+2"),
    ]
    for a, b in repl:
        t = t.replace(a, b)
    t = " ".join(t.split())
    # Tighten "Z 2" → "Z2" (RR keeps the space, e.g. "covers RR 6")
    t = re.sub(r'\bZ\s+(\d)', r'Z\1', t)
    return f"covers {t}"

def cover_tag(text: str, glyph: str = "g-link") -> str:
    """Inline pill: SVG glyph + short text. Returns empty string if no text.
    Retained for backward compat; v2.4 uses alert_strip() for prominent
    coverage callouts instead."""
    if not text: return ""
    return (f'<span class="cover-tag">'
            f'<svg><use href="#{glyph}"/></svg>{esc(text)}</span>')

def alert_banner_text(text: str) -> str:
    """Format compute_alerts() text as the bold uppercase banner string.
    'And Zone 2' -> 'AND ZONE 2'; 'And Z9 Smoking Room' -> 'AND Z9 SMOKING ROOM'."""
    if not text: return ""
    return text.strip().upper()

def alert_strip(text: str, glyph: str = "g-alert") -> str:
    """Bottom-of-card crimson banner. Reads from the floor."""
    if not text: return ""
    return (f'<div class="alert-strip">'
            f'<svg><use href="#{glyph}"/></svg>'
            f'<span>{esc(alert_banner_text(text))}</span></div>')

def count_summary(day: dict) -> dict:
    zones    = sum(1 for n in day["zones"] if n)
    rr_sides = sum(1 for n in day["rr_mens"] if n) + sum(1 for n in day["rr_womens"] if n)
    # Aux base = 6 always-present slots (Z9 SR + Admin + Trash 1/2 + Support 1/2).
    # Support 3 + Z9 SR Buddy are overflow indicators — included only when populated.
    base_aux = ("trash_1_5", "trash_6_10", "admin", "z9_sr", "support_1", "support_2")
    aux_filled = sum(1 for k in base_aux if day["aux"].get(k))
    aux_total  = len(base_aux)
    if day["aux"].get("support_3"):
        aux_filled += 1; aux_total += 1
    if day["aux"].get("z9_sr_buddy"):
        aux_filled += 1; aux_total += 1
    overlaps = sum(1 for n in day["pm_ol"] if n) + sum(1 for n in day["am_ol"] if n)
    return {
        "zones":    (zones,    10),
        "rr":       (rr_sides, 10),
        "aux":      (aux_filled, aux_total),
        "overlaps": (overlaps, 12),
    }

def count_break_groups(day: dict) -> tuple:
    """(g1_count, g2_count, g3_count) of staffed slots per break wave."""
    g = [0, 0, 0]
    rr_nums = [1, 6, 7, 8, 10]
    for i, name in enumerate(day["zones"]):
        if name and BG_ZONE.get(i + 1) in (1, 2, 3):
            g[BG_ZONE[i + 1] - 1] += 1
    for i, name in enumerate(day["rr_mens"]):
        if name and BG_RR_M.get(rr_nums[i]) in (1, 2, 3):
            g[BG_RR_M[rr_nums[i]] - 1] += 1
    for i, name in enumerate(day["rr_womens"]):
        if name and BG_RR_W.get(rr_nums[i]) in (1, 2, 3):
            g[BG_RR_W[rr_nums[i]] - 1] += 1
    for k, name in day["aux"].items():
        if name and BG_AUX.get(k) in (1, 2, 3):
            g[BG_AUX[k] - 1] += 1
    return tuple(g)

def total_in_rotation(day: dict) -> int:
    return (sum(1 for n in day["zones"] if n)
            + sum(1 for n in day["rr_mens"] if n)
            + sum(1 for n in day["rr_womens"] if n)
            + sum(1 for n in day["aux"].values() if n))

# Aggressive abbreviations applied only on the break sheet rows (where space
# is tight and the daily page already shows full task names).
ABBREVS = [
    ("Outdoor Smoking Area",     "Outdoor Smoking"),
    ("Elevators & Stairwells",   "Elevators"),
    ("Family Restroom",          "Family RR"),
    ("Buffet RR",                "Buffet"),
    ("Family RR",                "Family"),
    ("Lobby Trash Pull",         "Lobby Trash"),
    ("Lobby Restrooms",          "Lobby RR"),
    ("Poker Room Drink Trays",   "Poker Drink Trays"),
    ("High Limit Table Games",   "HL Tables"),
    ("Indoor TM Smoking Room",   "TM Smoking"),
    ("Smoking Room",             "Smoking Rm"),
    ("Pit 1 & 2",                "Pit 1&2"),
    ("South Door Glass",         "S Door Glass"),
    ("East Door Glass",          "E Door Glass"),
    ("Restrooms",                "RR"),
    ("Assist with Smoking Room", "Smoking assist"),
    ("High Limit Slots",         "HL Slots"),
    ("131 Restroom",             "131 RR"),
    ("TDR Restroom",             "TDR"),
    ("TMBR Locker Room",         "TMBR"),
]

def abbrev_task(t: str) -> str:
    for a, b in ABBREVS:
        t = t.replace(a, b)
    return t

def join_assigns(tasks) -> str:
    if not tasks: return ""
    return " · ".join(abbrev_task(t) for t in tasks if t)

# --------------------------------------------------------------------------
# Break-sheet row builder — buckets every staffed slot into Group 1/2/3
# --------------------------------------------------------------------------

_AUX_ORDER = ["z9_sr", "admin", "trash_1_5", "trash_6_10",
              "support_1", "support_2", "support_3"]
_AUX_COLOR = {"z9_sr": "red", "z9_sr_buddy": "red", "admin": "purple",  # Admin is purple per Brian
              "trash_1_5": "orange", "trash_6_10": "orange",
              "support_1": "grey", "support_2": "grey", "support_3": "teal"}

def build_break_groups(day: dict, sweeper_add: dict) -> dict:
    """Return {1: [row, ...], 2: [...], 3: [...]} where each row is a dict:
       {section, name, badge, color, assign}.
    Section ordering within a group: Zones → Auxiliary → Restrooms.
    Within section: zones by number, aux by canonical order, RR by number/men/women."""
    groups = {1: [], 2: [], 3: []}

    def assign_for_zone(z):
        base = list(TASKS_ZONE[z])
        for t in sweeper_add.get(f"zone_{z}", []):
            if t not in base: base.append(t)
        return join_assigns(base)

    def assign_for_aux(k):
        _, sub = TASKS_AUX[k]
        items = []
        if sub: items.append(sub)
        items.extend(sweeper_add.get(f"aux_{k}", []))
        return join_assigns(items)

    def assign_for_rr_side(rr_num, side_label):
        items = [side_label]
        items.extend(TASKS_RR.get(rr_num, []))
        items.extend(sweeper_add.get(f"rr_{rr_num}", []))
        return join_assigns(items)

    # Zones
    for i, name in enumerate(day["zones"]):
        if not name: continue
        g = BG_ZONE.get(i + 1)
        if g not in (1, 2, 3): continue
        groups[g].append({
            "section": "Zones",
            "name":    name,
            "badge":   f"Zone {i + 1}",
            "color":   ZONE_COLOR[i + 1],
            "assign":  assign_for_zone(i + 1),
        })

    # Auxiliary
    for k in _AUX_ORDER:
        name = day["aux"].get(k)
        if not name: continue
        g = BG_AUX.get(k)
        if g not in (1, 2, 3): continue
        label, _sub = TASKS_AUX[k]
        groups[g].append({
            "section": "Auxiliary",
            "name":    name,
            "badge":   label,
            "color":   _AUX_COLOR[k],
            "assign":  assign_for_aux(k),
        })
    # Z9 SR Buddy (F/Sa overflow paired in the smoke room) — added separately so it
    # doesn't pollute _AUX_ORDER for the daily-page strip layout.
    buddy = day["aux"].get("z9_sr_buddy")
    if buddy:
        g = BG_AUX.get("z9_sr_buddy", 2)
        if g in (1, 2, 3):
            label, sub = TASKS_AUX["z9_sr_buddy"]
            groups[g].append({
                "section": "Auxiliary",
                "name":    buddy,
                "badge":   label,
                "color":   _AUX_COLOR.get("z9_sr_buddy", "red"),
                "assign":  sub or "",
            })

    # Restrooms — RR# ascending, Men's then Women's
    rr_nums = [1, 6, 7, 8, 10]
    for i, n in enumerate(rr_nums):
        nm = day["rr_mens"][i]
        if nm:
            g = BG_RR_M.get(n)
            if g in (1, 2, 3):
                groups[g].append({
                    "section": "Restrooms",
                    "name":    nm,
                    "badge":   f"RR {n}" if n != 1 else "RR 1+2",
                    "color":   RR_COLOR[n],
                    "assign":  assign_for_rr_side(n, "Men's"),
                })
        nw = day["rr_womens"][i]
        if nw:
            g = BG_RR_W.get(n)
            if g in (1, 2, 3):
                groups[g].append({
                    "section": "Restrooms",
                    "name":    nw,
                    "badge":   f"RR {n}" if n != 1 else "RR 1+2",
                    "color":   RR_COLOR[n],
                    "assign":  assign_for_rr_side(n, "Women's"),
                })

    return groups

def render_break_col(group_num: int, rows: list) -> str:
    n = len(rows)
    parts = []
    cur_section = None
    for row in rows:
        if row["section"] != cur_section:
            parts.append(f'<div class="break-section-divider">{esc(row["section"])}</div>')
            cur_section = row["section"]
        assign_html = (f'<div class="assign">{esc(row["assign"])}</div>'
                       if row["assign"] else "")
        parts.append(
            f'<div class="break-row" style="--row-color: var(--c-{row["color"]});">'
            f'<div class="who-line"><div class="name">{esc(row["name"])}</div>{assign_html}</div>'
            f'<div class="slot-badge">{esc(row["badge"])}</div>'
            f'</div>'
        )
    rows_html = "".join(parts)
    return f"""<div class="break-col" data-group="{group_num}">
        <div class="col-head">
          <div class="col-num">{group_num}</div>
          <div class="col-info">
            <div class="col-label">Break {group_num}</div>
            <div class="col-meta"><span class="num">{n}</span> people</div>
          </div>
        </div>
        <div class="break-rows">{rows_html}</div>
      </div>"""

# Utility Porters per-day data — read from Rules/Utility Porters.json.
# Schema:
#   {
#     "2026-04-28": [
#       {"role": "PM",  "name": "Alistair"},
#       {"role": "PM",  "name": "Doug"},
#       ...up to 8 entries...
#     ],
#     "2026-04-29": [...],
#     ...
#   }
# Brian's intent (4/29/26): "It will come from the schedules uploaded from now on."
# Until the weekly schedule format includes Utility Porter rows, this JSON
# is manually populated (or left empty). When the schedule format adds
# Utility Porter info, fill_engine.py can write into this file directly.
UTILITY_PORTERS_BY_DATE: dict = {}

def load_utility_porters() -> dict:
    """Per-day Utility Porter assignments — currently empty.

    Brian (4/29/26): "It will come from the schedules uploaded from now on."
    Until the weekly schedule format includes Utility Porter rows, no source
    of truth exists for this data. Previously the JSON file was manually
    populated; that workflow was abandoned. Returns {} unconditionally so
    the renderer's per-day attach loop just sees no entries.

    When the schedule format adds Utility Porter info, fill_engine.py
    can populate a `utility_porters` table (one row per (date, role, tm_id))
    and this function can be wired to read it.
    """
    return {}

# Sweeper routes — keyed by sweeper code; values are (display label, route zones)
SWEEPER_ROUTES = {
    "sweeper_5_8_hl":  ("Sweeper 5 / 8 / HL",  [5, 8]),
    "sweeper_9_10_sr": ("Sweeper 9 / 10 / SR", [9, 10]),
}

# --------------------------------------------------------------------------
# Workbook reader — pulls 7 days of names from Week Overview
# --------------------------------------------------------------------------

def read_week(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Week Overview"]

    def name(r, c):
        v = ws.cell(row=r, column=c).value
        return "" if v is None else str(v).strip()

    start_cell = ws.cell(row=3, column=1).value
    if isinstance(start_cell, dt.datetime):
        start = start_cell.date()
    elif isinstance(start_cell, dt.date):
        start = start_cell
    else:
        stem = xlsx_path.stem
        try:
            we = dt.date.fromisoformat(stem.split(" - ")[-1])
            start = we - dt.timedelta(days=6)
        except Exception:
            start = dt.date.today()

    days = []
    for i in range(7):
        zone_row = 3 + i
        rr_row   = 13 + i
        aux_row  = 23 + i
        d = start + dt.timedelta(days=i)
        days.append({
            "date":       d,
            "label":      d.strftime("%A, %B %-d, %Y"),
            "weekday":    d.strftime("%A"),
            "date_short": d.strftime("%B %-d, %Y"),
            "day_num":    d.day,
            "zones":      [name(zone_row, 2 + z) for z in range(10)],
            "rr_mens":    [name(rr_row,   2 + z) for z in range(5)],
            "rr_womens":  [name(rr_row,   7 + z) for z in range(5)],
            "pm_ol":      [name(zone_row, 14 + z) for z in range(6)],
            "am_ol":      [name(rr_row,   14 + z) for z in range(6)],
            "aux": {
                "trash_1_5":   name(aux_row, 2),
                "trash_6_10":  name(aux_row, 3),
                "admin":       name(aux_row, 4),
                "z9_sr":       name(aux_row, 5),
                "support_1":   name(aux_row, 6),
                "support_2":   name(aux_row, 7),
                "support_3":   name(aux_row, 8),
                "z9_sr_buddy": name(aux_row, 9),
            },
        })
    return days

# --------------------------------------------------------------------------
# Gender info + sweeper auto-assignment
# --------------------------------------------------------------------------

def load_training_config() -> dict:
    """Read training schedule from Supabase via get_training_schedule_from_db().
    Returns {iso_date: {trainee, trainer, day}} so the renderer can pair the
    trainee onto the trainer's zone card with a 'TRAINING D{n}' pill."""
    try:
        cfg = get_training_schedule_from_db()  # {"schedule": {...}}
        return cfg.get("schedule", {}) or {}
    except Exception as e:
        print(f"  [warn] training_schedule DB read failed ({e}) — skipping training pairs")
        return {}

def load_gender_info() -> tuple[set, set, set]:
    """(males, females, no_sweeper) derived from Supabase tm_profiles +
    tm_eligibility (formerly read from Eligibility Roster.xlsx + TM
    Profiles.json under Rules/).

    Gender derivation: `Mens X` columns true → male; `Womens X` true → female.
    Mirrors the old Eligibility Roster-derived logic so the deployment book
    renders identically to the file-based version.
    """
    males, females, no_sweeper = set(), set(), set()
    try:
        roster, _ = get_engine_roster_from_db()
        for _key, info in (roster or {}).items():
            dn = (info.get("display_name") or "").strip()
            if not dn:
                continue
            elig = info.get("eligibility") or {}
            is_female = any(elig.get(f"Womens {n}", False) for n in ("1 + 2", 6, 7, 8, 10))
            is_male   = any(elig.get(f"Mens {n}",   False) for n in ("1 + 2", 6, 7, 8, 10))
            if is_female:
                females.add(dn)
            elif is_male:
                males.add(dn)
    except Exception as e:
        print(f"  [warn] roster DB read failed ({e}) — gender info empty")

    try:
        profiles = get_engine_profiles_from_db().get("profiles", {}) or {}
        for dn, p in profiles.items():
            if isinstance(p, dict) and p.get("slot_preference") == "no_sweeper":
                no_sweeper.add(dn)
    except Exception as e:
        print(f"  [warn] profiles DB read failed ({e}) — no_sweeper info empty")

    return males, females, no_sweeper


def assign_sweepers(day: dict, males: set, no_sweeper: set) -> dict:
    """Auto-assign each sweeper route to an eligible male per Brian's priority:
      1. Trash positions (if male and eligible)
      2. Male in a zone the sweeper covers (Z5/Z8 for 5/8/HL; Z9/Z10 for 9/10/SR)
      3. Any other eligible male anywhere except Support, Admin, Z9 SR (already taken)
    Returns: {sweeper_key: (target_slot_key, label, name)}."""
    def is_eligible(name):
        return bool(name) and name in males and name not in no_sweeper

    eligible = {}
    for i, name in enumerate(day["zones"]):
        if is_eligible(name):
            eligible[f"zone_{i+1}"] = (f"Zone {i+1}", name)
    rr_nums = [1, 6, 7, 8, 10]
    for i, name in enumerate(day["rr_mens"]):
        if is_eligible(name):
            label = f"RR {rr_nums[i]}" if rr_nums[i] != 1 else "RR 1+2"
            eligible[f"rr_{rr_nums[i]}"] = (label, name)
    for k, label in (("trash_1_5", "Trash 1"), ("trash_6_10", "Trash 2")):
        name = day["aux"].get(k)
        if is_eligible(name):
            eligible[f"aux_{k}"] = (label, name)
    z9sr = day["aux"].get("z9_sr")
    if is_eligible(z9sr):
        eligible["aux_z9_sr"] = ("Z9 SR", z9sr)

    assignments = {}
    used = set()
    for sweeper_key, (label, route_zones) in SWEEPER_ROUTES.items():
        target = None
        for tk in ("aux_trash_1_5", "aux_trash_6_10"):
            if tk in eligible and tk not in used:
                target = tk; break
        if not target:
            for z in route_zones:
                k = f"zone_{z}"
                if k in eligible and k not in used:
                    target = k; break
        if not target:
            order = ([f"zone_{i}" for i in range(1, 11)]
                     + [f"rr_{n}" for n in (1, 6, 7, 8, 10)]
                     + ["aux_z9_sr"])
            for k in order:
                if k in eligible and k not in used:
                    target = k; break
        if target:
            assignments[sweeper_key] = (target, label, eligible[target][1])
            used.add(target)
    return assignments

# --------------------------------------------------------------------------
# HTML helpers
# --------------------------------------------------------------------------

def esc(s: str) -> str:
    if s is None: return ""
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

def name_or_unfilled(s: str, classes: str = "name") -> str:
    if s:
        return f'<span class="{classes}">{esc(s)}</span>'
    return f'<span class="{classes} is-unfilled">Unfilled</span>'

def render_zone_card(num, name_str, color, tasks, alert="", group=None,
                     trainee_name="", training_day=None, trainer_marker_only=False):
    """v2.4 zone card. 3px top stripe, meta row (zone-num + group-pill),
    big name, icon-prefixed task list, optional bottom-strip alert banner.
    Empty zones with no canonical tasks render as italic 'Support as needed'.

    Training mode (5/1/26):
      - Standard mode (D1/D2/D3/D6): trainee_name set + trainer_marker_only=False
        → renders TRAINING pill + trainee name beside trainer (overlay fiction
        for days when trainee isn't in any board slot)
      - Trainer-marker-only mode (D4/D5): trainer_marker_only=True
        → renders TRAINING pill alone, no trainee name (trainee is visible in
        their actual RR/zone slot — the marker just flags Tawnya as the trainer)
    """
    name_html = (f'<div class="zone-name">{esc(name_str)}</div>'
                 if name_str else
                 '<div class="zone-name is-unfilled">Unfilled</div>')
    pill = (f'<span class="group-pill" data-group="{group}">{group}</span>'
            if (group and name_str) else "")
    trainee_html = ""
    if trainee_name and not trainer_marker_only:
        # Day marker removed from the displayed pill per Brian (5/1/26) — the pill
        # just reads 'TRAINING'. The training_day value is still tracked in Training
        # Config for engine placement logic (Day 1=RR, Day 2=Zone, etc.) but isn't
        # surfaced on the printed sheet.
        trainee_html = (f'<div class="zone-trainee">'
                        f'<span class="train-pill">TRAINING</span>'
                        f'<span class="trainee-name">{esc(trainee_name)}</span>'
                        f'</div>')
    elif trainer_marker_only:
        # Pill only — trainee is rendered elsewhere, this just flags the trainer.
        trainee_html = ('<div class="zone-trainee zone-trainee--marker-only">'
                        '<span class="train-pill">TRAINING</span>'
                        '</div>')
    if tasks:
        items = "".join(_render_task_li(t) for t in tasks)
        tasks_html = f'<ul class="zone-tasks">{items}</ul>'
    else:
        tasks_html = ('<ul class="zone-tasks zone-tasks--asneeded">'
                      '<li><em>Support as needed</em></li></ul>')
    empty_cls = "" if name_str else " is-empty"
    alert_cls = " has-alert" if alert else ""
    train_cls = " has-trainee" if trainee_html else ""  # fires in both modes
    # Crowded modifier: 4+ tasks triggers tighter layout so all tasks stay visible
    # (Z10 with HL Slots / E Door Glass / Outdoor Smoking / Pit 4 was clipping).
    # is-extra-crowded fires at 5+ tasks (typically when sweeper task is added),
    # shrinking the trainer name as well to give the task block more room.
    n_tasks = len(tasks) if tasks else 0
    crowded_cls = ""
    if n_tasks >= 5:
        crowded_cls = " is-crowded is-extra-crowded"
    elif n_tasks >= 4:
        crowded_cls = " is-crowded"
    alert_html = alert_strip(alert)
    return f"""<div class="zone-card c-{color}{empty_cls}{alert_cls}{train_cls}{crowded_cls}">
        <div class="zone-meta"><span class="zone-num"><svg class="zone-shape" aria-hidden="true"><use href="#sh-{num}"/></svg>Zone {num}</span>{pill}</div>
        {name_html}
        {trainee_html}
        {tasks_html}
        {alert_html}
      </div>"""

def render_aux_card(key, name_str, color, extra_tasks=None, alert="", group=None,
                    buddy_name="", buddy_group=None, conditional=False):
    """v2.4 aux card. Now supports:
      - buddy_name: equal-partner pairing rendered side-by-side at the same
        baseline; BUDDIED pill sits in the meta header (Z9 SR F/Sa)
      - alert: bottom-strip crimson banner
      - conditional: card carries 'is-conditional' class so empty Support 3 hides
    """
    label, sub = TASKS_AUX[key]
    # Phase 4g polish — semantic dedup. The default `sub` for trash_1_5 is
    # "Zones 1–5, plus Annex after 5am" which is functionally the same as the
    # supervisor-set custom_tasks ["Zones 1-5", "Annex after 5am"]. Rendering
    # both produced redundant 3-line stacks. Strategy: if any extra-task is a
    # substring of `sub` (after normalizing punctuation/whitespace), drop the
    # base `sub` and let the extras tell the full story line-by-line.
    items = []
    if extra_tasks:
        items.extend(extra_tasks)
    if sub:
        def _norm(s: str) -> str:
            return (s or "").replace("–", "-").replace(",", " ").lower().split()
        sub_words = set(_norm(sub))
        # If extras collectively cover most of sub's words, sub is redundant.
        extras_words = set()
        for e in (extra_tasks or []):
            extras_words.update(_norm(e))
        sub_is_redundant = (
            bool(extras_words) and
            len(sub_words & extras_words) >= max(1, int(len(sub_words) * 0.6))
        )
        if not sub_is_redundant:
            items.insert(0, sub)
    if items:
        tasks_html = ('<ul class="aux-tasks">'
                      + "".join(_render_task_li(t) for t in items)
                      + '</ul>')
    else:
        tasks_html = ""
    if buddy_name:
        # Side-by-side level rendering — both names equal weight, each with
        # their own group pill. Brian (5/1/26): BUDDIED pill removed from the
        # meta header — the two-name layout is self-explanatory and the meta
        # stays clean. The break group pill on each name now does the work.
        primary_pill = (f'<span class="group-pill" data-group="{group}">{group}</span>'
                        if (group and name_str) else "")
        buddy_pill   = (f'<span class="group-pill" data-group="{buddy_group}">{buddy_group}</span>'
                        if buddy_group else "")
        primary_render = (f'<div class="aux-name">{esc(name_str)}{primary_pill}</div>'
                          if name_str else
                          '<div class="aux-name is-unfilled">Unfilled</div>')
        buddy_render = (f'<div class="aux-name">{esc(buddy_name)}{buddy_pill}</div>')
        primary_html = (f'<div class="aux-buddy-row">{primary_render}{buddy_render}</div>')
        buddy_html = ""
        meta_pill = ""  # BUDDIED pill removed per Brian 5/1/26
    else:
        primary_html = (f'<div class="aux-name">{esc(name_str)}</div>'
                        if name_str else
                        '<div class="aux-name is-unfilled">Unfilled</div>')
        buddy_html = ""
        meta_pill = (f'<span class="group-pill" data-group="{group}">{group}</span>'
                     if (group and name_str) else "")
    empty_cls = "" if name_str else " is-empty"
    alert_cls = " has-alert" if alert else ""
    buddy_cls = " has-buddy" if buddy_name else ""
    cond_cls  = " is-conditional" if conditional else ""
    alert_html = alert_strip(alert)
    return f"""<div class="aux-card c-{color}{empty_cls}{alert_cls}{buddy_cls}{cond_cls}">
        <div class="aux-meta"><span class="aux-label">{esc(label)}</span>{meta_pill}</div>
        {primary_html}
        {buddy_html}
        {tasks_html}
        {alert_html}
      </div>"""

def render_rr_card(rr_num, mens, womens, color, extra_tasks=None, alert="",
                   mens_group=None, womens_group=None,
                   mens_training=False, womens_training=False):
    """v2.4 RR card — split men's/women's with per-side group pill, hairline-divided
    task list at bottom. Coverage callouts render as bottom alert strip.

    Training markers (5/1/26): when {mens,womens}_training is True, render a
    small TRAINING pill next to that side's name. Used on D4 (Shadow Restroom)
    so the trainee's RR slot carries the pairing indicator alongside the
    trainer-marker pill on the corresponding zone card."""
    label = f"RR {rr_num}" if rr_num != 1 else "RR 1 + 2"
    base_tasks = list(TASKS_RR.get(rr_num, []))
    if extra_tasks:
        for t in extra_tasks:
            if t not in base_tasks:
                base_tasks.append(t)
    if base_tasks:
        tasks_html = ('<ul class="rr-tasks">'
                      + "".join(_render_task_li(t) for t in base_tasks)
                      + '</ul>')
    else:
        tasks_html = ""

    def side(name_val, who_label, group, is_training):
        train_pill = ('<span class="train-pill rr-train-pill">TRAINING</span>'
                      if (is_training and name_val) else "")
        n = (f'<div class="name">{esc(name_val)}{train_pill}</div>' if name_val
             else '<div class="name is-unfilled">Unfilled</div>')
        pill = (f'<span class="group-pill" data-group="{group}">{group}</span>'
                if (group and name_val) else "")
        return (f'<div class="rr-side">'
                f'<span class="who">{who_label} {pill}</span>'
                f'{n}</div>')

    empty_cls = "" if (mens or womens) else " is-empty"
    alert_cls = " has-alert" if alert else ""
    alert_html = alert_strip(alert)
    return f"""<div class="rr-card c-{color}{empty_cls}{alert_cls}">
        <div class="rr-head"><span><svg class="zone-shape" aria-hidden="true"><use href="#sh-{rr_num}"/></svg>{label}</span></div>
        <div class="rr-split">
          {side(mens, "Men's", mens_group, mens_training)}
          {side(womens, "Women's", womens_group, womens_training)}
        </div>
        {tasks_html}
        {alert_html}
      </div>"""

def render_overlap_mini(name, task=""):
    """v2.3 overlap mini-card — name above, task with optional icon below."""
    who = (f'<div class="who">{esc(name)}</div>' if name
           else '<div class="who is-unfilled">Unfilled</div>')
    if task:
        what = f'<div class="what">{icon_for_task(task)}{esc(task)}</div>'
    else:
        what = '<div class="what">&nbsp;</div>'
    empty_cls = "" if name else " is-empty"
    return f'<div class="overlap-mini{empty_cls}">{who}{what}</div>'

# --------------------------------------------------------------------------
# Per-day page render
# --------------------------------------------------------------------------

def _compute_sweeper_add(day, males, no_sweeper_tms):
    """Resolve auto-sweeper + manual override into {target_slot: [labels...]}."""
    iso = day["date"].isoformat()
    auto_sweepers = assign_sweepers(day, males, no_sweeper_tms)
    sweeper_overrides = SWEEPER_REASSIGNMENTS.get(iso, {}) or {}
    sweeper_add = {}
    for sweeper_key, (label, _route) in SWEEPER_ROUTES.items():
        target_slot = sweeper_overrides.get(sweeper_key)
        if target_slot is None:
            auto = auto_sweepers.get(sweeper_key)
            if auto: target_slot = auto[0]
        if target_slot:
            sweeper_add.setdefault(target_slot, []).append(label)
    return sweeper_add

def render_day_page(day, idx, total, days, current_idx, males=None, no_sweeper_tms=None,
                    training_pair=None):
    """LANDSCAPE daily deployment page (Phase 6, 5/6/26).

    `idx` is 1-based day index (Friday=1 ... Thursday=7); `total` is total days
    in the schedule (always 7). Page numbering at the foot reflects the 14-page
    book: this page is `2*idx - 1` of `2*total`.

    Renders as 11in × 8.5in landscape with masthead, 4-row body grid (zones 2×5,
    restrooms, auxiliary, overlaps), and footer slug."""
    males = males or set()
    no_sweeper_tms = no_sweeper_tms or set()
    weekday   = day["weekday"]
    day_color = DAY_COLOR.get(weekday, "#444444")
    iso       = day["date"].isoformat()

    # Coverage alerts (auto + manual)
    alerts = compute_alerts(day)
    for slot_key, banner in (MANUAL_ALERTS.get(iso) or {}).items():
        alerts[slot_key] = banner

    sweeper_add = _compute_sweeper_add(day, males, no_sweeper_tms)

    def task_list_for(slot_key, base):
        extra = sweeper_add.get(slot_key, [])
        if not extra: return list(base)
        merged, seen = [], set()
        for t in list(base) + list(extra):
            if t not in seen:
                merged.append(t); seen.add(t)
        return merged

    # Counts
    cs = count_summary(day)
    g1, g2, g3 = count_break_groups(day)

    # Zones — training pair attaches the trainee onto the trainer's zone card.
    # Two visual modes (5/1/26):
    #   - Standard overlay (D1/D2/D3/D6): trainee not in any board slot →
    #     trainer's zone shows TRAINING pill + trainee name beside trainer
    #   - Trainer-marker-only (D4/D5): trainee placed in actual RR/zone slot →
    #     trainer's zone shows just the TRAINING pill (no name); the trainee's
    #     own slot card carries a TRAINING pill next to their name
    trainer_name = (training_pair or {}).get("trainer")
    trainee_name = (training_pair or {}).get("trainee")
    training_day = (training_pair or {}).get("day")
    trainee_placed_elsewhere = False
    trainee_rr_side = None  # 'mens' or 'womens' — index into rr arrays
    trainee_rr_idx  = None
    if trainee_name:
        if trainee_name in day["zones"]:
            trainee_placed_elsewhere = True
        elif trainee_name in day["rr_mens"]:
            trainee_placed_elsewhere = True
            trainee_rr_side = "mens"
            trainee_rr_idx = day["rr_mens"].index(trainee_name)
        elif trainee_name in day["rr_womens"]:
            trainee_placed_elsewhere = True
            trainee_rr_side = "womens"
            trainee_rr_idx = day["rr_womens"].index(trainee_name)
        elif trainee_name in day["aux"].values():
            trainee_placed_elsewhere = True
    zone_cards = []
    for z in range(10):
        zone_holder = day["zones"][z]
        is_trainer_zone = bool(trainer_name and zone_holder == trainer_name)
        if is_trainer_zone and trainee_placed_elsewhere:
            # D4/D5 mode: pill only on trainer's zone, no overlay name
            attach_trainee = ""
            marker_only   = True
        elif is_trainer_zone and trainee_name:
            # D1/D2/D3/D6 mode: pill + overlay name
            attach_trainee = trainee_name
            marker_only   = False
        else:
            attach_trainee = ""
            marker_only   = False
        zone_cards.append(
            render_zone_card(z + 1, zone_holder, ZONE_COLOR[z + 1],
                             task_list_for(f"zone_{z + 1}", TASKS_ZONE[z + 1]),
                             alert=alerts.get(f"zone_{z + 1}", ""),
                             group=BG_ZONE.get(z + 1),
                             trainee_name=attach_trainee,
                             training_day=training_day if attach_trainee else None,
                             trainer_marker_only=marker_only)
        )
    zones_html = "\n".join(zone_cards)

    # Auxiliary
    # F/Sa Z9 SR can carry a buddy (overflow #1 on weekend nights).
    z9sr_buddy = day["aux"].get("z9_sr_buddy", "")
    support_3  = day["aux"].get("support_3", "")
    aux_cards = [
        render_aux_card("z9_sr",      day["aux"]["z9_sr"],      "red",
                        extra_tasks=sweeper_add.get("aux_z9_sr", []),
                        alert=alerts.get("aux_z9_sr", ""),
                        group=BG_AUX.get("z9_sr"),
                        buddy_name=z9sr_buddy,
                        buddy_group=BG_AUX.get("z9_sr_buddy")),
        render_aux_card("admin",      day["aux"]["admin"],      "purple",  # Admin is purple per Brian
                        alert=alerts.get("aux_admin", ""),
                        group=BG_AUX.get("admin")),
        render_aux_card("trash_1_5",  day["aux"]["trash_1_5"],  "orange",
                        extra_tasks=sweeper_add.get("aux_trash_1_5", []),
                        alert=alerts.get("aux_trash_1_5", ""),
                        group=BG_AUX.get("trash_1_5")),
        render_aux_card("trash_6_10", day["aux"]["trash_6_10"], "orange",
                        extra_tasks=sweeper_add.get("aux_trash_6_10", []),
                        alert=alerts.get("aux_trash_6_10", ""),
                        group=BG_AUX.get("trash_6_10")),
        render_aux_card("support_1",  day["aux"]["support_1"],  "grey",
                        alert=alerts.get("aux_support_1", ""),
                        group=BG_AUX.get("support_1")),
        render_aux_card("support_2",  day["aux"]["support_2"],  "grey",
                        alert=alerts.get("aux_support_2", ""),
                        group=BG_AUX.get("support_2")),
    ]
    # Support 3 is a conditional overflow slot — only render the card when populated.
    if support_3:
        aux_cards.append(
            render_aux_card("support_3", support_3, "teal",
                            alert=alerts.get("aux_support_3", ""),
                            group=BG_AUX.get("support_3"),
                            conditional=True))
    aux_html = "\n".join(aux_cards)
    aux_strip_cls = "aux-strip" + (" has-support-3" if support_3 else "")

    # Restrooms — pass per-side training markers when the trainee is placed in
    # one of the RR slots (D4 Shadow Restroom).
    rr_nums = [1, 6, 7, 8, 10]
    rr_html = "\n".join(
        render_rr_card(n, day["rr_mens"][i], day["rr_womens"][i], RR_COLOR[n],
                       extra_tasks=sweeper_add.get(f"rr_{n}", []),
                       alert=alerts.get(f"rr_{n}", ""),
                       mens_group=BG_RR_M.get(n),
                       womens_group=BG_RR_W.get(n),
                       mens_training=(trainee_rr_side == "mens" and trainee_rr_idx == i),
                       womens_training=(trainee_rr_side == "womens" and trainee_rr_idx == i))
        for i, n in enumerate(rr_nums)
    )

    # Overlaps
    pm_tasks = list(TASKS_PM_OL); am_tasks = list(TASKS_AM_OL)
    overrides_today = OVERLAP_OVERRIDES.get(iso, {}) or {}

    def _apply(task_list, slot_prefix, ovr):
        for k, v in (ovr or {}).items():
            if isinstance(k, int) and 0 <= k < len(task_list):
                task_list[k] = v
            elif isinstance(k, str) and k.upper().startswith(slot_prefix):
                idx2 = int(k[len(slot_prefix):]) - 1
                if 0 <= idx2 < len(task_list): task_list[idx2] = v

    _apply(pm_tasks, "PMOL", overrides_today.get("PM") or overrides_today.get("pm"))
    _apply(am_tasks, "AMOL", overrides_today.get("AM") or overrides_today.get("am"))
    pm_minis = "".join(render_overlap_mini(day["pm_ol"][i], pm_tasks[i]) for i in range(6))
    am_minis = "".join(render_overlap_mini(day["am_ol"][i], am_tasks[i]) for i in range(6))

    month_name = day["date"].strftime("%B %Y")
    page_num   = 2 * idx - 1
    page_total = 2 * total

    return f"""<article class="page" style="--day-color:{day_color};">
  <header class="mast">
    <div class="mast-day">{day['day_num']}</div>
    <div class="mast-meta">
      <div class="day-name">{weekday}</div>
      <div class="month-str">{month_name} · Day {idx} of {total}</div>
      <div class="status-row">
        <span class="stat"><span class="num">{sum(v[0] for v in cs.values())}</span><span class="lbl">Filled</span></span>
        <span class="stat"><span class="num">{sum(v[1] - v[0] for v in cs.values())}</span><span class="lbl">Open</span></span>
        <span class="break-bar">
          <span class="lbl">Breaks</span>
          <span class="break-dot g1">{g1}</span>
          <span class="break-dot g2">{g2}</span>
          <span class="break-dot g3">{g3}</span>
        </span>
      </div>
    </div>
    <div class="mast-right">
      <div class="shift-label">Grave · 11pm – 7am</div>
      <div class="week-dots">
        {''.join(
          f'<div class="week-dot{"  cur" if i == current_idx else ""}">{d}</div>'
          for i, d in enumerate(['F','S','S','M','T','W','T'])
        )}
      </div>
      <div class="group-key">
        Group <span class="gp g1">1</span><span class="gp g2">2</span><span class="gp g3">3</span>
      </div>
    </div>
  </header>
  <div class="body">
    <section>
      <h2 class="section-lbl">
        Zones <span class="meta">{cs["zones"][0]} / {cs["zones"][1]} filled</span>
      </h2>
      <div class="zones-grid">
{zones_html}
      </div>
    </section>
    <section>
      <h2 class="section-lbl">
        Restrooms <span class="meta">{cs["rr"][0]} / {cs["rr"][1]} filled</span>
      </h2>
      <div class="rr-grid">
{rr_html}
      </div>
    </section>
    <section>
      <h2 class="section-lbl">
        Auxiliary <span class="meta">{cs["aux"][0]} / {cs["aux"][1]} filled</span>
      </h2>
      <div class="{aux_strip_cls}">
{aux_html}
      </div>
    </section>
    <section>
      <h2 class="section-lbl">
        Overlaps <span class="meta">{cs["overlaps"][0]} / {cs["overlaps"][1]} filled</span>
      </h2>
      <div class="overlap-row">
        <div class="overlap-time">11p – 1a<span class="kind">Late evening</span></div>
        <div class="overlap-mini-grid">{pm_minis}</div>
      </div>
      <div class="overlap-row">
        <div class="overlap-time">5a – 7a<span class="kind">Early AM</span></div>
        <div class="overlap-mini-grid">{am_minis}</div>
      </div>
    </section>
  </div>
  <footer class="page-foot">
    <span class="foot-mark"><span class="swatch"></span>GLCR · Grave</span>
    <span class="foot-center"><span class="now">{weekday}</span> {esc(day['date_short'])} · Zone Deployment</span>
    <span class="foot-pn"><span class="cur">{page_num}</span> / {page_total}</span>
  </footer>
</article>"""

def render_break_sheet_page(day, idx, total, males=None, no_sweeper_tms=None, current_idx=None):
    """LANDSCAPE break sheet (Phase 6, 5/6/26 — page 2 of each day).

    Buckets every staffed slot into Group 1/2/3 break waves and renders three
    columns of rows. Uses break-cols 3-column grid layout at landscape scale."""
    males = males or set()
    no_sweeper_tms = no_sweeper_tms or set()
    weekday   = day["weekday"]
    day_color = DAY_COLOR.get(weekday, "#444444")

    sweeper_add = _compute_sweeper_add(day, males, no_sweeper_tms)
    groups      = build_break_groups(day, sweeper_add)
    g1, g2, g3  = count_break_groups(day)
    in_rotation = total_in_rotation(day)

    cols_html = "\n".join(render_break_col(g, groups[g]) for g in (1, 2, 3))
    month_name = day["date"].strftime("%B %Y")
    page_num   = 2 * idx
    page_total = 2 * total

    # Phase 4g: overlaps on break sheet — same task-lookup logic as render_day_page
    iso_bs       = day["date"].isoformat()
    pm_tasks_bs  = list(TASKS_PM_OL); am_tasks_bs = list(TASKS_AM_OL)
    ovr_bs       = OVERLAP_OVERRIDES.get(iso_bs, {}) or {}
    def _apply_bs(task_list, slot_prefix, ovr):
        for k, v in (ovr or {}).items():
            if isinstance(k, int) and 0 <= k < len(task_list): task_list[k] = v
            elif isinstance(k, str) and k.upper().startswith(slot_prefix):
                idx2 = int(k[len(slot_prefix):]) - 1
                if 0 <= idx2 < len(task_list): task_list[idx2] = v
    _apply_bs(pm_tasks_bs, "PMOL", ovr_bs.get("PM") or ovr_bs.get("pm"))
    _apply_bs(am_tasks_bs, "AMOL", ovr_bs.get("AM") or ovr_bs.get("am"))
    pm_minis_bs = "".join(render_overlap_mini(day["pm_ol"][i], pm_tasks_bs[i]) for i in range(6))
    am_minis_bs = "".join(render_overlap_mini(day["am_ol"][i], am_tasks_bs[i]) for i in range(6))

    week_dots_bs = ''.join(
        f'<div class="week-dot{"  cur" if current_idx is not None and i == current_idx else ""}">{d}</div>'
        for i, d in enumerate(['F','S','S','M','T','W','T'])
    )

    return f"""<article class="page" style="--day-color:{day_color};">
  <header class="mast">
    <div class="mast-day outline">{day['day_num']}</div>
    <div class="mast-meta">
      <div class="day-name">Break Sheet</div>
      <div class="month-str">{weekday} · {month_name}</div>
      <div class="status-row">
        <span class="stat"><span class="num">{in_rotation}</span><span class="lbl">In Rotation</span></span>
        <span class="break-bar">
          <span class="lbl">Breaks</span>
          <span class="break-dot g1">{g1}</span>
          <span class="break-dot g2">{g2}</span>
          <span class="break-dot g3">{g3}</span>
        </span>
      </div>
    </div>
    <div class="mast-right">
      <div class="shift-label">By Break Wave</div>
      <div class="week-dots" style="margin-top:4px;">
        {week_dots_bs}
      </div>
    </div>
  </header>
  <div class="body">
    <div class="break-cols" style="max-height:5.4in; overflow:hidden;">
{cols_html}
    </div>
    <!-- Phase 4g polish: full-width hairline separator pulls the OVERLAPS
         section visually away from the Break 1-2-3 columns above. Matches the
         deployment-page treatment so both pages feel like the same family. -->
    <section class="overlaps-section" style="margin-top:18px; padding-top:14px; border-top:1px solid var(--hairline-strong);">
      <h2 class="section-lbl" style="margin-bottom:6px;">
        Overlaps <span class="meta">11p–1a &amp; 5a–7a</span>
      </h2>
      <div class="overlap-row">
        <div class="overlap-time">11p – 1a<span class="kind">Late evening</span></div>
        <div class="overlap-mini-grid">{pm_minis_bs}</div>
      </div>
      <div class="overlap-row">
        <div class="overlap-time">5a – 7a<span class="kind">Early AM</span></div>
        <div class="overlap-mini-grid">{am_minis_bs}</div>
      </div>
    </section>
  </div>
  <footer class="page-foot">
    <span class="foot-mark"><span class="swatch"></span>GLCR · Grave</span>
    <span class="foot-center"><span class="now">{weekday}</span> {esc(day['date_short'])} · Break Sheet</span>
    <span class="foot-pn"><span class="cur">{page_num}</span> / {page_total}</span>
  </footer>
</article>"""

# --------------------------------------------------------------------------
# CSS — LANDSCAPE REDESIGN (Phase 6, 5/6/26)
# Lifted directly from design/zds_redesign/zds_print_landscape_template.html
# Operational color taxonomy (DAY_COLORS, ZONE_COLOR, RR_COLOR) UNCHANGED.
# Page model: 11in × 8.5in landscape. CSS variables inject --day-color and
# --card-color per-page and per-card. Renders as-is with no mediator layer.
# Casino scatter background: skipped for v1 (see template lines 219-234).
# --------------------------------------------------------------------------

CSS = r"""
/* Landscape redesign (5/6/26): lift Barlow weights; drop Atkinson Hyperlegible
   (portrait accessibility face). Landscape is read at table distance, not floor
   distance, so viewport is larger. Landscape uses the template's ink palette. */
/* Phase 4g polish: switched primary type from Barlow → Atkinson Hyperlegible.
   Atkinson is purpose-built for low-vision accessibility — distinct letterforms
   (no I/l confusion, open apertures, asymmetric b/d/p/q) and prints crisply at
   small sizes. Barlow kept as fallback for browsers that fail the @import. */
@import url('https://fonts.googleapis.com/css2?family=Atkinson+Hyperlegible:wght@400;700&family=Barlow:wght@400;500;600;700;800&display=swap');

:root {
  --safe: 0.32in;
  --page-w: 11in;
  --page-h: 8.5in;

  --ink-900: #0b1a2a;
  --ink-700: #2c3e54;
  --ink-500: #5a6b7d;
  --ink-300: #94a2b1;
  --ink-200: #c8d3dc;
  --ink-100: #eef1f4;
  --gold:    #c9a96e;
  --gold-lt: #e8d5b4;
  --hairline: rgba(11,26,42,0.10);
  --hairline-strong: rgba(11,26,42,0.20);

  /* ── Zone / RR / Aux color palette (Phase 4g hotfix) ─────────────────
     Sometime before Phase 4g these definitions were lost from this file.
     The .c-yellow / .c-red / etc. classes farther down all reference these
     vars (e.g. `.c-yellow { --card-color: var(--c-yellow); }`). Without
     them, --card-color resolved to nothing and the .zone-card::before
     5px top stripe rendered as transparent — which is exactly what Brian
     saw: cards with no colored top accents.

     Palette matches what was working in the Week-of-2026-05-07 PDF.    */
  --c-yellow: #B89708;   /* Z1, Z2, RR 1+2, Admin                 */
  --c-red:    #E53935;   /* Z3, Z4, Z5, Z9, Z9 SR                 */
  --c-pink:   #B7679A;   /* Z6, RR 6                              */
  --c-blue:   #1E88E5;   /* Z7, RR 7                              */
  --c-brown:  #6B5346;   /* Z8, RR 8                              */
  --c-green:  #43A047;   /* Z10, RR 10, Support 3                 */
  --c-orange: #FB8C00;   /* Trash 1, Trash 2                      */
  --c-purple: #8E24AA;   /* Admin                                 */
  --c-grey:   #4a5568;   /* Support 1, Support 2                  */
  --c-teal:   #14b8a6;   /* fallback                              */
  --c-alert:  #e53935;   /* used by alert dot bg                  */

  /* Body text uses Atkinson (legibility); display headings keep Barlow's
     condensed feel via --font-display. */
  --font: 'Atkinson Hyperlegible', 'Barlow', 'Helvetica Neue', Arial, sans-serif;
  --font-display: 'Barlow', 'Atkinson Hyperlegible', 'Helvetica Neue', Arial, sans-serif;
}

html, body { margin: 0; padding: 0; background: #d0d4d8;
  font-family: var(--font); -webkit-font-smoothing: antialiased; }
body { display: flex; flex-direction: column; align-items: center; padding: 32px 24px; gap: 28px; }

.page {
  width: var(--page-w); height: var(--page-h);
  background: #fff;
  box-shadow: 0 12px 48px rgba(0,0,0,0.22);
  display: grid;
  grid-template-rows: auto 1fr auto;
  overflow: hidden;
  font-feature-settings: "tnum","ss01";
  position: relative;
}
/* Phase 4g polish: casino-scatter pattern (hearts/diamonds/clubs/spades)
   layered behind everything via .page::before. The cards have solid white
   backgrounds so they cover the pattern where they sit; the scatter is
   visible only in the gaps + bottom-of-page whitespace. Matches the IDEAL
   Friday May 1 mockup. */
.page::before {
  content: "";
  position: absolute;
  inset: 0;
  opacity: 0.7;  /* Phase 4j: dial scatter to 70% of baked levels — adjust this one knob */
  background: url("data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHZpZXdCb3g9IjAgMCAxMDU2IDgxNiIgd2lkdGg9IjEwNTYiIGhlaWdodD0iODE2Ij4KICA8ZGVmcz4KICAgIAogICAgPHBhdGggaWQ9ImhlYXJ0IiBkPSJNMCwtMTAgQzAsLTE4IC0xMiwtMjAgLTEyLC0xMCBDLTEyLDAgMCwxMCAwLDEwIEMwLDEwIDEyLDAgMTIsLTEwIEMxMiwtMjAgMCwtMTggMCwtMTB6Ij48L3BhdGg+CiAgICAKICAgIDxwYXRoIGlkPSJkaWFtb25kIiBkPSJNMCwtMTQgTDEwLDAgTDAsMTQgTC0xMCwweiI+PC9wYXRoPgogICAgCiAgICA8cGF0aCBpZD0iY2x1YiIgZD0iTTAsMTIgTC01LDEyIEwtMyw2IEMtOCw1IC0xMCwwIC03LC00IEMtNCwtOCAyLC04IDQsLTUgQzQsLTkgNywtMTIgMTAsLTEwIEMxMywtOCAxMywtMyAxMCwwIEMxMywzIDEzLDggMTAsMTAgQzcsMTIgMywxMCAyLDcgQzEsMTAgLTEsMTIgMCwxMnogTS02LDEyIEw2LDEyIj48L3BhdGg+CiAgICAKICAgIDxwYXRoIGlkPSJzcGFkZSIgZD0iTTAsLTEyIEMtOCwtNCAtMTIsNCAtOCw4IEMtNSwxMSAtMiwxMCAwLDggQy0xLDExIC0zLDEzIC01LDEzIEw1LDEzIEMzLDEzIDEsMTEgMCw4IEMyLDEwIDUsMTEgOCw4IEMxMiw0IDgsLTQgMCwtMTJ6Ij48L3BhdGg+CiAgPC9kZWZzPgoKICAKCiAgCiAgPGcgdHJhbnNmb3JtPSJ0cmFuc2xhdGUoNTgsNzIwKSByb3RhdGUoLTE1KSI+CiAgICA8cGF0aCBkPSJNMCwtMjIgTDE2LDAgTDAsMjIgTC0xNiwweiIgZmlsbD0iIzk0YTJiMSIgb3BhY2l0eT0iMC4zNSI+PC9wYXRoPgogIDwvZz4KCiAgCiAgPGcgdHJhbnNmb3JtPSJ0cmFuc2xhdGUoMzgsODEwKSByb3RhdGUoMTApIHNjYWxlKDAuNikiPgogICAgPHBhdGggZD0iTTAsLTE0IEMtOSwtNSAtMTQsNCAtOSw5IEMtNiwxMiAtMiwxMSAwLDkgQy0xLDEyIC0zLDE0IC02LDE0IEw2LDE0IEMzLDE0IDEsMTIgMCw5IEMyLDExIDYsMTIgOSw5IEMxNCw0IDksLTUgMCwtMTR6IiBmaWxsPSIjOTRhMmIxIiBvcGFjaXR5PSIwLjI1Ij48L3BhdGg+CiAgPC9nPgoKICAKICA8ZyB0cmFuc2Zvcm09InRyYW5zbGF0ZSgxNjgsNzYwKSByb3RhdGUoLTMpIj4KICAgIDxyZWN0IHg9Ii0yNiIgeT0iLTM2IiB3aWR0aD0iNTIiIGhlaWdodD0iNzIiIHJ4PSI1IiByeT0iNSIgZmlsbD0id2hpdGUiIHN0cm9rZT0iIzMwYjJmZiIgc3Ryb2tlLXdpZHRoPSIyLjUiIG9wYWNpdHk9IjAuNTUiPjwvcmVjdD4KICAgIDx0ZXh0IHg9Ii0yMCIgeT0iLTIyIiBmb250LWZhbWlseT0iQmFybG93LHNhbnMtc2VyaWYiIGZvbnQtc2l6ZT0iMTMiIGZvbnQtd2VpZ2h0PSI3MDAiIGZpbGw9IiMzMGIyZmYiIG9wYWNpdHk9IjAuNTUiPkE8L3RleHQ+CiAgICA8cGF0aCBkPSJNMCwtNCBDLTUsLTQgLTYsMCAtMywzIEMtMSw1IDEsNSAzLDMgQzYsMCA1LC00IDAsLTR6IiBmaWxsPSIjMzBiMmZmIiBvcGFjaXR5PSIwLjUiIHRyYW5zZm9ybT0idHJhbnNsYXRlKDAsMTYpIj48L3BhdGg+CiAgPC9nPgoKICAKICA8ZyB0cmFuc2Zvcm09InRyYW5zbGF0ZSgyOTgsNzkzKSByb3RhdGUoNSkgc2NhbGUoMC44KSI+CiAgICA8Y2lyY2xlIGN4PSIwIiBjeT0iLTYiIHI9IjYiIGZpbGw9IiMzMGIyZmYiIG9wYWNpdHk9IjAuMjUiPjwvY2lyY2xlPgogICAgPGNpcmNsZSBjeD0iLTYiIGN5PSIwIiByPSI1IiBmaWxsPSIjMzBiMmZmIiBvcGFjaXR5PSIwLjI1Ij48L2NpcmNsZT4KICAgIDxjaXJjbGUgY3g9IjYiIGN5PSIwIiByPSI1IiBmaWxsPSIjMzBiMmZmIiBvcGFjaXR5PSIwLjI1Ij48L2NpcmNsZT4KICAgIDxyZWN0IHg9Ii0yIiB5PSI0IiB3aWR0aD0iNCIgaGVpZ2h0PSI4IiBmaWxsPSIjMzBiMmZmIiBvcGFjaXR5PSIwLjI1Ij48L3JlY3Q+CiAgICA8cmVjdCB4PSItNiIgeT0iMTEiIHdpZHRoPSIxMiIgaGVpZ2h0PSIyIiBmaWxsPSIjMzBiMmZmIiBvcGFjaXR5PSIwLjI1Ij48L3JlY3Q+CiAgPC9nPgoKICAKICA8ZyB0cmFuc2Zvcm09InRyYW5zbGF0ZSg0NjAsNzYwKSByb3RhdGUoMikiPgogICAgPHJlY3QgeD0iLTIyIiB5PSItMzIiIHdpZHRoPSI0NCIgaGVpZ2h0PSI2NCIgcng9IjUiIHJ5PSI1IiBmaWxsPSJ3aGl0ZSIgc3Ryb2tlPSIjOTRhMmIxIiBzdHJva2Utd2lkdGg9IjEuNSIgb3BhY2l0eT0iMC40MCI+PC9yZWN0PgogICAgPHRleHQgeD0iLTE1IiB5PSItMTgiIGZvbnQtZmFtaWx5PSJCYXJsb3csc2Fucy1zZXJpZiIgZm9udC1zaXplPSIxMiIgZm9udC13ZWlnaHQ9IjYwMCIgZmlsbD0iIzk0YTJiMSIgb3BhY2l0eT0iMC40MCI+QTwvdGV4dD4KICA8L2c+CgogIAoKICAKICA8ZyB0cmFuc2Zvcm09InRyYW5zbGF0ZSg0MzAsNzAwKSByb3RhdGUoLTgpIHNjYWxlKDIuOCkiPgogICAgPGNpcmNsZSBjeD0iMCIgY3k9Ii04IiByPSI4IiBmaWxsPSIjZTBjYmI2IiBvcGFjaXR5PSIwLjE4Ij48L2NpcmNsZT4KICAgIDxjaXJjbGUgY3g9Ii04IiBjeT0iMCIgcj0iNyIgZmlsbD0iI2UwY2JiNiIgb3BhY2l0eT0iMC4xOCI+PC9jaXJjbGU+CiAgICA8Y2lyY2xlIGN4PSI4IiBjeT0iMCIgcj0iNyIgZmlsbD0iI2UwY2JiNiIgb3BhY2l0eT0iMC4xOCI+PC9jaXJjbGU+CiAgICA8cmVjdCB4PSItMyIgeT0iNyIgd2lkdGg9IjYiIGhlaWdodD0iMTAiIGZpbGw9IiNlMGNiYjYiIG9wYWNpdHk9IjAuMTgiPjwvcmVjdD4KICAgIDxyZWN0IHg9Ii04IiB5PSIxNiIgd2lkdGg9IjE2IiBoZWlnaHQ9IjIuNSIgZmlsbD0iI2UwY2JiNiIgb3BhY2l0eT0iMC4xOCI+PC9yZWN0PgogIDwvZz4KCiAgCiAgPGcgdHJhbnNmb3JtPSJ0cmFuc2xhdGUoMzQ4LDc2MCkgcm90YXRlKC0yMCkgc2NhbGUoMS42KSI+CiAgICA8cGF0aCBkPSJNMCwxMiBDMCwxMiAtMTQsMiAtMTQsLTYgQy0xNCwtMTIgLTgsLTE0IDAsLTggQzgsLTE0IDE0LC0xMiAxNCwtNiBDMTQsMiAwLDEyIDAsMTJ6IiBmaWxsPSIjZTBjYmI2IiBvcGFjaXR5PSIwLjIyIj48L3BhdGg+CiAgPC9nPgoKICAKICA8ZyB0cmFuc2Zvcm09InRyYW5zbGF0ZSg1MTAsNzU2KSByb3RhdGUoMTIpIHNjYWxlKDEuMikiPgogICAgPHBhdGggZD0iTTAsMTAgQzAsMTAgLTEyLDEgLTEyLC01IEMtMTIsLTEwIC02LC0xMiAwLC03IEM2LC0xMiAxMiwtMTAgMTIsLTUgQzEyLDEgMCwxMCAwLDEweiIgZmlsbD0iIzk0YTJiMSIgb3BhY2l0eT0iMC4yMCI+PC9wYXRoPgogIDwvZz4KCiAgCiAgPGcgdHJhbnNmb3JtPSJ0cmFuc2xhdGUoNjAyLDY0MCkgcm90YXRlKDIyKSBzY2FsZSgwLjkpIj4KICAgIDxwYXRoIGQ9Ik0wLC0xNiBMMTIsMCBMMCwxNiBMLTEyLDB6IiBmaWxsPSIjZTBjYmI2IiBvcGFjaXR5PSIwLjIyIj48L3BhdGg+CiAgPC9nPgoKICAKICA8ZyB0cmFuc2Zvcm09InRyYW5zbGF0ZSg1NjYsNzEyKSByb3RhdGUoOCkgc2NhbGUoMS40KSI+CiAgICA8cGF0aCBkPSJNMCwtMTQgQy05LC01IC0xNCw0IC05LDkgQy02LDEyIC0yLDExIDAsOSBDLTEsMTIgLTMsMTQgLTYsMTQgTDYsMTQgQzMsMTQgMSwxMiAwLDkgQzIsMTEgNiwxMiA5LDkgQzE0LDQgOSwtNSAwLC0xNHoiIGZpbGw9IiM5NGEyYjEiIG9wYWNpdHk9IjAuMjIiPjwvcGF0aD4KICA8L2c+CgogIAogIDxnIHRyYW5zZm9ybT0idHJhbnNsYXRlKDU4NCw3ODgpIHJvdGF0ZSgtNSkgc2NhbGUoMS4zKSI+CiAgICA8cGF0aCBkPSJNMCwxMCBDMCwxMCAtMTIsMSAtMTIsLTUgQy0xMiwtMTAgLTYsLTEyIDAsLTcgQzYsLTEyIDEyLC0xMCAxMiwtNSBDMTIsMSAwLDEwIDAsMTB6IiBmaWxsPSIjZTBjYmI2IiBvcGFjaXR5PSIwLjE4Ij48L3BhdGg+CiAgPC9nPgoKICAKICA8ZyB0cmFuc2Zvcm09InRyYW5zbGF0ZSg1MTIsNjY4KSI+CiAgICA8Y2lyY2xlIGN4PSIwIiBjeT0iMCIgcj0iMjgiIGZpbGw9Im5vbmUiIHN0cm9rZT0iIzk0YTJiMSIgc3Ryb2tlLXdpZHRoPSIyIiBzdHJva2UtZGFzaGFycmF5PSI2IDUiIG9wYWNpdHk9IjAuMTgiPjwvY2lyY2xlPgogICAgPGNpcmNsZSBjeD0iMCIgY3k9IjAiIHI9IjE4IiBmaWxsPSJub25lIiBzdHJva2U9IiM5NGEyYjEiIHN0cm9rZS13aWR0aD0iMSIgb3BhY2l0eT0iMC4xNCI+PC9jaXJjbGU+CiAgICA8Y2lyY2xlIGN4PSIwIiBjeT0iMCIgcj0iOCIgZmlsbD0ibm9uZSIgc3Ryb2tlPSIjOTRhMmIxIiBzdHJva2Utd2lkdGg9IjEuNSIgb3BhY2l0eT0iMC4xNCI+PC9jaXJjbGU+CiAgPC9nPgoKICAKCiAgCiAgPGcgdHJhbnNmb3JtPSJ0cmFuc2xhdGUoNzU2LDY4MCkgcm90YXRlKDE4KSI+CiAgICA8cmVjdCB4PSItNDQiIHk9Ii02MCIgd2lkdGg9Ijg4IiBoZWlnaHQ9IjEyMCIgcng9IjgiIHJ5PSI4IiBmaWxsPSJ3aGl0ZSIgc3Ryb2tlPSIjMzBiMmZmIiBzdHJva2Utd2lkdGg9IjIuNSIgb3BhY2l0eT0iMC40NSI+PC9yZWN0PgogICAgPHRleHQgeD0iLTM0IiB5PSItNDAiIGZvbnQtZmFtaWx5PSJCYXJsb3csc2Fucy1zZXJpZiIgZm9udC1zaXplPSIxOCIgZm9udC13ZWlnaHQ9IjcwMCIgZmlsbD0iIzMwYjJmZiIgb3BhY2l0eT0iMC40NSI+QTwvdGV4dD4KICAgIAogICAgPGcgdHJhbnNmb3JtPSJ0cmFuc2xhdGUoOCwgMTApIiBvcGFjaXR5PSIwLjQiPgogICAgICA8Y2lyY2xlIGN4PSIwIiBjeT0iLTgiIHI9IjgiIGZpbGw9IiMzMGIyZmYiPjwvY2lyY2xlPgogICAgICA8Y2lyY2xlIGN4PSItOCIgY3k9IjAiIHI9IjciIGZpbGw9IiMzMGIyZmYiPjwvY2lyY2xlPgogICAgICA8Y2lyY2xlIGN4PSI4IiBjeT0iMCIgcj0iNyIgZmlsbD0iIzMwYjJmZiI+PC9jaXJjbGU+CiAgICAgIDxyZWN0IHg9Ii0zIiB5PSI3IiB3aWR0aD0iNiIgaGVpZ2h0PSI5IiBmaWxsPSIjMzBiMmZmIj48L3JlY3Q+CiAgICAgIDxyZWN0IHg9Ii03IiB5PSIxNSIgd2lkdGg9IjE0IiBoZWlnaHQ9IjIuNSIgZmlsbD0iIzMwYjJmZiI+PC9yZWN0PgogICAgPC9nPgogIDwvZz4KCiAgCiAgPGcgdHJhbnNmb3JtPSJ0cmFuc2xhdGUoNzA2LDY5MCkgcm90YXRlKC0xMikiPgogICAgPHJlY3QgeD0iLTQ0IiB5PSItNjAiIHdpZHRoPSI4OCIgaGVpZ2h0PSIxMjAiIHJ4PSI4IiByeT0iOCIgZmlsbD0id2hpdGUiIHN0cm9rZT0iIzMwYjJmZiIgc3Ryb2tlLXdpZHRoPSIyLjUiIG9wYWNpdHk9IjAuNTUiPjwvcmVjdD4KICAgIDx0ZXh0IHg9Ii0zNCIgeT0iLTQwIiBmb250LWZhbWlseT0iQmFybG93LHNhbnMtc2VyaWYiIGZvbnQtc2l6ZT0iMTgiIGZvbnQtd2VpZ2h0PSI3MDAiIGZpbGw9IiMzMGIyZmYiIG9wYWNpdHk9IjAuNTUiPkE8L3RleHQ+CiAgICA8dGV4dCB4PSItMzQiIHk9Ii0yMiIgZm9udC1mYW1pbHk9IkJhcmxvdyxzYW5zLXNlcmlmIiBmb250LXNpemU9IjExIiBmb250LXdlaWdodD0iNjAwIiBmaWxsPSIjMzBiMmZmIiBvcGFjaXR5PSIwLjU1Ij7imaM8L3RleHQ+CiAgICAKICAgIDxwYXRoIGQ9Ik0wLC0yMiBMMTYsMCBMMCwyMiBMLTE2LDB6IiBmaWxsPSIjMzBiMmZmIiBvcGFjaXR5PSIwLjMwIiB0cmFuc2Zvcm09InRyYW5zbGF0ZSgwLDE0KSI+PC9wYXRoPgogIDwvZz4KCiAgCgogIAogIDxnIHRyYW5zZm9ybT0idHJhbnNsYXRlKDg0MCw2NjApIHJvdGF0ZSgtMTApIHNjYWxlKDEuNSkiPgogICAgPHBhdGggZD0iTTAsMTIgQzAsMTIgLTE0LDIgLTE0LC02IEMtMTQsLTEyIC04LC0xNCAwLC04IEM4LC0xNCAxNCwtMTIgMTQsLTYgQzE0LDIgMCwxMiAwLDEyeiIgZmlsbD0iIzMwYjJmZiIgb3BhY2l0eT0iMC4zMCI+PC9wYXRoPgogIDwvZz4KCiAgCiAgPGcgdHJhbnNmb3JtPSJ0cmFuc2xhdGUoOTIwLDY3MCkgcm90YXRlKDgpIHNjYWxlKDEuMikiPgogICAgPHBhdGggZD0iTTAsLTE4IEwxNCwwIEwwLDE4IEwtMTQsMHoiIGZpbGw9IiNlMGNiYjYiIG9wYWNpdHk9IjAuMjgiPjwvcGF0aD4KICA8L2c+CgogIAogIDxnIHRyYW5zZm9ybT0idHJhbnNsYXRlKDkyMCw3NDgpIHJvdGF0ZSgtNSkgc2NhbGUoMS4xKSI+CiAgICA8cGF0aCBkPSJNMCwtMTYgTDEyLDAgTDAsMTYgTC0xMiwweiIgZmlsbD0iIzk0YTJiMSIgb3BhY2l0eT0iMC4yOCI+PC9wYXRoPgogIDwvZz4KCiAgCiAgPGcgdHJhbnNmb3JtPSJ0cmFuc2xhdGUoMTAxMCw3OTApIHNjYWxlKDIuMikiPgogICAgPGNpcmNsZSBjeD0iMCIgY3k9Ii05IiByPSI5IiBmaWxsPSIjMzBiMmZmIiBvcGFjaXR5PSIwLjI4Ij48L2NpcmNsZT4KICAgIDxjaXJjbGUgY3g9Ii05IiBjeT0iMCIgcj0iOCIgZmlsbD0iIzMwYjJmZiIgb3BhY2l0eT0iMC4yOCI+PC9jaXJjbGU+CiAgICA8Y2lyY2xlIGN4PSI5IiBjeT0iMCIgcj0iOCIgZmlsbD0iIzMwYjJmZiIgb3BhY2l0eT0iMC4yOCI+PC9jaXJjbGU+CiAgICA8cmVjdCB4PSItMy41IiB5PSI4IiB3aWR0aD0iNyIgaGVpZ2h0PSIxMSIgZmlsbD0iIzMwYjJmZiIgb3BhY2l0eT0iMC4yOCI+PC9yZWN0PgogICAgPHJlY3QgeD0iLTkiIHk9IjE4IiB3aWR0aD0iMTgiIGhlaWdodD0iMyIgZmlsbD0iIzMwYjJmZiIgb3BhY2l0eT0iMC4yOCI+PC9yZWN0PgogIDwvZz4KCiAgCiAgPGcgdHJhbnNmb3JtPSJ0cmFuc2xhdGUoMTA0MCw3MTApIHJvdGF0ZSgtMjApIHNjYWxlKDAuOSkiPgogICAgPHBhdGggZD0iTTAsLTE0IEMtOSwtNSAtMTQsNCAtOSw5IEMtNiwxMiAtMiwxMSAwLDkgQy0xLDEyIC0zLDE0IC02LDE0IEw2LDE0IEMzLDE0IDEsMTIgMCw5IEMyLDExIDYsMTIgOSw5IEMxNCw0IDksLTUgMCwtMTR6IiBmaWxsPSIjZTBjYmI2IiBvcGFjaXR5PSIwLjI1Ij48L3BhdGg+CiAgPC9nPgoKICAKCiAgCiAgPGcgdHJhbnNmb3JtPSJ0cmFuc2xhdGUoMjY4LDU5Nikgcm90YXRlKDI1KSBzY2FsZSgwLjkpIj4KICAgIDxwYXRoIGQ9Ik0wLDEwIEMwLDEwIC0xMSwyIC0xMSwtNCBDLTExLC05IC01LC0xMSAwLC02IEM1LC0xMSAxMSwtOSAxMSwtNCBDMTEsMiAwLDEwIDAsMTB6IiBmaWxsPSIjZTBjYmI2IiBvcGFjaXR5PSIwLjIwIj48L3BhdGg+CiAgPC9nPgoKICAKICA8ZyB0cmFuc2Zvcm09InRyYW5zbGF0ZSg2MzAsNTk4KSByb3RhdGUoLTMwKSBzY2FsZSgwLjgpIj4KICAgIDxwYXRoIGQ9Ik0wLC0xNCBMMTAsMCBMMCwxNCBMLTEwLDB6IiBmaWxsPSIjOTRhMmIxIiBvcGFjaXR5PSIwLjE4Ij48L3BhdGg+CiAgPC9nPgoKICAKICA8ZyB0cmFuc2Zvcm09InRyYW5zbGF0ZSgxNzYsNjU2KSByb3RhdGUoMTUpIHNjYWxlKDEuMSkiPgogICAgPHBhdGggZD0iTTAsLTE0IEMtOSwtNSAtMTQsNCAtOSw5IEMtNiwxMiAtMiwxMSAwLDkgQy0xLDEyIC0zLDE0IC02LDE0IEw2LDE0IEMzLDE0IDEsMTIgMCw5IEMyLDExIDYsMTIgOSw5IEMxNCw0IDksLTUgMCwtMTR6IiBmaWxsPSIjZTBjYmI2IiBvcGFjaXR5PSIwLjE2Ij48L3BhdGg+CiAgPC9nPgoKPC9zdmc+") no-repeat center / cover;
  pointer-events: none;
  z-index: 0;
}
/* Ensure the masthead, body, and footer all sit above the scatter. */
.page > * { position: relative; z-index: 1; }

/* ── MASTHEAD ── */
.mast {
  padding: 14px var(--safe) 10px;
  display: grid;
  grid-template-columns: auto 1fr auto;
  gap: 20px;
  align-items: end;
  border-bottom: 1px solid var(--hairline);
  position: relative;
}
.mast::after {
  content: '';
  position: absolute;
  left: 0; right: 0; bottom: -1px;
  height: 3px;
  background: var(--day-color);
}

.mast-day { font-weight: 800; font-size: 72px; line-height: 0.82; letter-spacing: -0.05em; color: var(--ink-900); font-variant-numeric: tabular-nums; }
.mast-day.outline { color: #fff; -webkit-text-stroke: 2.5px var(--ink-900); }

.mast-meta { display: grid; gap: 4px; padding-bottom: 8px; }
.day-name  { font-weight: 700; font-size: 28px; letter-spacing: -0.02em; color: var(--day-color); line-height: 1; }
.month-str { font-weight: 400; font-size: 12px; color: var(--ink-500); }
.status-row { display: flex; gap: 16px; margin-top: 5px; align-items: center; }
.stat { display: inline-flex; align-items: center; gap: 5px; }
.stat .num { font-weight: 700; font-size: 13px; color: var(--ink-900); font-variant-numeric: tabular-nums; }
.stat .lbl { font-size: 9.5px; letter-spacing: 0.10em; text-transform: uppercase; color: var(--ink-500); }

.break-bar { display: inline-flex; align-items: center; gap: 5px; }
.break-bar .lbl { font-size: 9.5px; letter-spacing: 0.10em; text-transform: uppercase; color: var(--ink-500); margin-right: 2px; }
.break-dot { width: 20px; height: 20px; border-radius: 50%; display: inline-flex; align-items: center; justify-content: center; font-weight: 800; font-size: 10.5px; font-variant-numeric: tabular-nums; }
.break-dot.g1 { background: #1a2332; color: #fff; }
.break-dot.g2 { background: #5a6b7d; color: #fff; }
.break-dot.g3 { background: #c8d3dc; color: var(--ink-900); }

.mast-right { display: grid; gap: 6px; padding-bottom: 8px; text-align: right; justify-items: end; }
.shift-label { font-weight: 600; font-size: 11px; letter-spacing: 0.06em; text-transform: uppercase; color: var(--ink-700); }
.week-dots { display: inline-flex; gap: 2px; }
.week-dot { width: 22px; height: 22px; border-radius: 5px; display: inline-flex; align-items: center; justify-content: center; font-size: 10px; font-weight: 500; color: var(--ink-300); font-variant-numeric: tabular-nums; }
.week-dot.cur { background: var(--day-color); color: #fff; font-weight: 700; }
.group-key { display: inline-flex; gap: 5px; align-items: center; font-size: 9px; letter-spacing: 0.10em; text-transform: uppercase; color: var(--ink-500); }
.gp { width: 16px; height: 16px; border-radius: 4px; display: inline-flex; align-items: center; justify-content: center; font-weight: 800; font-size: 10px; }
.gp.g1 { background: #1a2332; color: #fff; }
.gp.g2 { background: #5a6b7d; color: #fff; }
.gp.g3 { background: #c8d3dc; color: var(--ink-900); }

/* ── BODY ── */
.body { padding: 8px var(--safe) 6px; display: grid; grid-template-rows: minmax(0,1.4fr) minmax(0,0.85fr) auto auto; gap: 7px; min-height: 0; }

/* ── SECTION LABEL ── */
.section-lbl {
  /* Phase 4g polish: ink-500 → ink-700 weight bump for the section
     hierarchy. The previous color was washing out at print scale. */
  font-weight: 700; font-size: 10.5px; letter-spacing: 0.16em;
  text-transform: uppercase; color: var(--ink-700);
  margin: 0 0 8px 0;
  display: flex; align-items: center; gap: 7px;
  padding-bottom: 6px;
  border-bottom: 1px solid var(--gold-lt);
}
.section-lbl .meta { margin-left: auto; font-weight: 500; font-size: 9.5px; color: var(--ink-500); letter-spacing: 0.06em; }
.section-lbl svg { width: 12px; height: 12px; color: var(--ink-300); }

/* ── FOOTER ── */
.page-foot {
  display: grid; grid-template-columns: auto 1fr auto; align-items: center;
  gap: 14px; padding: 6px var(--safe) 7px;
  position: relative; border-top: 1px solid var(--hairline);
}
.page-foot::before { content: ''; position: absolute; left: 0; right: 0; top: -1px; height: 2px; background: var(--day-color); }
.foot-mark { display: inline-flex; align-items: center; gap: 7px; font-size: 9px; font-weight: 700; letter-spacing: 0.14em; text-transform: uppercase; color: var(--ink-500); }
.foot-mark .swatch { width: 8px; height: 8px; border-radius: 2px; background: var(--day-color); }
.foot-center { font-size: 9.5px; color: var(--ink-500); text-align: center; letter-spacing: 0.04em; font-variant-numeric: tabular-nums; }
.foot-center .now { color: var(--ink-900); font-weight: 700; }
.foot-pn { font-size: 9.5px; font-weight: 700; letter-spacing: 0.10em; color: var(--ink-500); font-variant-numeric: tabular-nums; }
.foot-pn .cur { color: var(--ink-900); }

/* Phase D — Night lock stamp in print footer */
.foot-lock-stamp {
  display: inline-flex; align-items: center; gap: 5px;
  font-size: 8.5px; font-weight: 700;
  color: #8a6a3c;   /* gold-on-white readable for print */
  letter-spacing: 0.10em; text-transform: uppercase;
}

/* Phase E — Notice badges in print (inline after TM name) */
.print-notice {
  display: inline-block;
  font-size: 8px; font-weight: 800;
  padding: 1px 5px; border-radius: 3px;
  margin-left: 3px; vertical-align: middle;
  line-height: 1.4;
}
.print-notice-alert    { background: #fef3c7; color: #b45309; }
.print-notice-info     { background: #dbeafe; color: #0065bf; }
.print-notice-training { background: #d1fae5; color: #047857; }
.print-notice-meeting  { background: #ede9fe; color: #5b21b6; }

/* MASTHEAD — bigger date, status row, group key */
.mast {
  padding: 12px var(--safe) 10px;
  display: grid;
  grid-template-columns: auto 1fr auto;
  gap: 18px;
  align-items: end;
  border-bottom: 1px solid var(--hairline);
  position: relative;
}
.mast::after {
  content:""; position:absolute; left:0; right:0; bottom:-1px; height:3px;
  background: var(--day-color);
}
.mast-day-num {
  font-weight: 800; font-size: 82px; line-height: 0.82;
  letter-spacing: -0.05em; color: var(--ink-900);
  font-variant-numeric: tabular-nums;
}
.mast-day-num.is-outline {
  font-weight: 800;
  color: #fff;
  -webkit-text-stroke: 2px var(--ink-900);
  text-stroke: 2px var(--ink-900);
}
.mast-meta { display: grid; gap: 4px; padding-bottom: 8px; }
.mast-meta .day-name {
  font-weight: 600; font-size: 26px; letter-spacing: -0.025em;
  color: var(--day-color); line-height: 1;
}
.mast-meta .month {
  font-weight: 400; font-size: 12px; color: var(--ink-500); line-height: 1;
}
.mast-meta .status {
  display: flex; gap: 14px; margin-top: 4px;
  font-size: 11px; font-weight: 500;
  align-items: center;
}
.mast-meta .status .stat { display: inline-flex; align-items: center; gap: 5px; }
.mast-meta .status .num {
  font-weight: 700; font-size: 13px; color: var(--ink-900);
  font-variant-numeric: tabular-nums; line-height: 1;
}
.mast-meta .status .lbl {
  font-size: 9.5px; letter-spacing: 0.10em; text-transform: uppercase; color: var(--ink-500);
}
.mast-meta .status .num.alert { color: #B91C1C; }

.break-bar { display: inline-flex; align-items: center; gap: 5px; }
.break-bar .lbl {
  font-size: 9.5px; letter-spacing: 0.10em; text-transform: uppercase; color: var(--ink-500);
  margin-right: 3px;
}
.break-bar .dot {
  width: 19px; height: 19px; border-radius: 50%;
  display: inline-flex; align-items: center; justify-content: center;
  font-weight: 700; font-size: 10.5px; line-height: 1;
  font-variant-numeric: tabular-nums;
  font-feature-settings: "tnum";
}
.break-bar .dot[data-group="1"] { background: #1d1d1f; color: #fff; }
.break-bar .dot[data-group="2"] { background: #6e6e73; color: #fff; }
.break-bar .dot[data-group="3"] { background: #d2d2d7; color: #1d1d1f; }

.mast-context { display: grid; gap: 6px; padding-bottom: 8px; text-align: right; justify-items: end; }
.mast-context .shift {
  font-weight: 500; font-size: 11px; letter-spacing: 0.06em;
  text-transform: uppercase; color: var(--ink-700); line-height: 1;
  white-space: nowrap;
}
.mast-context .week { display: inline-flex; gap: 2px; }
.mast-context .week .d {
  width: 22px; height: 22px; border-radius: 5px;
  display: inline-flex; align-items: center; justify-content: center;
  font-size: 10px; font-weight: 500; color: var(--ink-300);
  font-variant-numeric: tabular-nums;
}
.mast-context .week .d.is-current { background: var(--day-color); color: #fff; font-weight: 700; }
.mast-context .group-key {
  display: inline-flex; gap: 5px; align-items: center;
  font-size: 9px; letter-spacing: 0.10em; text-transform: uppercase; color: var(--ink-500);
}
.mast-context .group-key .gp {
  width: 16px; height: 16px; border-radius: 4px;
  display: inline-flex; align-items: center; justify-content: center;
  font-weight: 700; font-size: 10px; line-height: 1;
}
.gp[data-group="1"], .group-pill[data-group="1"] { background: #1d1d1f; color: #fff; }
.gp[data-group="2"], .group-pill[data-group="2"] { background: #6e6e73; color: #fff; }
.gp[data-group="3"], .group-pill[data-group="3"] { background: #d2d2d7; color: #1d1d1f; }

/* BODY — Phase F: tighter row ratios per handoff spec */
.body { padding: 10px var(--safe) 8px; display: grid;
  grid-template-rows: minmax(0, 1.4fr) minmax(0, 0.85fr) auto auto;
  /* Phase 4g polish: gap 7px → 11px so the ZONES → RESTROOMS → AUXILIARY
     section transitions breathe. Was bumping flush against each other. */
  gap: 11px; min-height: 0; }

.section-label {
  /* Phase 4g polish: bumped section header weight + color for stronger
     hierarchy between zones / restrooms / auxiliary. */
  font-weight: 600; font-size: 10px; letter-spacing: 0.16em;
  text-transform: uppercase; color: var(--ink-700);
  margin: 0 0 8px 0;
  display: flex; align-items: center; gap: 7px;
  white-space: nowrap;
}
.section-label.is-primary {
  font-weight: 600; font-size: 11px; letter-spacing: 0.14em;
  color: var(--ink-700);
  margin-bottom: 8px;
}
.section-label .glyph {
  width: 12px; height: 12px; flex: none;
  color: var(--ink-300);
}
.section-label.is-primary .glyph {
  width: 14px; height: 14px; color: var(--ink-500);
}
.section-label .meta {
  margin-left: auto; font-weight: 400;
  font-size: 9px; letter-spacing: 0.08em; color: var(--ink-300);
  font-variant-numeric: tabular-nums;
}

/* ── ZONE GRID ── */
.body > section:first-of-type { display: grid; grid-template-rows: auto 1fr; min-height: 0; }
.body > section:nth-of-type(2) { display: grid; grid-template-rows: auto 1fr; min-height: 0; }
.zones-grid { display: grid; grid-template-columns: repeat(5,1fr); grid-template-rows: 1fr 1fr; gap: 6px; min-height: 0; }

/* ── ZONE CARD ── */
.zone-card {
  background: #fff; border: 1px solid var(--hairline); border-radius: 6px;
  /* Phase 4g polish: padding-top 0 → 11px so the "ZONE 1" label has clear
     space below the 5px colored stripe (was crushed against it). */
  padding: 11px 10px 8px; display: grid; grid-template-rows: auto auto 1fr;
  gap: 2px; position: relative; overflow: hidden;
}
.zone-card::before { content: ''; position: absolute; inset: 0 0 auto 0; height: 5px; background: var(--card-color); }  /* Phase 4g: 3px→5px */
.zone-meta { display: flex; align-items: center; justify-content: space-between; gap: 6px; margin-top: 0; min-width: 0; }
.zone-meta .zone-num { min-width: 0; overflow: hidden; }
.zone-num {
  font-weight: 700; font-size: 11.5px; letter-spacing: 0.09em;
  text-transform: uppercase; color: var(--card-color);
  white-space: nowrap;
  display: inline-flex; align-items: center;
}
.zone-num .zone-alert {
  display: inline-flex; align-items: center; gap: 3px;
  margin-left: 6px;
  color: #B91C1C;
  font-weight: 600; font-size: 8.5px; letter-spacing: 0.05em;
}
.zone-name {
  font-weight: 700; font-size: 28px; letter-spacing: -0.025em;
  line-height: 1; color: var(--ink-900);
  white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
  max-width: 100%; min-width: 0;
}
.zone-name.is-unfilled { color: var(--ink-300); font-weight: 400; font-style: italic; }
.zone-tasks {
  list-style: none; margin: 1px 0 0; padding: 0;
  /* Phase 4g polish: bumped from 11px → 11.5px and ink-500 → ink-700 for
     reliable print legibility. The deployment book is read at 2am by tired
     supervisors; the original muted color washed out under fluorescent light. */
  font-size: 11.5px; line-height: 1.18; color: var(--ink-700);
  display: flex; flex-direction: column; gap: 0;
}
.zone-tasks li { display: flex; align-items: center; gap: 5px; white-space: nowrap;
  overflow: hidden; text-overflow: ellipsis; }
.zone-tasks li svg { width: 11.5px; height: 11.5px; flex: none; color: var(--ink-300); }
/* Crowded zone cards. Tiered shrink so the task block always fits:
   - is-crowded         (4 tasks)  → name 25px, tasks 10px
   - is-extra-crowded   (5+ tasks, often canonical + sweeper) →
                                    name 22px, tasks 9.5px, smaller icons
   The "shrink-to-fit name" is the key — at 28px the name eats most of the
   card; dropping to 22-25px gives the task list ~6-12px more vertical space.
   (5/3/26 update — Brian flagged that Z10 with sweeper was still clipping.) */
.zone-card.is-crowded .zone-name {
  font-size: 25px; letter-spacing: -0.025em;
}
.zone-card.is-crowded .zone-tasks { font-size: 10px; line-height: 1.1; }
.zone-card.is-crowded .zone-tasks li svg { width: 10.5px; height: 10.5px; }

.zone-card.is-extra-crowded .zone-name {
  font-size: 22px; letter-spacing: -0.025em;
}
.zone-card.is-extra-crowded .zone-tasks { font-size: 9.5px; line-height: 1.08; }
.zone-card.is-extra-crowded .zone-tasks li svg { width: 10px; height: 10px; }
.zone-card.is-extra-crowded li.sweeper-task { padding: 1px 4px; margin: 0; }
.zone-card.is-extra-crowded li.sweeper-task .sweeper-pill {
  font-size: 7.5px; padding: 0px 4px 1px;
}
.zone-tasks--asneeded li {
  font-size: 11.5px; color: var(--ink-300); font-style: italic;
  letter-spacing: 0.01em;
}
.zone-tasks--asneeded li em { font-style: italic; }

.group-pill {
  display: inline-flex; align-items: center; justify-content: center;
  min-width: 16px; height: 16px; padding: 0 3.5px;
  border-radius: 4px; font-weight: 700; font-size: 10px;
  font-variant-numeric: tabular-nums; line-height: 1;
  flex: none;
}

/* AUX */
/* Phase 4g polish: align-items stretch + auto-rows so every aux card grows
   to match the tallest in the row. Was producing a ragged bottom edge when
   Trash 1 had 3 task lines but Support 1 had none. */
.aux-strip { display: grid; grid-template-columns: repeat(6, 1fr);
  grid-auto-rows: 1fr; align-items: stretch; gap: 7px; min-width: 0; }
.aux-card {
  background: #fff; border: 1px solid var(--hairline); border-radius: 6px;
  /* Phase 4g polish: padding-top 6px → 9px so the "TRASH 1" label clears
     the 4px colored stripe with breathing room. */
  padding: 9px 10px 7px;
  /* Make tasks a flex region that absorbs spare vertical space so single-
     line cards no longer collapse to ~44px while neighbors stretch to 70px. */
  display: grid; grid-template-rows: auto auto 1fr; gap: 2px;
  position: relative; overflow: hidden;
  min-height: 44px;
}
.aux-card .aux-tasks { align-self: start; }
.aux-card::before {
  content: ""; position: absolute; inset: 0 0 auto 0; height: 4px;  /* Phase 4g: 2px→4px */
  background: var(--card-color);
}
.aux-meta { display: flex; align-items: center; justify-content: space-between; gap: 6px; margin-top: 2px; min-width: 0; }
.aux-meta .aux-label { min-width: 0; overflow: hidden; text-overflow: ellipsis; }
.aux-label {
  font-weight: 700; font-size: 10.5px; letter-spacing: 0.09em;
  text-transform: uppercase; color: var(--card-color);
  white-space: nowrap;
}
.aux-name { font-weight: 700; font-size: 16px; letter-spacing: -0.02em; color: var(--ink-900); line-height: 1;
  white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 100%; min-width: 0; }
.aux-name.is-unfilled { color: var(--ink-300); font-weight: 400; font-style: italic; }
.aux-tasks { list-style: none; margin: 0; padding: 0;
  /* Phase 4g polish: 10.5px → 11px, ink-500 → ink-700 for print contrast. */
  font-size: 11px; line-height: 1.25; color: var(--ink-700);
  display: flex; flex-direction: column; gap: 1px; }
/* Aux task lines wrap to a 2nd line when long (e.g., Trash 1's "Zones 1-5,
   plus Annex after 5am" was overflowing the narrow card width). */
.aux-tasks li { display: flex; align-items: flex-start; gap: 4px; }
.aux-tasks li svg { width: 11px; height: 11px; color: var(--ink-300); flex: none; margin-top: 1px; }

/* RESTROOMS — unified, tasks merged into header. Cards stretch to a uniform
   height across the row so the 5-card strip reads as a single band even when
   task counts vary (RR 7 has 1 task, RR 8 has 3). */
.rr-grid { display: grid; grid-template-columns: repeat(5, 1fr); gap: 7px; min-height: 0; align-items: stretch; }
.rr-card {
  background: #fff; border: 1px solid var(--hairline); border-radius: 6px;
  /* Phase 4g polish: padding-top 6px → 9px so the "RR 7" label clears the
     4px colored top stripe. Also bumped stripe 2px → 4px for visual parity
     with zone cards. */
  position: relative; overflow: hidden; padding: 9px 10px 7px;
  display: grid; grid-template-rows: auto auto auto; gap: 6px;
  align-content: start;
}
.rr-card::before {
  content: ""; position: absolute; inset: 0 0 auto 0; height: 4px;
  background: var(--card-color);
}
.rr-head {
  display: flex; align-items: center; justify-content: space-between; gap: 6px;
  font-weight: 700; font-size: 11.5px; letter-spacing: 0.09em;
  text-transform: uppercase; color: var(--card-color); line-height: 1;
  margin-top: 2px;
  white-space: nowrap;
}
.rr-head > span:first-child { white-space: nowrap; display: inline-flex; align-items: center; }
.rr-split { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; align-items: start; }
.rr-side { display: grid; gap: 2px; min-width: 0; }
.rr-side .who {
  font-size: 8.5px; font-weight: 500; letter-spacing: 0.06em;
  text-transform: uppercase; color: var(--ink-300); line-height: 1;
  display: inline-flex; align-items: center; gap: 4px;
}
.rr-side .who .group-pill {
  min-width: 12px; height: 12px; font-size: 8.5px; padding: 0 2.5px;
  border-radius: 3px;
}
.rr-side .name {
  font-weight: 700; font-size: 21px; letter-spacing: -0.025em;
  color: var(--ink-900); line-height: 1;
}
.rr-side .name.is-unfilled { color: var(--ink-300); font-weight: 400; font-style: italic; }

.rr-tasks {
  list-style: none; margin: 0; padding: 5px 0 0; border-top: 1px solid var(--hairline);
  /* Phase 4g polish: ink-500 → ink-700 for print contrast. */
  font-size: 11px; line-height: 1.25; color: var(--ink-700);
  display: flex; flex-direction: column; gap: 1px;
}
.rr-tasks li { display: flex; align-items: center; gap: 5px; white-space: nowrap; }
.rr-tasks li svg { width: 11px; height: 11px; flex: none; color: var(--ink-300); }

/* OVERLAPS */
.overlaps-section { display: grid; gap: 6px; }
.overlap-row { display: grid; grid-template-columns: 70px 1fr; gap: 12px; align-items: start; }
.overlap-window {
  font-weight: 600; font-size: 12px; color: var(--ink-900); line-height: 1.2;
  padding-top: 4px;
}
.overlap-window .kind {
  display: block; font-size: 9px; font-weight: 500; letter-spacing: 0.10em;
  text-transform: uppercase; color: var(--ink-500); margin-top: 1px;
}
.overlap-mini-grid { display: grid; grid-template-columns: repeat(6, 1fr); gap: 5px; }
.overlap-mini {
  background: #fff; border: 1px solid var(--hairline); border-radius: 5px;
  padding: 4px 8px; display: grid; gap: 0; min-height: 28px;
}
.overlap-mini .who {
  font-weight: 700; font-size: 13px; letter-spacing: -0.02em;
  color: var(--ink-900); line-height: 1.1;
}
.overlap-mini .who.is-unfilled { color: var(--ink-300); font-weight: 400; font-style: italic; }
.overlap-mini .what {
  font-size: 11px; line-height: 1.25; color: var(--ink-500);
  display: inline-flex; align-items: center; gap: 4px;
}
.overlap-mini .what svg { width: 11px; height: 11px; flex: none; color: var(--ink-300); }

/* PORTERS — collapsed to single line (kept for forward-compat; not currently emitted) */
.porter-line {
  margin-top: 0; padding: 7px 10px;
  border: 1px solid var(--hairline); border-radius: 5px;
  background: var(--ink-100);
  display: flex; align-items: center; gap: 12px; min-height: 32px;
}
.porter-line .label {
  font-weight: 600; font-size: 10px; letter-spacing: 0.14em;
  text-transform: uppercase; color: var(--ink-500);
  display: inline-flex; align-items: center; gap: 6px;
}
.porter-line .label svg { width: 12px; height: 12px; }
.porter-line .names {
  flex: 1; color: var(--ink-700); font-weight: 500; font-size: 11.5px;
}
.porter-line .empty-tag {
  font-weight: 500; font-size: 10.5px; color: var(--ink-300);
  font-style: italic;
  display: inline-flex; align-items: center; gap: 5px;
}
.porter-line .empty-tag svg { width: 11px; height: 11px; color: #aeaeb2; }

/* COVERAGE CALLOUT — subtle chip treatment for legibility at print distance.
   Faint crimson tint background + slightly bolder weight + larger glyph
   keep the tag quiet on the page but lift it enough to register as an alert
   rather than read as decorative type. */
.cover-tag {
  display: inline-flex; align-items: center; gap: 4px;
  margin-left: 6px;
  padding: 1px 5px 1.5px;
  border-radius: 3px;
  background: rgba(185, 28, 28, 0.09);
  color: #B91C1C;
  font-weight: 700; font-size: 9px; letter-spacing: 0.06em;
  text-transform: uppercase; line-height: 1;
  white-space: nowrap;
}
.cover-tag svg { width: 11px; height: 11px; flex: none; }
.zone-num .cover-tag { font-size: 8.5px; padding: 1px 4px 1.5px; }
.zone-num .cover-tag svg { width: 10px; height: 10px; }

/* === BREAK SHEET (page 2) === */
.break-page .mast {
  border-bottom: 1px solid var(--hairline);
}
.break-page .mast::after {
  content:""; position:absolute; left:0; right:0; top:0; bottom:auto; height:6px;
  background: var(--day-color);
}
.break-page .body {
  padding: 10px var(--safe) 10px;
  display: grid; grid-template-rows: minmax(0, 1fr); gap: 10px;
  min-height: 0; overflow: hidden;
}
.break-cols {
  display: grid;
  grid-template-columns: 1fr 1fr 1fr;
  gap: 14px;
  align-items: stretch;
  min-height: 0;
  height: 100%;
}
.break-col {
  display: grid; grid-template-rows: auto 1fr;
  gap: 10px; min-width: 0;
  border-top: 3px solid var(--col-color);
  padding-top: 12px;
  padding-bottom: 12px;
  position: relative;
}
.break-col::after {
  content:"";
  position: absolute; left: 0; right: 0; bottom: 0;
  height: 1px; background: var(--col-color); opacity: 0.45;
}
.break-col .col-head {
  display: grid; grid-template-columns: auto 1fr;
  align-items: end; gap: 12px;
  padding-bottom: 4px;
}
.break-col .col-num {
  font-weight: 800; font-size: 56px; line-height: 0.82;
  letter-spacing: -0.05em; color: var(--col-color);
  font-variant-numeric: tabular-nums;
}
.break-col .col-info {
  display: grid; gap: 3px; padding-bottom: 6px;
}
.break-col .col-label {
  font-size: 11px; letter-spacing: 0.14em; text-transform: uppercase;
  font-weight: 600; color: var(--ink-700); line-height: 1;
}
.break-col .col-meta {
  font-size: 10px; color: var(--ink-500); font-variant-numeric: tabular-nums;
  letter-spacing: 0.04em; line-height: 1;
}
.break-col .col-meta .num { color: var(--ink-900); font-weight: 700; }
.break-col .col-title { display: none; }

.break-rows { display: grid; gap: 2px; align-content: start; }
.break-row {
  display: grid;
  grid-template-columns: 1fr auto;
  gap: 8px; align-items: center;
  padding: 3px 9px 4px;
  border-radius: 5px;
  background: #fff;
  border: 1px solid var(--hairline);
  position: relative; overflow: hidden;
}
.break-row::before {
  content: ""; position: absolute; inset: 0 auto 0 0; width: 3px;
  background: var(--row-color, var(--ink-200));
}
.break-row .who-line {
  display: grid; gap: 1px; min-width: 0;
}
.break-row .name {
  font-weight: 700; font-size: 12.5px; letter-spacing: -0.02em;
  color: var(--ink-900); line-height: 1;
}
.break-row .assign {
  font-size: 9px; line-height: 1.18; color: var(--ink-500);
  letter-spacing: 0;
}
.break-row .slot-badge {
  font-weight: 600; font-size: 9px; letter-spacing: 0.10em;
  text-transform: uppercase; color: var(--row-color, var(--ink-300));
  white-space: nowrap; line-height: 1;
}
.break-section-divider {
  margin: 4px 0 -1px;
  font-size: 8.5px; font-weight: 600; letter-spacing: 0.14em;
  text-transform: uppercase; color: var(--ink-300);
  display: flex; align-items: center; gap: 6px;
}
.break-section-divider::after {
  content: ""; flex: 1; height: 1px; background: var(--hairline);
}
.break-section-divider:first-child { margin-top: 0; }

.break-col[data-group="1"] { --col-color: #1d1d1f; }
.break-col[data-group="2"] { --col-color: #6e6e73; }
.break-col[data-group="3"] { --col-color: #aeaeb2; }

/* COLOR CLASSES — operational zone-family taxonomy */
.c-yellow { --card-color: var(--c-yellow); }
.c-red    { --card-color: var(--c-red); }
.c-pink   { --card-color: var(--c-pink); }
.c-blue   { --card-color: var(--c-blue); }
.c-brown  { --card-color: var(--c-brown); }
.c-green  { --card-color: var(--c-green); }
.c-orange { --card-color: var(--c-orange); }
.c-purple { --card-color: var(--c-purple); }
.c-grey   { --card-color: var(--c-grey); }
.c-teal   { --card-color: var(--c-teal); }

.is-empty { opacity: 0.55; }

/* === SWEEPER TASK STYLING (5/3/26) ===
   Sweeper tasks (Sweeper 5/8/HL, Sweeper 9/10/SR) get a distinct treatment so
   they stand out against the regular task list. Sweeper is a ROUTE (multi-zone)
   not a static task, and operationally critical — Brian wanted it unmistakable
   at floor distance. Pattern: small frosted-orange "SWEEPER" pill before the
   route detail, faint orange tint on the row, slight border-left accent. */
li.sweeper-task {
  display: flex !important; align-items: center;
  gap: 5px;
  background: rgba(251, 140, 0, 0.10);
  border-left: 2px solid var(--c-orange);
  border-radius: 0 3px 3px 0;
  padding: 2px 5px 2px 5px;
  margin: 1px 0;
}
li.sweeper-task .sweeper-pill {
  display: inline-flex; align-items: center; justify-content: center;
  padding: 1px 5px 1.5px;
  border-radius: 3px;
  background: linear-gradient(180deg, #FB8C00 0%, #C56C00 100%);
  color: #fff;
  font-weight: 800; font-size: 8.5px; letter-spacing: 0.10em;
  text-transform: uppercase; line-height: 1;
  white-space: nowrap;
  box-shadow: inset 0 1px 0 rgba(255,255,255,0.25);
  flex: none;
}
li.sweeper-task .sweeper-route {
  font-weight: 700; color: #C56C00;
  font-variant-numeric: tabular-nums;
}

/* === ZONE SHAPE ICONS (5/1/26 accessibility) ===
   Each zone (1-10) has a unique geometric mark; RR cards inherit the shape
   of their corresponding zone. Renders inline before the "Zone N" / "RR N"
   label, tinted card-color. Adds a fourth identifier channel (number, color,
   text, shape) so TMs with reading or color-perception challenges have
   multiple cues to recognize their card. */
.zone-shape {
  width: 14.5px; height: 14.5px; display: inline-block;
  vertical-align: middle; margin-right: 6px;
  color: var(--card-color);
  flex: none;
}
.zone-num .zone-shape  { width: 14.5px; height: 14.5px; margin-right: 6px; }
.rr-head  .zone-shape  { width: 14px; height: 14px; margin-right: 6px; }

/* === CARD WATERMARK (5/1/26) ===
   Giant translucent number behind each zone and restroom card. Primary
   identifier for TMs who struggle with reading — the number is unmistakable
   at distance even when small text is hard to parse. Tinted with the card's
   color so it reinforces the zone/RR color identity at a glance.
   Sits at z-index 0 (behind all content); cards have overflow:hidden so the
   number is clipped by card bounds. */
/* Phase 4j: .card-watermark removed entirely — ghost numerals ghosted on
   B&W copies and muddied card edges. Color stripe + zone-shape SVG icon
   carry zone identity without the visual noise. */
/* Ensure interactive content layers above card background */
.zone-card > * { position: relative; z-index: 1; }
.rr-card   > * { position: relative; z-index: 1; }
.aux-card  > * { position: relative; z-index: 1; }

/* === BOTTOM-STRIP ALERT BANNER (v2.4) ===
   Coverage callout that reads from the floor. Sits flush at the bottom of any
   card whose `.has-alert` modifier is set. Tuned (4/5 update) so multiple
   alerts on a Mon-Thu page don't overwhelm — 14px tall, weight 700, slightly
   warmer crimson with a subtle inset highlight. */
.alert-strip {
  position: absolute;
  left: 0; right: 0; bottom: 0;
  height: 14px;
  background: var(--c-alert);
  color: #fff;
  font-weight: 700; font-size: 9px;
  letter-spacing: 0.09em;
  text-transform: uppercase;
  display: flex; align-items: center; justify-content: center; gap: 4px;
  white-space: nowrap;
  line-height: 1;
  z-index: 2;
  box-shadow: inset 0 1px 0 rgba(255,255,255,0.16);
}
.alert-strip svg { width: 9px; height: 9px; flex: none; color: #fff; }

/* Bottom-strip clearance — content above the strip needs enough padding so
   the strip never overlaps the last visible task. Uniform 18px across all
   card types so the banner sits in the same relative position regardless of
   card content. (5/3/26 — fixes Z5/Z10 banner clip + RR row alignment.) */
.zone-card.has-alert { padding-bottom: 18px; }
.aux-card.has-alert  { padding-bottom: 18px; min-height: 52px; }
.rr-card.has-alert   { padding-bottom: 18px; }
/* Extra cushion on crowded cards so a 5th task or SWEEPER pill can't bleed */
.zone-card.is-extra-crowded.has-alert { padding-bottom: 20px; }

/* Conditional aux card (Support 3) — hidden when empty so quiet nights stay clean */
.aux-card.is-conditional.is-empty { display: none; }
.aux-strip.has-support-3 { grid-template-columns: repeat(7, 1fr); }

/* Z9 SR Buddy — names side by side at the same baseline; BUDDIED moves to meta */
.aux-card.has-buddy { min-height: 52px; }
.aux-card.has-buddy.has-alert { padding-bottom: 17px; min-height: 64px; }
.aux-buddy-row {
  display: grid;
  grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
  align-items: center;
  gap: 6px;
  margin-top: 0;
}
.aux-buddy-row .aux-name {
  display: inline-flex; align-items: center; gap: 4px;
  font-weight: 700; font-size: 14px;
  letter-spacing: -0.02em; color: var(--ink-900);
  line-height: 1;
  white-space: nowrap;
  min-width: 0;
}
.aux-buddy-row .aux-name > :first-child { overflow: hidden; text-overflow: ellipsis; min-width: 0; }
.aux-buddy-row .group-pill {
  min-width: 13px; height: 13px; font-size: 8.5px; padding: 0 2.5px;
  border-radius: 3px; flex: none;
}
/* Generic BUDDIED pill — frosted gold, used in both meta and old buddy slots */
.buddy-pill {
  display: inline-flex; align-items: center; justify-content: center;
  padding: 1px 5px 1.5px;
  border-radius: 3px;
  background: linear-gradient(180deg, #C8A85A 0%, #A88838 100%);
  color: #fff;
  font-weight: 800; font-size: 8px; letter-spacing: 0.10em;
  text-transform: uppercase; line-height: 1;
  white-space: nowrap;
  box-shadow: inset 0 1px 0 rgba(255,255,255,0.25);
}
.aux-meta .meta-buddy-pill { font-size: 8px; padding: 1px 5px 1.5px; }

/* Training pair on a zone card — trainee rendered below trainer, with
   a 'TRAINING D{n}' frosted-blue pill so it reads differently than BUDDIED.
   This is a teacher/learner relationship, not equal partners. The trainer
   name shrinks slightly when paired so the trainee row + the canonical task
   list still fit cleanly inside the fixed-height grid cell. */
.zone-card.has-trainee { padding-bottom: 7px; }
.zone-card.has-trainee.has-alert { padding-bottom: 17px; }
.zone-card.has-trainee .zone-name { font-size: 23px; }
.zone-card.has-trainee .zone-tasks { font-size: 10px; line-height: 1.13; gap: 0; }
.zone-card.has-trainee .zone-tasks li svg { width: 10.5px; height: 10.5px; }
.zone-trainee {
  display: inline-flex; align-items: baseline; gap: 5px;
  margin-top: 1px; line-height: 1;
}
.zone-trainee .train-pill {
  display: inline-flex; align-items: center; justify-content: center;
  padding: 1px 5px 1.5px;
  border-radius: 3px;
  background: linear-gradient(180deg, #4F86C6 0%, #2F5C99 100%);
  color: #fff;
  font-weight: 800; font-size: 8px; letter-spacing: 0.10em;
  text-transform: uppercase; line-height: 1;
  white-space: nowrap;
  box-shadow: inset 0 1px 0 rgba(255,255,255,0.25);
}
.zone-trainee .trainee-name {
  font-weight: 600; font-size: 13px;
  letter-spacing: -0.02em; color: var(--ink-900);
}

/* Trainer-marker-only mode (D4/D5): just the pill, no trainee name */
.zone-trainee--marker-only { gap: 0; }

/* Training pill rendered next to a name inside the RR card (D4 trainee side) */
.rr-train-pill {
  display: inline-flex; align-items: center; justify-content: center;
  margin-left: 5px;
  padding: 1px 5px 1.5px;
  border-radius: 3px;
  background: linear-gradient(180deg, #4F86C6 0%, #2F5C99 100%);
  color: #fff;
  font-weight: 800; font-size: 7.5px; letter-spacing: 0.10em;
  text-transform: uppercase; line-height: 1;
  white-space: nowrap;
  box-shadow: inset 0 1px 0 rgba(255,255,255,0.25);
  vertical-align: middle;
}
.rr-side .name { display: inline-flex; align-items: center; flex-wrap: wrap; gap: 0; }

/* === PERSON-FIRST PAGE (kept for forward-compat) === */
.people-page .body {
  padding: 14px var(--safe) 14px;
  display: grid; grid-template-rows: auto 1fr; gap: 12px;
}
.people-grid {
  display: grid;
  grid-template-columns: repeat(6, 1fr);
  grid-auto-rows: min-content;
  gap: 5px 7px;
  align-content: start;
}
.person-card {
  border: 1px solid var(--hairline); border-radius: 5px;
  padding: 5px 8px 6px; background: #fff;
  display: grid; gap: 1px;
  position: relative; overflow: hidden;
}
.person-card::before {
  content: ""; position: absolute; inset: 0 auto 0 0; width: 3px;
  background: var(--card-color, var(--ink-200));
}
.person-card .name-row {
  display: flex; align-items: center; gap: 5px;
}
.person-card .name {
  font-weight: 600; font-size: 12px; letter-spacing: -0.02em;
  color: var(--ink-900); line-height: 1; flex: 1;
}
.person-card .group-pill { min-width: 14px; height: 14px; font-size: 9px; padding: 0 3px; }
.person-card .assign {
  font-size: 9.5px; color: var(--ink-500);
  letter-spacing: 0; line-height: 1.25;
}
.people-letter-divider {
  grid-column: 1 / -1;
  margin: 6px 0 -2px 0;
  font-weight: 700; font-size: 11px;
  letter-spacing: 0.16em; text-transform: uppercase;
  color: var(--ink-300); padding-bottom: 3px;
  border-bottom: 1px solid var(--hairline);
}
.people-letter-divider:first-child { margin-top: 0; }

/* === PRINT HELPER (screen-only) ===
   Floating affordance shown when viewing the HTML in a browser.
   Hidden in print via @media print { .screen-only { display: none } }.
   Brian can dismiss via the × — preference persists in localStorage and
   the helper collapses to a small "Print" pill in the same corner.
   Shift+Click the pill to bring the full helper back. */
.print-helper {
  position: fixed;
  top: 16px; right: 16px;
  z-index: 100;
  background: #fff;
  border: 1px solid var(--hairline);
  border-radius: 10px;
  padding: 14px 14px 12px 14px;
  box-shadow: 0 8px 24px rgba(0,0,0,0.12);
  display: flex; flex-direction: column; gap: 8px;
  max-width: 280px;
  font-family: var(--font);
  font-size: 11px; color: var(--ink-700); line-height: 1.4;
}
.print-helper .print-helper-close {
  position: absolute;
  top: 4px; right: 6px;
  appearance: none; background: transparent; border: 0;
  width: 20px; height: 20px;
  display: inline-flex; align-items: center; justify-content: center;
  cursor: pointer;
  font-size: 18px; font-weight: 400; line-height: 1;
  color: var(--ink-300);
  border-radius: 4px;
  padding: 0;
  transition: color 0.15s, background 0.15s;
}
.print-helper .print-helper-close:hover {
  color: var(--ink-700);
  background: var(--ink-100);
}
.print-helper .print-btn {
  appearance: none; background: var(--ink-900); color: #fff;
  border: 0; border-radius: 6px;
  padding: 8px 14px; cursor: pointer;
  font-family: var(--font); font-weight: 700; font-size: 12px;
  letter-spacing: 0.02em;
  transition: background 0.15s;
}
.print-helper .print-btn:hover { background: #000; }
.print-helper .print-tips strong { color: var(--ink-900); display: block; margin-bottom: 2px; font-size: 10.5px; letter-spacing: 0.04em; }

/* Mini print button — shown after the helper is dismissed.
   Click prints. Shift+Click reopens the full helper. */
.print-mini {
  position: fixed;
  top: 16px; right: 16px;
  z-index: 100;
  appearance: none;
  background: var(--ink-900); color: #fff;
  border: 0; border-radius: 18px;
  padding: 8px 16px;
  display: none;
  align-items: center; justify-content: center;
  cursor: pointer;
  font-family: var(--font);
  font-weight: 700; font-size: 11px;
  letter-spacing: 0.04em;
  box-shadow: 0 4px 12px rgba(0,0,0,0.18);
  transition: background 0.15s, transform 0.1s;
}
.print-mini:hover { background: #000; }
.print-mini:active { transform: scale(0.96); }

/* ============================================================
   PRINT STYLES — How to print from a browser:
     1. Open the HTML file in Chrome (Safari/Firefox also work but
        Chrome's Save-as-PDF is most reliable)
     2. Cmd+P  (or File → Print)
     3. Destination: Save as PDF (or your printer)
     4. Layout: Landscape  ·  Paper size: Letter (8.5 × 11)
     5. Margins: None (or "Default" — @page rule below sets to 0)
     6. Scale: 100%  /  "Default"  (do NOT use Fit-to-page)
     7. ☑ Background graphics  (REQUIRED — turns on color stripes,
        watermarks, alert banners. Without this it prints monochrome.)
   The result should be exactly 14 pages, 11×8.5 landscape, every page
   self-contained (no spillover).
   ============================================================ */
@media print {
  /* 1. Force every element to honor explicit colors / backgrounds.
        Without this, browsers strip backgrounds in print to save toner. */
  *, *::before, *::after {
    -webkit-print-color-adjust: exact !important;
    print-color-adjust: exact !important;
    color-adjust: exact !important;
  }

  /* 2. Page geometry: Letter landscape, no printer margins.
        @page must come early so browsers apply it before laying out. */
  @page { size: 11in 8.5in landscape; margin: 0; }

  /* 3. Reset the screen-mode body container (the grey backdrop + side
        padding don't apply in print). */
  html, body {
    background: #fff !important;
    padding: 0 !important; margin: 0 !important;
    gap: 0 !important;
  }
  body { display: block !important; }

  /* 4. Each <article class="page"> becomes one physical page.
        page-break-after enforces a new sheet between articles.
        page-break-inside prevents a single article from splitting. */
  .page {
    box-shadow: none !important;
    page-break-after: always;
    break-after: page;
    page-break-inside: avoid;
    break-inside: avoid;
    /* Hard size match — guards against print engines that ignore the
       on-screen height/width if they treat the article as flow content. */
    width: 11in !important;
    height: 8.5in !important;
    margin: 0 !important;
  }
  .page:last-child { page-break-after: avoid; break-after: avoid; }

  /* 5. Break sheet only — clip overflow to the page so a tall column
        doesn't spill onto a 3rd physical page. Deployment page keeps
        its natural sizing (its content is designed to fit). */
  .break-page { max-height: 8.5in; overflow: hidden; }

  /* 6. Hide any on-screen-only chrome (none defined yet, but reserved
        for future "Print this book" buttons or screen-only hints). */
  .screen-only, .print-hidden { display: none !important; }
  :root { --hairline: rgba(0,0,0,0.22); }
}
"""

# Inline SVG sprite — every glyph the daily and break-sheet pages reference.
# Lives in the <body> once, referenced via <use href="#g-..."/> by all pages.
SVG_SPRITE = """<svg width="0" height="0" style="position:absolute" aria-hidden="true">
  <defs>
    <symbol id="g-zones" viewBox="0 0 16 16"><rect x="2" y="2" width="5" height="5" rx="1" fill="none" stroke="currentColor" stroke-width="1.5"/><rect x="9" y="2" width="5" height="5" rx="1" fill="none" stroke="currentColor" stroke-width="1.5"/><rect x="2" y="9" width="5" height="5" rx="1" fill="none" stroke="currentColor" stroke-width="1.5"/><rect x="9" y="9" width="5" height="5" rx="1" fill="none" stroke="currentColor" stroke-width="1.5"/></symbol>
    <symbol id="g-aux" viewBox="0 0 16 16"><circle cx="6" cy="5" r="2.2" fill="none" stroke="currentColor" stroke-width="1.5"/><circle cx="11" cy="6" r="1.8" fill="none" stroke="currentColor" stroke-width="1.5"/><path d="M2.5 13c.6-2.2 2-3.5 3.5-3.5s2.9 1.3 3.5 3.5" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round"/><path d="M9.5 13c.4-1.6 1.4-2.5 2.5-2.5s2.1.9 2.5 2.5" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round"/></symbol>
    <symbol id="g-restroom" viewBox="0 0 16 16"><path d="M8 2c1 1.6 3 3.6 3 6.5a3 3 0 0 1-6 0C5 5.6 7 3.6 8 2z" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linejoin="round"/></symbol>
    <symbol id="g-overlap" viewBox="0 0 16 16"><path d="M4 3v4l-2 2 2 2v4" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/><path d="M12 3v4l2 2-2 2v4" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/><path d="M4 9h8" stroke="currentColor" stroke-width="1.5" stroke-linecap="round"/></symbol>
    <symbol id="g-walk" viewBox="0 0 16 16"><circle cx="9" cy="3" r="1.4" fill="none" stroke="currentColor" stroke-width="1.5"/><path d="M6 14l2-4-1-3 3-1 2 3 2 1" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></symbol>
    <symbol id="g-smoke" viewBox="0 0 16 16"><rect x="2" y="9" width="9" height="2" rx="0.5" fill="none" stroke="currentColor" stroke-width="1.5"/><path d="M12.5 9v2M9 5c0-1 1-1.5 1-2.5s-.7-1.4-.7-1.4M11.5 5c0-1 1-1.5 1-2.5s-.7-1.4-.7-1.4" fill="none" stroke="currentColor" stroke-width="1.2" stroke-linecap="round"/></symbol>
    <symbol id="g-toilet" viewBox="0 0 16 16"><path d="M3 3h10v4H3z" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linejoin="round"/><path d="M4.5 7l1 6h5l1-6" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linejoin="round"/></symbol>
    <symbol id="g-trash" viewBox="0 0 16 16"><path d="M3 4h10M5 4V2.5h6V4M4 4l1 10h6l1-10" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></symbol>
    <symbol id="g-glass" viewBox="0 0 16 16"><rect x="2.5" y="2.5" width="11" height="11" fill="none" stroke="currentColor" stroke-width="1.5"/><path d="M2.5 8h11M8 2.5v11" stroke="currentColor" stroke-width="1.2"/></symbol>
    <symbol id="g-table" viewBox="0 0 16 16"><path d="M2 6h12M3 6v8M13 6v8M5 6V3.5h6V6" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></symbol>
    <symbol id="g-broom" viewBox="0 0 16 16"><path d="M11 2l-7 7M3 13l3-3M3 13l-1 1M9 7l4-4M3 13c2 0 5-1 6-3l1-1" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round"/></symbol>
    <symbol id="g-pit" viewBox="0 0 16 16"><path d="M8 2c1.5 2 4 4 4 7a4 4 0 0 1-8 0c0-3 2.5-5 4-7z" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linejoin="round"/><path d="M8 13v1.5M6.5 14.5h3" stroke="currentColor" stroke-width="1.5" stroke-linecap="round"/></symbol>
    <symbol id="g-elevator" viewBox="0 0 16 16"><rect x="3" y="2" width="10" height="12" rx="0.5" fill="none" stroke="currentColor" stroke-width="1.5"/><path d="M8 2v12M6 6l-1 1.5h2zM10 10l1-1.5h-2z" fill="none" stroke="currentColor" stroke-width="1.2" stroke-linejoin="round"/></symbol>
    <symbol id="g-alert" viewBox="0 0 16 16"><path d="M8 2L1.5 13.5h13z" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linejoin="round"/><path d="M8 6.5v3.5M8 11.5v.5" stroke="currentColor" stroke-width="1.5" stroke-linecap="round"/></symbol>
    <symbol id="g-question" viewBox="0 0 16 16"><circle cx="8" cy="8" r="6" fill="none" stroke="currentColor" stroke-width="1.5"/><path d="M6 6.5c0-1.4 1-2 2-2s2 .6 2 1.8c0 1.6-2 1.6-2 3M8 11.5v.3" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round"/></symbol>
    <symbol id="g-key" viewBox="0 0 16 16"><circle cx="11" cy="5" r="3" fill="none" stroke="currentColor" stroke-width="1.5"/><path d="M9 7L3 13l2 2 1-1-1-1 2-2-1-1 2-2" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linejoin="round" stroke-linecap="round"/></symbol>
    <symbol id="g-link" viewBox="0 0 16 16"><path d="M6.5 9.5l3-3M5 8L3 10a2.5 2.5 0 003.5 3.5L8 12M11 8l2-2A2.5 2.5 0 009.5 2.5L8 4" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round"/></symbol>
    <!-- Zone shape icons (5/1/26 accessibility): one geometric mark per zone, distinct
         enough to read at a glance for low-vision TMs. Solid fills for maximum contrast
         at small sizes. RR cards inherit the shape of their corresponding zone. -->
    <symbol id="sh-1"  viewBox="0 0 24 24"><path d="M12 2 L14.6 9 L22 9 L16 13.4 L18.2 21 L12 16.5 L5.8 21 L8 13.4 L2 9 L9.4 9 Z" fill="currentColor"/></symbol>
    <symbol id="sh-2"  viewBox="0 0 24 24"><path d="M12 2 L22 12 L12 22 L2 12 Z" fill="currentColor"/></symbol>
    <symbol id="sh-3"  viewBox="0 0 24 24"><path d="M12 3 L22 20 L2 20 Z" fill="currentColor"/></symbol>
    <symbol id="sh-4"  viewBox="0 0 24 24"><rect x="3" y="3" width="18" height="18" rx="1.5" fill="currentColor"/></symbol>
    <symbol id="sh-5"  viewBox="0 0 24 24"><path d="M12 2 L21.5 8.5 L17.8 20 L6.2 20 L2.5 8.5 Z" fill="currentColor"/></symbol>
    <symbol id="sh-6"  viewBox="0 0 24 24"><path d="M12 21 C12 21 2.5 14.2 2.5 8.4 A5.4 5.4 0 0 1 12 5.6 A5.4 5.4 0 0 1 21.5 8.4 C21.5 14.2 12 21 12 21 Z" fill="currentColor"/></symbol>
    <symbol id="sh-7"  viewBox="0 0 24 24"><circle cx="12" cy="12" r="9.5" fill="currentColor"/></symbol>
    <symbol id="sh-8"  viewBox="0 0 24 24"><path d="M12 2 L22 8 L22 16 L12 22 L2 16 L2 8 Z" fill="currentColor"/></symbol>
    <symbol id="sh-9"  viewBox="0 0 24 24"><path d="M19 12 A7.5 7.5 0 1 1 12 4.5 A6 6 0 1 0 19 12 Z" fill="currentColor"/></symbol>
    <symbol id="sh-10" viewBox="0 0 24 24"><path d="M9.5 2 H14.5 V9.5 H22 V14.5 H14.5 V22 H9.5 V14.5 H2 V9.5 H9.5 Z" fill="currentColor"/></symbol>
  </defs>
</svg>"""

HTML_SHELL = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Zone Deployment Book — {week_end_short}</title>
<style>{css}</style>
</head>
<body>
{sprite}
<aside class="print-helper screen-only" id="ph-helper">
  <button class="print-helper-close" type="button" onclick="hidePrintHelper()" aria-label="Hide print helper" title="Hide">×</button>
  <button class="print-btn" onclick="window.print()" type="button">Print Book</button>
  <div class="print-tips">
    <strong>Print Settings:</strong>
    Landscape · Letter (8.5×11) · No margins · 100% scale · ☑ Background graphics
  </div>
</aside>
<button class="print-mini screen-only" id="ph-mini" type="button"
        onclick="if(event.shiftKey){{showPrintHelper()}}else{{window.print()}}"
        title="Click to print · Shift+Click to show settings">Print</button>
<script>
(function(){{
  var helper = document.getElementById('ph-helper');
  var mini = document.getElementById('ph-mini');
  var hidden = false;
  try {{ hidden = localStorage.getItem('glcr_print_helper_hidden') === '1'; }} catch(e) {{}}
  if (helper && mini) {{
    helper.style.display = hidden ? 'none' : '';
    mini.style.display   = hidden ? 'inline-flex' : 'none';
  }}
}})();
function hidePrintHelper(){{
  var h = document.getElementById('ph-helper'), m = document.getElementById('ph-mini');
  if (h) h.style.display = 'none';
  if (m) m.style.display = 'inline-flex';
  try {{ localStorage.setItem('glcr_print_helper_hidden','1'); }} catch(e) {{}}
}}
function showPrintHelper(){{
  var h = document.getElementById('ph-helper'), m = document.getElementById('ph-mini');
  if (h) h.style.display = '';
  if (m) m.style.display = 'none';
  try {{ localStorage.removeItem('glcr_print_helper_hidden'); }} catch(e) {{}}
}}
</script>
{pages}
</body>
</html>
"""

# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main(argv):
    if len(argv) < 2:
        print("Usage: render_deployment_book.py <Week Overview - Filled - YYYY-MM-DD.xlsx> [output.html]")
        return 1
    xlsx = Path(argv[1])
    if not xlsx.exists():
        print(f"Input not found: {xlsx}")
        return 1

    days = read_week(xlsx)
    week_end = days[-1]["date"]

    # Phase F closure (5/6/26): all four config sources now read from Supabase
    # via shared/db.py engine helpers. Rules/ folder no longer touched here.
    global TASKS_PM_OL, TASKS_AM_OL, OVERLAP_OVERRIDES
    TASKS_PM_OL, TASKS_AM_OL, OVERLAP_OVERRIDES = load_overlap_tasks()

    # Phase 4k.1: rebuild zone/RR task lists from DB via list_tasks().
    # Hardcoded dicts stay as fallback when DB returns nothing for a slot.
    _load_tasks_from_db()

    males, females, no_sweeper = load_gender_info()

    training_schedule = load_training_config()

    global UTILITY_PORTERS_BY_DATE
    UTILITY_PORTERS_BY_DATE = load_utility_porters()

    # Attach utility porters per-day
    for d in days:
        d["utility_porters"] = UTILITY_PORTERS_BY_DATE.get(d["date"].isoformat(), [])

    # 14-page book: each day emits a Daily Deployment page followed by a
    # Break Sheet page. Page numbers in the footer reflect the 14-page total.
    page_blocks = []
    for idx, d in enumerate(days):
        day_idx = idx + 1
        iso = d["date"].isoformat()
        page_blocks.append(
            render_day_page(d, day_idx, len(days), days, idx,
                            males=males, no_sweeper_tms=no_sweeper,
                            training_pair=training_schedule.get(iso))
        )
        page_blocks.append(
            render_break_sheet_page(d, day_idx, len(days),
                                    males=males, no_sweeper_tms=no_sweeper,
                                    current_idx=idx)  # Phase 4g: pass current_idx for week-dot highlight
        )
    pages_html = "\n".join(page_blocks)

    out_html = HTML_SHELL.format(
        week_end_short=week_end.strftime("%B %-d, %Y"),
        css=CSS,
        sprite=SVG_SPRITE,
        pages=pages_html,
    )

    if len(argv) >= 3:
        out_path = Path(argv[2])
    else:
        out_path = xlsx.parent / f"Zone Deployment Book - {week_end.isoformat()}.html"

    out_path.write_text(out_html, encoding="utf-8")
    print(f"Wrote: {out_path}")
    print(f"  {2 * len(days)} pages ({len(days)} days × 2) · "
          f"{sum(1 for d in days for n in d['zones'] if n)} zone names · "
          f"week ending {week_end}")
    return 0

if __name__ == "__main__":
    sys.exit(main(sys.argv))
