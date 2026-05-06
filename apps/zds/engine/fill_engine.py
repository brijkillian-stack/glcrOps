"""
GLCR Grave Deployment Fill Engine v8
Updated 2026-05-01.

5/1/26 update — overflow routing + new conditional slots:
  - Master template extended: Support 1/2 (renamed from MP 1/2 in headers),
    new Support 3 (col H) + Z9 SR Buddy (col I) conditional columns
  - Engine internal keys still use MP1/MP2 (writes to F/G columns which now
    hold Support 1/2 labels) for backward compatibility
  - New slots: Support3 (col H, all days), Z9SRBuddy (col I, F/Sa overflow)
  - Overflow logic: any unplaced grave-pool TM after AM OL fill goes to:
      F/Sa: 1st → Z9 SR Buddy, 2nd → Support 3
      M-Th: 1st → Support 3
  - Trainees (skill_score ≤ 3) excluded from overflow — they should always
    land in canonical Support 1/2 slots first per training pre-pass

Fill order:
  FRI-SUN:  Admin → All Restrooms → Zone9 → Zone9SR → Zones 1,2,3,4,5,6,7,8,10 → Trash 1,2 → MP
  MON-THU:  Admin → Restrooms (non-1+2) → Zones 1,2,3,4,5,8,9,10 → Restrooms 1+2 → Zones 6,7 → Trash 1,2 → MP

  Admin:    prefer Sheri O, Sherry B, Jamie, Tawnya
  Zone 9/9SR (Fri/Sat): prefer Joy, Melissa, Mike S, Silvia
  Zone 1: prefer Womens-1+2-capable | Zone 2: prefer Mens-1+2-capable
  Z9SR: weekends only | Trash: both day types, always before MP

Rotation: within-week soft preference + 8-week archive cross-week rotation.
Skill matching: soft penalty when TM skill > 2pts below slot difficulty.

Training pre-pass (v7):
  Runs FIRST each day. Overrides all general scheduling rules.
  Config loaded from Rules/Training Config.json — update through conversation.
  TRAINEE_DISPLAY derived dynamically from TM Profiles (score ≤ 3).
  Days 1-6 per GLCR training SOP.

Week config auto-derived from the most recent schedule file — no manual
date edits needed each week.
"""
import shutil, json, warnings, re, sys
from datetime import datetime, date, timedelta
from pathlib import Path
import openpyxl
warnings.filterwarnings("ignore")

# 5/2/26: Multi-objective placement scoring extracted to glcr_engine.scorecard
# (Phase 1 unified-system refactor). fill_engine constructs context + state,
# scorecard.py owns the math. Same module is used by swap.py (Phase 2).
sys.path.insert(0, str(Path(__file__).resolve().parent))
from glcr_engine import scorecard as _sc
from glcr_engine.config import (
    SLOT_CATEGORY as _SLOT_CATEGORY_SHARED,
    ZONE_ADJACENCY as _ZONE_ADJACENCY_SHARED,
)

# 5/5/26: DB helpers — replace Rules/*.json + Eligibility Roster.xlsx reads.
# shared/ lives at repo root (4 parents above this file).
_REPO_ROOT = Path(__file__).resolve().parent.parent.parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
from shared.db import (
    get_engine_roster_from_db,
    get_engine_profiles_from_db,
    get_slot_difficulty as _get_slot_difficulty_db,
    get_slot_load_scores as _get_slot_load_scores_db,
    get_scorecard_config as _get_scorecard_config_db,
    get_overlap_tasks_for_engine,
    get_training_schedule_from_db,
    create_new_tm_stub_in_db,
    get_engine_overrides,
    mark_engine_overrides_applied,
    get_client as _get_db_client,
)

# BASE = the GLCR working directory. Resolved from this script's location so
# the engine runs unchanged in any cowork session. (The script lives at
# GLCR/fill_engine.py; everything else — Rules/, Templates/, Outputs/,
# Inputs/ — is a sibling.)
BASE            = Path(__file__).resolve().parent
# Templates + Archive stay file-based (output artifacts, not rule data).
TEMPLATE_PATH   = BASE / "Templates" / "Zone Deployment Master Template.xlsx"
ARCHIVE_PATH    = BASE / "Archive" / "Grave Placement Archive.xlsx"
# Rules/*.json paths kept for reference but are no longer read by the engine.
# Data now lives in Supabase. Files will be removed after the next verified run.
ROSTER_PATH          = BASE / "Rules" / "Eligibility Roster.xlsx"   # obsolete
PROFILES_PATH        = BASE / "Rules" / "TM Profiles.json"          # obsolete
DIFFICULTY_PATH      = BASE / "Rules" / "Slot Difficulty.json"      # obsolete
TRAINING_CONFIG_PATH = BASE / "Rules" / "Training Config.json"      # obsolete
OVERLAP_TASKS_PATH   = BASE / "Rules" / "Overlap Tasks.json"        # obsolete
SLOT_LOAD_PATH       = BASE / "Rules" / "Slot Load Scores.json"     # obsolete
SCORECARD_WEIGHTS_PATH = BASE / "Rules" / "Scorecard Weights.json"  # obsolete

# ── AUTO-DETECT SCHEDULE FILE ─────────────────────────────────────────
# Default: most recently modified .xlsx in Weekly Schedules.
# Override: pass the filename as a CLI arg, e.g.:
#   python fill_engine.py "Weekly TM EOW 4-30.xlsx"
_sched_dir  = BASE / "Inputs" / "Weekly Schedules"
if len(sys.argv) > 1:
    SCHEDULE_PATH = _sched_dir / sys.argv[1]
    if not SCHEDULE_PATH.exists():
        raise FileNotFoundError(f"Specified schedule not found: {SCHEDULE_PATH}")
else:
    _candidates = sorted(_sched_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not _candidates:
        raise FileNotFoundError(f"No schedule file found in {_sched_dir}")
    SCHEDULE_PATH = _candidates[0]
print(f"  Schedule file: {SCHEDULE_PATH.name}")

# ── AUTO-DERIVE WEEK CONFIG FROM SCHEDULE ────────────────────────────
# Row 7 of Sheet3 contains day names (row 6) and dates (row 7).
# Columns with parseable dates become DATE_COLS.
DAY_ORDER = ["Friday","Saturday","Sunday","Monday","Tuesday","Wednesday","Thursday"]
_swb_cfg = openpyxl.load_workbook(SCHEDULE_PATH, data_only=True, read_only=True)
_ws_cfg  = _swb_cfg["Sheet3"]
_rows    = list(_ws_cfg.iter_rows(min_row=6, max_row=7, values_only=True))
_day_row, _date_row = _rows[0], _rows[1]

DATE_COLS = {}
for _ci, (_dname, _dval) in enumerate(zip(_day_row, _date_row)):
    if not _dval: continue
    try:
        if isinstance(_dval, (date, datetime)):
            _d = _dval.date() if isinstance(_dval, datetime) else _dval
        else:
            _s = str(_dval).strip()
            _m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", _s)
            if not _m: continue
            _d = date(int(_m.group(3)), int(_m.group(1)), int(_m.group(2)))
        DATE_COLS[_d] = _ci
    except: continue
_swb_cfg.close()

if not DATE_COLS:
    raise ValueError("Could not parse any dates from schedule header row 7.")

_all_dates   = sorted(DATE_COLS.keys())
ANCHOR_DATE  = _all_dates[0]
WEEK_ENDING  = str(_all_dates[-1])

# ── WEEK_ID — resolve the DB week record for engine_overrides queries ─────────
# engine_overrides are keyed by week_id (UUID), but the engine only knows
# WEEK_ENDING (date string). One small query bridges them.  If the week isn't
# found (e.g. running stand-alone locally without a DB week record), engine
# overrides are silently skipped — the engine still fills normally.
_WEEK_ID: str = ""
_tmid_to_dn: dict[str, str] = {}   # {tm_id: display_name} for override filtering
try:
    _sb_wk = _get_db_client()
    _wk_row = (
        _sb_wk.table("weeks")
        .select("id")
        .eq("week_ending", WEEK_ENDING)
        .maybe_single()
        .execute()
    )
    _WEEK_ID = ((_wk_row.data or {}).get("id") or "") if _wk_row else ""
    if _WEEK_ID:
        # Build tm_id → display_name lookup for override pool filtering
        _prof_rows = (
            _sb_wk.table("tm_profiles")
            .select("tm_id,display_name")
            .execute()
        ).data or []
        _tmid_to_dn = {r["tm_id"]: r["display_name"] for r in _prof_rows if r.get("tm_id")}
except Exception as _wk_err:
    print(f"  [warn] WEEK_ID lookup failed ({_wk_err}) — engine_overrides will be skipped.")
if _WEEK_ID:
    print(f"  Week ID (overrides): {_WEEK_ID[:8]}…  ({len(_tmid_to_dn)} TM display names indexed)")
else:
    print(f"  Week ID: not found — engine_overrides skipped for this run.")

# Sort days by the order they appear in the schedule (Fri first → Thu last)
DAYS = sorted(DAY_ORDER, key=lambda d: _all_dates.index(
    next((dt for dt in _all_dates if dt.strftime("%A") == d), _all_dates[0]))
    if d in [dt.strftime("%A") for dt in _all_dates] else 99)

RUN_TS      = datetime.now().isoformat()
DAY_DATES   = {d: _all_dates[i] for i, d in enumerate(DAYS)}
DATE_TO_DAY = {v: k for k, v in DAY_DATES.items()}
WEEKEND     = {"Friday","Saturday","Sunday"}
ROTATION_WEEKS = 8

# ── Headcount targets + slot priority for unresolved-severity tagging ────────
# Brian's expected grave headcount per day-type. Anything beyond these counts
# the master template flags as a slot but no body is scheduled to fill it —
# those are *expected* unfilled, not real shortages.
TARGET_GRAVE_BY_DAY = {
    "Friday":   25,
    "Saturday": 25,
    "Sunday":   20,
    "Monday":   18,
    "Tuesday":  18,
    "Wednesday":18,
    "Thursday": 18,
}
# Slots that are auxiliary by design — overflow / surge capacity. If unfilled,
# audit them as severity=low (expected unfilled), not "couldn't fill."
# PMOL/AMOL are included because their fill rate is bounded by pmpool/ampool
# size (often 3–4 TMs against 6 slots) — slots beyond pool size are expected
# to go unfilled and are now logged via the per-slot tracker in the PMOL/AMOL
# fill loop (previously they were silently skipped; see pm_idx watermark fix).
LOW_PRIORITY_SLOTS = {
    "Trash1", "Trash2",
    "MP1", "MP2",
    "Support3",
    "Z9SRBuddy",
    "PMOL1","PMOL2","PMOL3","PMOL4","PMOL5","PMOL6",
    "AMOL1","AMOL2","AMOL3","AMOL4","AMOL5","AMOL6",
}

OUTPUT_DIR  = BASE / "Outputs" / WEEK_ENDING
OUTPUT_WB   = OUTPUT_DIR / f"Week Overview - Filled - {WEEK_ENDING}.xlsx"
AUDIT_JSON  = OUTPUT_DIR / f"Grave Deployment Audit - {WEEK_ENDING}.json"
AUDIT_MD    = OUTPUT_DIR / f"Grave Deployment Audit - {WEEK_ENDING}.md"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── SPECIALIST PREFERENCES ───────────────────────────────────────────
# Admin: prefer these four first — highest difficulty slot, best TMs should own it.
ADMIN_PREFERRED = {"Sheri O", "Sherry B", "Jamie", "Tawnya"}

# Zone 9/9SR: prefer specialists on Fri/Sat only (not Sunday).
ZONE9_WEEKEND_PREFERRED = {"Silvia", "Joy", "Mike S", "Melissa"}
ZONE9_PREF_DAYS = {"Friday", "Saturday"}

# Staffing goals per day type (grave pool bodies on the floor)
STAFFING_GOAL = {
    "Friday": 24, "Saturday": 24,
    "Sunday": 20,
    "Monday": 18, "Tuesday": 18, "Wednesday": 18, "Thursday": 18,

}

# ── TRAINING SCHEDULE (from Supabase public.training_schedule) ───────
# Edit via the webapp or through conversation — never edit this file.
TRAINING_SCHEDULE = {}
_tc = get_training_schedule_from_db()
TRAINING_SCHEDULE = _tc.get("schedule", {})
print(f"  Training config: {len(TRAINING_SCHEDULE)} scheduled day(s) loaded (DB)")


# Zone adjacency — sourced from glcr_engine/config.py (5/3/26 consolidation).
# Auto-symmetrized at config import (#15) so one-way entries can't sneak in.
ZONE_ADJACENCY = _ZONE_ADJACENCY_SHARED

# Restroom slot → corresponding Zone (for Day 4 same-eligibility fallback)
RESTROOM_TO_ZONE = {
    "MRR1":"Zone1",  "WRR1":"Zone1",
    "MRR6":"Zone6",  "WRR6":"Zone6",
    "MRR7":"Zone7",  "WRR7":"Zone7",
    "MRR8":"Zone8",  "WRR8":"Zone8",
    "MRR10":"Zone10","WRR10":"Zone10",
}

# Double-coverage slots: when their paired position can't be filled, the TM
# in this slot is operationally responsible for both. Per the GLCR rules
# (references/rules.md → "Zone coverage responsibility"):
#   Zone 1 ↔ MRR1/WRR1, Zone 2 ↔ MRR1/WRR1
#   Zone N ↔ MRRn/WRRn for N=6,7,8,10
# These slots use skill-priority ranking — strongest available TM wins.
# Z9/Z9SR are excluded; they use specialist pool preference instead.
DOUBLE_COVERAGE_SLOTS = {
    "Zone1", "Zone2", "Zone6", "Zone7", "Zone8", "Zone10",
    "MRR1", "WRR1", "MRR6", "WRR6", "MRR7", "WRR7",
    "MRR8", "WRR8", "MRR10", "WRR10",
}

# Slots that get the hard "no same slot back-to-back" rotation guard.
# Per Brian (4/29/26): a TM cannot be in the same Zone or RR-side on two
# consecutive calendar days. Aux (Z9 SR / Trash / MP / Admin) and Overlaps
# are not in scope — those are role/specialty slots where rotation already
# happens via the soft within_repeat penalty.
BACKTOBACK_SLOTS = {
    "Zone1","Zone2","Zone3","Zone4","Zone5",
    "Zone6","Zone7","Zone8","Zone9","Zone10",
    "MRR1","WRR1","MRR6","WRR6","MRR7","WRR7","MRR8","WRR8","MRR10","WRR10",
}

# Floor area grouping — used to encourage cross-day variety so a TM doesn't
# spend back-to-back nights in the same physical part of the casino.
# Areas reflect the actual floor geography (Brian's confirmed mapping):
#   Z1 + Z2 + RR1+2 = Lobby north
#   Z6 + RR6        = south outdoor
#   Z7 + RR7        = south smoking
#   Z8 + RR8        = table games east
#   Z9 + Z9SR       = high limit slots / smoke room
#   Z10 + RR10      = high limit corner
#   Z3 / Z4 / Z5    = three distinct center-floor areas
#   Trash / MP / Admin / Overlaps — floor-spanning, not area-tracked
SLOT_TO_AREA = {
    "Zone1":"Lobby",  "Zone2":"Lobby",  "MRR1":"Lobby",   "WRR1":"Lobby",
    "Zone3":"C3", "Zone4":"C4", "Zone5":"C5",
    "Zone6":"Z6",   "MRR6":"Z6",   "WRR6":"Z6",
    "Zone7":"Z7",   "MRR7":"Z7",   "WRR7":"Z7",
    "Zone8":"Z8",   "MRR8":"Z8",   "WRR8":"Z8",
    "Zone9":"Z9",   "Zone9SR":"Z9", "Z9SRBuddy":"Z9",  # buddy paired in smoke room
    "Zone10":"Z10", "MRR10":"Z10", "WRR10":"Z10",
}

# Slot code → eligibility roster column name
SLOT_TO_ELIG = {
    "Zone1":"Zone 1","Zone2":"Zone 2","Zone3":"Zone 3","Zone4":"Zone 4","Zone5":"Zone 5",
    "Zone6":"Zone 6","Zone7":"Zone 7","Zone8":"Zone 8","Zone9":"Zone 9","Zone10":"Zone 10",
    "Zone9SR":"Zone 9 SR",
    "MRR1":"Mens 1 + 2","MRR6":"Mens 6","MRR7":"Mens 7","MRR8":"Mens 8","MRR10":"Mens 10",
    "WRR1":"Womens 1 + 2","WRR6":"Womens 6","WRR7":"Womens 7","WRR8":"Womens 8","WRR10":"Womens 10",
    "Trash1":"Trash 1","Trash2":"Trash 2","Admin":"Admin","MP1":"MP 1","MP2":"MP 2",
    # Support 3 + Z9 SR Buddy (5/1/26 — overflow slots; no roster eligibility col)
    "Support3":None, "Z9SRBuddy":None,
}

audit_items = []
placements  = []

print("=" * 62)
print(f"GLCR Grave Deployment v8  |  Week ending {WEEK_ENDING}")
print(f"Run: {RUN_TS[:19]}")
print("=" * 62)

# ── 1. ROSTER ────────────────────────────────────────────────────────
# 5/5/26: reads from Supabase public.tm_profiles + public.tm_eligibility.
print("\n[1/7] Loading Eligibility Roster (DB)...")
roster, fn_lookup = get_engine_roster_from_db()
print(f"  {len(roster)} TMs loaded")

# ── TRAINEE_DISPLAY + profiles_data from Supabase (5/5/26) ──────────
# Load profiles once; reused by 1b, 1c, 1d, and drift-check below.
print("  Loading TM profiles from Supabase...")
profiles_data = get_engine_profiles_from_db()
_profile_trainees = {
    dn for dn, p in profiles_data.get("profiles", {}).items()
    if isinstance(p, dict) and p.get("skill_score", 5) <= 3
}
TRAINEE_DISPLAY = _profile_trainees if _profile_trainees else {"Seth", "Trenidee"}

def match(first, last):
    full = f"{first} {last}".lower().strip()
    if full in roster: return full
    fn = first.lower()
    cands = fn_lookup.get(fn, [])
    if len(cands) == 1: return cands[0]
    if cands:
        ln = last.lower()
        for c in cands:
            if any(ln in p for p in c.split()): return c
        return cands[0]
    return None

def elig(rkey, zone_name):
    return roster.get(rkey, {}).get("eligibility", {}).get(zone_name, False)

def rkey_for_name(display_name):
    """Look up roster key by display name. Whitespace + case-insensitive
    (5/3/26 #3) — prevents silent lookup failures from trailing spaces or
    case mismatches between TM Profiles / Training Config / Roster."""
    if not display_name: return None
    target = str(display_name).strip().lower()
    return next((k for k, v in roster.items()
                 if str(v.get("display_name","")).strip().lower() == target), None)

# ── 1b. TM PROFILES & SLOT DIFFICULTY ───────────────────────────────
tm_skill = {}   # {display_name: skill_score 1-10}
slot_difficulty = {}  # {slot_code: difficulty 1-10}

RESTROOM_PREFERRED = set()   # TMs with slot_preference = "restroom" (physical accommodation)
AVOID_PHYSICAL    = set()    # TMs who should skip Zone9/9SR/Trash
NO_SWEEPER        = set()    # TMs with slot_preference = "no_sweeper" — explicit individual flag (e.g. Daryl)
SWEEPER_CAPABLE   = set()    # Women with slot_preference = "sweeper_capable" — override default no-sweeper
FEMALE_TMS        = set()    # Derived from roster — used for women-no-sweeper default

# Per-TM slot avoidance lists, derived from slot_preference values + gender.
# The fill loop reads SLOT_AVOID_BY_TM[display_name] (a set of slot codes) and
# excludes the TM from those slots. Keeps constraints encoded in data, not code.
SLOT_AVOID_BY_TM = {}

# Sweeper-heavy zones: heavy floor sweeper required.
# Default rule (Brian, 4/28/26): women cannot do sweeper unless explicitly
# marked sweeper_capable. Daryl-style no_sweeper extends to RR 1+2 too.
SWEEPER_SLOTS    = {"Zone7", "Zone8"}
NO_SWEEPER_SLOTS = SWEEPER_SLOTS | {"MRR1", "WRR1"}

# profiles_data already loaded above (see TRAINEE_DISPLAY block).
for dn, p in profiles_data.get("profiles", {}).items():
    if not isinstance(p, dict): continue
    tm_skill[dn] = p.get("skill_score", 5)
    pref = p.get("slot_preference")
    if pref == "restroom":
        RESTROOM_PREFERRED.add(dn)
        AVOID_PHYSICAL.add(dn)
    elif pref == "no_sweeper":
        NO_SWEEPER.add(dn)
        SLOT_AVOID_BY_TM.setdefault(dn, set()).update(NO_SWEEPER_SLOTS)
    elif pref == "sweeper_capable":
        SWEEPER_CAPABLE.add(dn)

# ── 1c. PREFERENCES + PAIR AFFINITIES + ACCOMMODATIONS (5/2/26) ──────
# Per-TM preference, pair_affinity, and accommodation arrays from TM Profiles.
# Empty list if the TM has no entries. Engine reads these at scoring time.
#   - Hard preferences AND accommodations act as candidate-pool filters
#   - Soft preferences nudge the scorecard's preference_fit component
#   - Pair affinities fire when two TMs land in ADJACENT or SAME-AREA slots
TM_PREFERENCES     = {}   # {dn: [{stance, strength, target, ...}, ...]}
TM_PAIR_AFFINITIES = {}   # {dn: [{with, stance, strength, ...}, ...]}
TM_ACCOMMODATIONS  = {}   # {dn: [{type, severity, target, note, status, ...}, ...]}
# profiles_data already loaded above (see TRAINEE_DISPLAY block).
for dn, p in profiles_data.get("profiles", {}).items():
    if not isinstance(p, dict): continue
    prefs = p.get("preferences") or []
    if prefs: TM_PREFERENCES[dn] = prefs
    pairs = p.get("pair_affinities") or []
    if pairs: TM_PAIR_AFFINITIES[dn] = pairs
    accs = p.get("accommodations") or []
    if accs: TM_ACCOMMODATIONS[dn] = accs

# ── 1d. SLOT LOAD SCORES + SCORECARD WEIGHTS ─────────────────────────
SLOT_LOADS = {}
SWEEPER_TAG_BONUS = 2
TRAINING_ROLE_BONUS = {"trainer": 1, "trainee": 1}
try:
    sl_cfg = _get_slot_load_scores_db()
    SLOT_LOADS = sl_cfg.get("loads", {}) or {}
    SWEEPER_TAG_BONUS = sl_cfg.get("sweeper_tag_bonus", 2)
    TRAINING_ROLE_BONUS = sl_cfg.get("training_role_bonus", TRAINING_ROLE_BONUS)
except Exception as _e:
    print(f"  [warn] Slot load scores unavailable ({_e}) — fatigue index disabled")

SCORECARD_WEIGHTS = {
    "skill_match": 1.0, "preference_fit": 1.5, "pair_affinity": 1.0,
    "within_repeat": 1.0, "cross_week_rotation": 0.5, "area_diversity": 0.7,
    "fatigue_index": 0.8, "soft_prefer_set": 0.6,
}
FATIGUE_WINDOW_DAYS = 7
try:
    sw_cfg = _get_scorecard_config_db()
    SCORECARD_WEIGHTS.update(sw_cfg.get("weights", {}) or {})
    FATIGUE_WINDOW_DAYS = sw_cfg.get("fatigue_index_window_days", 7)
except Exception as _e:
    print(f"  [warn] Scorecard config unavailable ({_e}) — using defaults")

# Derive gender from roster (Womens-* eligibility = female). Used to track who
# can do the sweeper TASK (a separate concern from placement). Per Brian
# (4/28/26 latest): women CAN be placed in Z7/Z8 — the sweeper task within
# those zones can be retasked to a male elsewhere on the floor. The placement
# engine treats Z7/Z8 like any other zone; the renderer/supervisor handles
# sweeper task reassignment via per-day override.
for rk, info in roster.items():
    elig_map = info.get("eligibility", {})
    is_female = any(elig_map.get(col) for col in
                    ("Womens 1 + 2", "Womens 6", "Womens 7", "Womens 8", "Womens 10"))
    if is_female:
        FEMALE_TMS.add(info["display_name"])

if RESTROOM_PREFERRED:
    print(f"  Restroom-preferred TMs: {', '.join(sorted(RESTROOM_PREFERRED))}")
if NO_SWEEPER:
    print(f"  No-sweeper TMs (individual flag, e.g. Daryl): {', '.join(sorted(NO_SWEEPER))}")
if SWEEPER_CAPABLE:
    print(f"  Sweeper-capable women: {', '.join(sorted(SWEEPER_CAPABLE))}")
print(f"  Female TMs in roster: {len(FEMALE_TMS)} (Z7/Z8 placement OK; sweeper TASK assigned separately)")

# 5/3/26 #4: Profile vs Roster drift validation. Surfaces silent drift between
# the two TM sources before it bites (separated TMs left in profiles, profile
# entries with no matching roster row, etc.).
_active_roster_dn = {info["display_name"] for info in roster.values()}
# profiles_data loaded from DB above — no file-existence guard needed.
_profile_dn = {dn for dn, p in profiles_data.get("profiles", {}).items()
               if isinstance(p, dict)}
_roster_only = _active_roster_dn - _profile_dn
_profile_only_active = {dn for dn in (_profile_dn - _active_roster_dn)
                        if not (isinstance(profiles_data.get("profiles", {}).get(dn), dict)
                                and profiles_data["profiles"][dn].get("status") in
                                    ("separated", "transferred_to_utility_porter"))}
if _roster_only:
    audit_items.append({"severity":"warning","type":"PROFILE_DRIFT_ROSTER_ONLY",
        "tm_name": None,
        "detail": f"Active roster TMs with no TM Profile entry: {sorted(_roster_only)}. "
                  f"Engine will use defaults (skill_score=5, no preferences/accommodations)."})
if _profile_only_active:
    audit_items.append({"severity":"warning","type":"PROFILE_DRIFT_PROFILE_ONLY",
        "tm_name": None,
        "detail": f"TM Profile entries (non-separated/transferred) with no active roster row: "
                  f"{sorted(_profile_only_active)}. Lookups via rkey_for_name will return None."})
if _roster_only or _profile_only_active:
    print(f"  ⚠ Profile drift detected: {len(_roster_only)} roster-only, "
          f"{len(_profile_only_active)} profile-only. See audit warnings.")

_diff_cfg = _get_slot_difficulty_db()
for slot, d in _diff_cfg.get("slots", {}).items():
    slot_difficulty[slot] = d.get("difficulty", 5)

# Overlap task assignments — slot code → canonical task string.
# Task per slot is stable; people rotate. Used to record the implied task
# on each overlap placement (audit trail) and shared with the deployment
# book renderer so both stay in sync.
overlap_tasks = {}
_ot_data = get_overlap_tasks_for_engine()
for shift_key in ("PM", "AM"):
    for slot_code, task in (_ot_data.get(shift_key) or {}).items():
        overlap_tasks[slot_code] = task
print(f"  Overlap tasks loaded: {len(overlap_tasks)} slot mappings (DB)")

# ── 2. ARCHIVE ───────────────────────────────────────────────────────
print("[2/7] Loading placement archive (rotation history)...")
archive_history = {}
cutoff_date = ANCHOR_DATE - timedelta(weeks=ROTATION_WEEKS)

if ARCHIVE_PATH.exists():
    awb_read = openpyxl.load_workbook(ARCHIVE_PATH, data_only=True)
    if "Placements" in awb_read.sheetnames:
        pws = awb_read["Placements"]
        headers = [str(c.value).strip() if c.value else "" for c in list(pws.iter_rows(min_row=1, max_row=1))[0]]
        col_map = {h: i for i, h in enumerate(headers)}
        for row in pws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            try:
                row_date_raw = row[col_map.get("Date", 1)]
                if isinstance(row_date_raw, datetime): row_date = row_date_raw.date()
                elif isinstance(row_date_raw, date):   row_date = row_date_raw
                else:                                  row_date = date.fromisoformat(str(row_date_raw)[:10])
                if row_date < cutoff_date: continue
                if row_date >= ANCHOR_DATE: continue  # exclude current week
                tm_name   = str(row[col_map.get("TM", 3)]).strip()
                zone_slot = str(row[col_map.get("Zone Slot", 4)]).strip()
                if not tm_name or not zone_slot: continue
                if tm_name not in archive_history: archive_history[tm_name] = {}
                prev = archive_history[tm_name].get(zone_slot)
                if prev is None or row_date > prev:
                    archive_history[tm_name][zone_slot] = row_date
            except Exception as _e:
                # 5/3/26 #13: surface parse failures instead of silently dropping rows.
                # If many fire, archive corruption is likely.
                audit_items.append({"severity":"warning","type":"ARCHIVE_DATE_PARSE_FAIL",
                    "tm_name": str(row[3]) if len(row) > 3 else None,
                    "detail": f"Row date parse failed ({type(_e).__name__}: {_e}); row dropped from archive_history"})
    awb_read.close()
    print(f"  {len(archive_history)} TMs with prior-week history")
else:
    print("  No archive — rotation starts fresh")

# ── 3. SCHEDULE ──────────────────────────────────────────────────────
print("[3/7] Parsing TM Schedule...")
swb = openpyxl.load_workbook(SCHEDULE_PATH, data_only=True)
# DATE_COLS auto-derived from schedule header above — no hardcoding needed.
DATA_ROW = 8

def working(v):
    if v is None: return False
    s = str(v).strip().upper()
    return bool(s) and s not in ("OFF","PTO HOURLY","PTO","MDL","PTO HOURLY MDL") and "PTO" not in s

def is_5am(v): return v is not None and " 5:" in str(v)
def is_1am(v): return v is not None and "1:00A" in str(v)

daily_pools = {d: {"grave":[], "pm_ol":[], "am_ol":[]} for d in DAY_DATES.values()}

def parse(ws, ptype):
    for row in ws.iter_rows(min_row=DATA_ROW, values_only=True):
        if not row or len(row) < 13: continue
        first = str(row[2]).strip() if row[2] else ""
        last  = str(row[3]).strip() if row[3] else ""
        if not first or first in ("First Name","None"): continue
        if "headcount" in first.lower() or "headcount" in last.lower(): continue
        rk = match(first, last)
        if not rk:
            if first not in ("","None"):
                full_name = f"{first} {last}".strip()
                # ── NEW TM AUTO-DETECTION (grave pool only) ────────────
                # Only flag unrostered TMs from the grave pool sheet.
                # PM/AM OL workers are swing/day shift with partial overlap
                # — they don't need roster entries or training configs.
                if ptype == "grave":
                    # Auto-stub into entities + tm_profiles. Idempotent — returns
                    # True if newly inserted, False if TM already existed.
                    # NOTE: even when stub succeeds, the engine's in-memory roster
                    # was loaded once at startup, so the new TM can't be placed
                    # in THIS run. They become eligible next run after Brian
                    # configures their tm_eligibility flags.
                    _stub_threw = False
                    _stub_created = False
                    try:
                        _stub_created = create_new_tm_stub_in_db(full_name, first, WEEK_ENDING)
                        if _stub_created:
                            print(f"  ★ NEW GRAVE TM detected: {full_name} — stub added to DB")
                    except Exception as _e:
                        # Real exception (vs. "already exists" False return).
                        # Previously: silent `pass` swallowed the error completely.
                        _stub_threw = True
                        print(f"  ⚠ NEW GRAVE TM {full_name}: stub failed with {type(_e).__name__}: {_e}")

                    if _stub_threw:
                        audit_items.append({
                            "severity": "error",
                            "type":     "NEW_TM_STUB_FAILED",
                            "tm_name":  full_name,
                            "detail":   "On grave schedule, stub creation threw — see engine stderr. Add to roster manually.",
                        })
                    else:
                        # Either stub_created=True (just inserted) or False
                        # (idempotent no-op because already in tm_profiles).
                        # Either way: TM exists in tm_profiles but needs
                        # eligibility configured before they can be placed.
                        audit_items.append({
                            "severity": "warning",
                            "type":     "NEW_TM_NEEDS_ELIGIBILITY",
                            "tm_name":  full_name,
                            "detail":   (
                                "On grave schedule, stub exists in tm_profiles. "
                                "Configure tm_eligibility flags + training pair "
                                "(if applicable) so they're placeable next run."
                            ),
                        })
                elif ptype != "grave":
                    # OL workers not in roster are expected — suppress warning
                    pass
                else:
                    audit_items.append({"severity":"warning","type":"UNROSTERED_TM_IN_SCHEDULE",
                        "tm_name": full_name, "detail": f"Not placed (pool:{ptype})"})
            continue
        for sdate, ci in DATE_COLS.items():
            cv = row[ci] if ci < len(row) else None
            if not working(cv): continue
            if ptype == "grave":   pk, pd = "grave", sdate
            elif ptype == "pm_ol":
                if not is_1am(cv): continue
                pk, pd = "pm_ol", sdate
            elif ptype == "am_ol":
                if not is_5am(cv): continue
                pk, pd = "am_ol", sdate - timedelta(days=1)
            if pd not in daily_pools: continue
            if rk not in daily_pools[pd][pk]:
                daily_pools[pd][pk].append(rk)

parse(swb["Sheet3"], "grave")
parse(swb["Sheet2"], "pm_ol")
parse(swb["Sheet1"], "am_ol")
for d in daily_pools:
    for pk in ["grave","pm_ol","am_ol"]:
        daily_pools[d][pk].sort(key=lambda k: roster.get(k,{}).get("rank",999))

# ── 4. CELL MAP ──────────────────────────────────────────────────────
print("[4/7] Building cell map...")
twb = openpyxl.load_workbook(TEMPLATE_PATH)
DAY_PREFIX = {"Friday":"dayFriday","Saturday":"daySaturday","Sunday":"daySunday",
              "Monday":"dayMonday","Tuesday":"dayTuesday","Wednesday":"dayWednesday",
              "Thursday":"dayThursday"}

def nr_cell(name):
    dn = twb.defined_names.get(name)
    if not dn: return None
    try:
        dests = list(dn.destinations)
        if not dests: return None
        _, coord = dests[0]
        if not coord: return None
        from openpyxl.utils.cell import coordinate_to_tuple
        return coordinate_to_tuple(coord.replace("$",""))
    except: return None

cell_map = {}
# 5/1/26: Support 3 (col H) and Z9 SR Buddy (col I) are new conditional slots
# added to the master template. The named ranges for them haven't been defined
# yet — engine writes via direct row/col fallback. Map: day → aux row.
AUX_ROW_BY_DAY = {"Friday":23, "Saturday":24, "Sunday":25, "Monday":26,
                  "Tuesday":27, "Wednesday":28, "Thursday":29}
SUPPORT3_COL  = 8   # H
Z9SRBUDDY_COL = 9   # I
BUDDY_DAYS    = {"Friday", "Saturday"}  # F/Sa only — overflow #1 routes to Z9 SR Buddy

for day in DAYS:
    pfx = DAY_PREFIX[day]
    cell_map[day] = {}
    slots = ["Zone1","Zone2","Zone3","Zone4","Zone5","Zone6","Zone7","Zone8","Zone9","Zone10",
             "Zone9SR","MRR1","MRR6","MRR7","MRR8","MRR10","WRR1","WRR6","WRR7","WRR8","WRR10",
             "Trash1","Trash2","Admin","MP1","MP2",
             "PMOL1","PMOL2","PMOL3","PMOL4","PMOL5","PMOL6",
             "AMOL1","AMOL2","AMOL3","AMOL4","AMOL5","AMOL6"]
    for slot in slots:
        nr = ("dayThurday" if day == "Thursday" and "AMOL" in slot else pfx) + slot
        rc = nr_cell(nr)
        if rc: cell_map[day][slot] = rc
    # Direct cell mapping for Support 3 + Z9 SR Buddy (no named range).
    aux_row = AUX_ROW_BY_DAY[day]
    cell_map[day]["Support3"]  = (aux_row, SUPPORT3_COL)
    cell_map[day]["Z9SRBuddy"] = (aux_row, Z9SRBUDDY_COL)

# ── 5. COPY TEMPLATE ─────────────────────────────────────────────────
print("[5/7] Copying template...")
shutil.copy2(TEMPLATE_PATH, OUTPUT_WB)
owb = openpyxl.load_workbook(OUTPUT_WB)
ows = owb["Week Overview"]
anchor = nr_cell("key_wo_d1_date")
if anchor:
    ows.cell(row=anchor[0], column=anchor[1]).value = ANCHOR_DATE

# ── 6. FILL ──────────────────────────────────────────────────────────
print("[6/7] Filling board (v6: training pre-pass + rotation)...")

total_slots  = 0
filled_count = 0
unresolved   = []
week_zone_history = {}  # {display_name: set of zone_slots placed this week}

# Cross-day area diversity tracking. Populated by _record_placement so the
# fill loop knows which floor area each TM was in on each prior day.
tm_areas_by_date = {}   # {iso_date_str: {display_name: area_code}}
# Cross-day slot tracking. Drives the hard "no same slot back-to-back" guard
# in pick() — keyed by date + display name → set of slot codes filled that day.
tm_slots_by_date = {}   # {iso_date_str: {display_name: set(slot, ...)}}
_current_day_iso = None  # set by the daily fill loop before placements

# Adjacent placements per day, populated by write_cell() as fill progresses.
# Passed into _sc.init() so the scorecard can read it during pair-affinity checks.
day_placements = {}  # {day_name: {slot_code: display_name}}

# Initialize the shared scorecard module with all the state it needs.
# Mutable dicts (week_zone_history, archive_history, tm_areas_by_date,
# day_placements) are passed by reference — the engine mutates them as it
# fills, scorecard reads the live state on every score call.
_sc.init(
    roster=roster,
    tm_skill=tm_skill,
    slot_difficulty=slot_difficulty,
    tm_preferences=TM_PREFERENCES,
    tm_accommodations=TM_ACCOMMODATIONS,
    tm_pair_affinities=TM_PAIR_AFFINITIES,
    slot_loads=SLOT_LOADS,
    weights=SCORECARD_WEIGHTS,
    fatigue_window_days=FATIGUE_WINDOW_DAYS,
    slot_to_area=SLOT_TO_AREA,
    zone_adjacency=ZONE_ADJACENCY,
    week_zone_history=week_zone_history,
    archive_history=archive_history,
    tm_areas_by_date=tm_areas_by_date,
    day_placements=day_placements,
    day_dates=DAY_DATES,
    anchor_date=ANCHOR_DATE,
)

# ── SCORECARD (5/2/26 — extracted to glcr_engine.scorecard) ──────────
# All scoring math lives in glcr_engine/scorecard.py and is shared with the
# swap engine (Phase 2). fill_engine constructs context + state, the scorecard
# module owns the math. Module-level aliases below preserve the existing call
# sites without changing behavior.

# === LEGACY INLINE SCORECARD CODE — REMOVED 5/3/26 ===================
# What used to live here (~190 lines of helpers + score_placement + rotation_key)
# has been extracted to glcr_engine/scorecard.py. The block below is the original
# SLOT_CATEGORY map kept for any legacy fill-engine internal that still references
# it directly. Scorecard's own copy lives in scorecard.py.
SLOT_CATEGORY = _SLOT_CATEGORY_SHARED  # 5/3/26 — single source via glcr_engine.config

# === Inline scorecard helpers REMOVED 5/3/26 — see glcr_engine/scorecard.py ===
# fill_engine constructs context + state, scorecard.py owns the math. The
# wrappers below preserve the existing call sites (has_hard_block, score_placement,
# rotation_key) without behavior change.
def _resolve_day_name():
    if not _current_day_iso: return None
    for dn_, d_ in DAY_DATES.items():
        if d_.isoformat() == _current_day_iso: return dn_
    return None

def has_hard_block(dn, slot_code):
    return _sc.has_hard_block(dn, slot_code)

def score_placement(rk, slot_code, *, skill_priority=False, soft_prefer_set=None, day_name=None):
    return _sc.score_placement(rk, slot_code, skill_priority=skill_priority,
                               soft_prefer_set=soft_prefer_set,
                               day_name=day_name or _resolve_day_name())

def rotation_key(rk, zone_slot, skill_priority=False, soft_prefer_set=None, day_name=None):
    return _sc.rotation_key(rk, zone_slot, skill_priority=skill_priority,
                            soft_prefer_set=soft_prefer_set,
                            day_name=day_name or _resolve_day_name())

def write_cell(day, slot, dn):
    global filled_count
    rc = cell_map.get(day, {}).get(slot)
    if rc:
        ows.cell(row=rc[0], column=rc[1]).value = dn
        filled_count += 1
        # Track for pair-affinity adjacency lookups during scoring of subsequent slots
        day_placements.setdefault(day, {})[slot] = dn

def _record_placement(day, slot, dn, pool_type, priority):
    placed_ref = DAY_DATES[day]
    if dn not in week_zone_history: week_zone_history[dn] = set()
    week_zone_history[dn].add(slot)
    skill  = tm_skill.get(dn, 5)
    diff   = slot_difficulty.get(slot, 5)
    stretch = (diff - skill) > 2   # flagged as a skill-stretch assignment
    record = {"date":str(placed_ref),"zone_slot":slot,
              "priority":priority,"tm_display_name":dn,"pool_type":pool_type,
              "tm_skill":skill,"slot_difficulty":diff,"stretch":stretch}
    # If this is an overlap slot and we have a canonical task for it,
    # attach it to the audit record so the audit shows who's doing what.
    if slot in overlap_tasks:
        record["overlap_task"] = overlap_tasks[slot]
    # 5/2/26: Attach scorecard breakdown so the audit shows WHY this TM was chosen.
    rk = rkey_for_name(dn)
    if rk:
        try:
            sc = score_placement(rk, slot, day_name=day)
            record["scorecard"] = {
                "total": round(sc["total"], 3),
                "fatigue_pts": sc["fatigue_pts"],
                "components": {k: round(v, 3) for k, v in sc["components"].items()},
            }
        except Exception:
            pass  # scorecard is enrichment — never block a placement on it
    placements.append(record)

    # Track which area this TM ended up in today, so subsequent days can
    # apply the area-diversity soft penalty in rotation_key.
    area = SLOT_TO_AREA.get(slot)
    if area:
        tm_areas_by_date.setdefault(str(placed_ref), {})[dn] = area

    # Track which slot(s) this TM filled today — used by the back-to-back
    # rotation guard in pick() so the same TM doesn't land in the same slot
    # two consecutive days.
    tm_slots_by_date.setdefault(str(placed_ref), {}).setdefault(dn, set()).add(slot)

def pick(pool_list, placed_today, elig_col, zone_slot, prefer_elig=None, prefer_names=None,
         avoid_names=None, skip_trainees=True, skill_priority=False, soft_prefer_names=None,
         day=None):
    # Yesterday's date-string (used by the back-to-back rotation guard).
    # Only computed when the slot is in BACKTOBACK_SLOTS so we don't pay
    # the date math on every aux/overlap pick.
    yest_iso = None
    if zone_slot in BACKTOBACK_SLOTS and _current_day_iso:
        try:
            yest_iso = (date.fromisoformat(_current_day_iso) - timedelta(days=1)).isoformat()
        except (ValueError, TypeError):
            yest_iso = None

    candidates = []
    fallback_btb = []  # TMs blocked only by the back-to-back guard — kept as a
                       # last-resort pool if the hard rule leaves the slot empty.
    fallback_hardpref = []  # TMs blocked only by a hard-avoid preference
    for rk in pool_list:
        dn = roster[rk]["display_name"]
        if dn in placed_today: continue
        if skip_trainees and dn in TRAINEE_DISPLAY: continue
        if avoid_names and dn in avoid_names: continue   # call-site physical restriction
        # Per-TM slot avoidance from profile slot_preference (Daryl no_sweeper, women no_sweeper, etc.)
        if zone_slot in SLOT_AVOID_BY_TM.get(dn, ()): continue
        if not elig(rk, elig_col): continue
        # 5/2/26: Hard preferences (avoid:hard) filter from the candidate pool.
        # If the slot is must-fill and the pool empties, we'll override below
        # and audit-log it as HARD_PREFERENCE_OVERRIDDEN.
        hard_block = has_hard_block(dn, zone_slot)
        if hard_block:
            fallback_hardpref.append((rk, hard_block))
            continue
        # Hard back-to-back guard: same Zone or RR-side as yesterday is not
        # allowed. If this excludes everyone we'll fall back below.
        if yest_iso and zone_slot in tm_slots_by_date.get(yest_iso, {}).get(dn, ()):
            fallback_btb.append(rk)
            continue
        candidates.append(rk)
    # Fallback chain: prefer back-to-back violation over hard-preference override
    # (least bad first). Hard-preference override is logged as a warning.
    if not candidates and fallback_btb:
        candidates = fallback_btb
    if not candidates and fallback_hardpref:
        # Override the hard preference — audit it loudly so the supervisor sees.
        candidates = [rk for rk, _ in fallback_hardpref]
        for rk, hard_block in fallback_hardpref:
            audit_items.append({"severity":"warning","type":"HARD_PREFERENCE_OVERRIDDEN",
                "tm_name":roster[rk]["display_name"],"zone_slot":zone_slot,
                "detail":f"Hard avoid preference overridden — no other eligible candidates. "
                         f"Preference note: {hard_block.get('note','(no note)')}"})
    if not candidates: return None
    # Sort using the multi-objective scorecard.
    candidates.sort(key=lambda rk: rotation_key(rk, zone_slot, skill_priority,
                                                soft_prefer_names, day_name=day))
    # prefer_names: HARD pick from this set first (e.g. Z9 SR weekend specialists)
    if prefer_names:
        named = [rk for rk in candidates if roster[rk]["display_name"] in prefer_names]
        if named: return named[0]
    if prefer_elig:
        preferred = [rk for rk in candidates if elig(rk, prefer_elig)]
        if preferred: return preferred[0]
    return candidates[0]

def place(day, slot, elig_col, pool_list, placed_today,
          pool_type="grave", priority="auto", prefer_elig=None, prefer_names=None,
          avoid_names=None, skip_trainees=True, skill_priority=False, soft_prefer_names=None):
    if slot not in cell_map.get(day, {}): return None
    if slot in filled_slots: return None          # already claimed by training pre-pass
    rk = pick(pool_list, placed_today, elig_col, slot, prefer_elig, prefer_names,
               avoid_names, skip_trainees, skill_priority, soft_prefer_names, day=day)
    if rk:
        dn = roster[rk]["display_name"]
        write_cell(day, slot, dn)
        placed_today.add(dn)
        filled_slots.add(slot)
        _record_placement(day, slot, dn, pool_type, priority)
        return dn
    unresolved.append({"date":str(DAY_DATES[day]),"zone_slot":slot,"priority":priority})
    return None

# ── TRAINING PRE-PASS ────────────────────────────────────────────────
def place_training(day, slot, display_name, placed_today, tag):
    """Force-place a TM into a slot (training override). Bypasses normal pick logic."""
    if slot not in cell_map.get(day, {}):
        audit_items.append({"severity":"warning","type":"TRAINING_SLOT_NOT_IN_TEMPLATE",
            "tm_name":display_name,"detail":f"{slot} not in cell map for {day}"})
        return False
    if display_name in placed_today:
        audit_items.append({"severity":"warning","type":"TRAINING_CONFLICT",
            "tm_name":display_name,"detail":f"Already placed on {day}, cannot force to {slot}"})
        return False
    # 5/3/26 #1: Eligibility guard. Training pre-pass force-places by design,
    # but should never put a TM in a slot they're not Y for in the roster.
    # If they're ineligible, log a warning and skip the placement.
    elig_col = SLOT_TO_ELIG.get(slot)
    rk_for_check = rkey_for_name(display_name)
    if rk_for_check and elig_col and not elig(rk_for_check, elig_col):
        audit_items.append({"severity":"warning","type":"TRAINING_TM_INELIGIBLE",
            "tm_name":display_name,
            "detail":f"{tag}: {display_name} not eligible for {slot} (roster: {elig_col}=N). Skipping."})
        return False
    write_cell(day, slot, display_name)
    placed_today.add(display_name)
    filled_slots.add(slot)                        # lock slot against subsequent grave fill
    _record_placement(day, slot, display_name, "training", tag)
    return True

def training_prepass(day, placed):
    d      = DAY_DATES[day]
    config = TRAINING_SCHEDULE.get(str(d))
    if not config: return

    trainee_name = config["trainee"]
    trainer_name = config["trainer"]
    training_day = config["day"]
    is_wkend     = day in WEEKEND

    trainee_rk = rkey_for_name(trainee_name)
    trainer_rk = rkey_for_name(trainer_name)

    if not trainee_rk:
        audit_items.append({"severity":"error","type":"TRAINING_TM_NOT_FOUND",
            "tm_name":trainee_name,"detail":"Trainee not in roster — skipping training pre-pass"})
        return
    if not trainer_rk:
        audit_items.append({"severity":"error","type":"TRAINING_TM_NOT_FOUND",
            "tm_name":trainer_name,"detail":"Trainer not in roster — skipping training pre-pass"})
        return

    print(f"    ★ Training Day {training_day}: {trainee_name} (trainee) ↔ {trainer_name} (trainer)")

    if training_day == 1:
        # Day 1: Restroom Training — Restrooms 1+2
        trainee_mens = elig(trainee_rk, "Mens 1 + 2")
        trainer_mens = elig(trainer_rk, "Mens 1 + 2")
        opposite = trainee_mens != trainer_mens

        if opposite:
            # Each in their respective Restrooms 1+2
            t_slot  = "MRR1" if trainee_mens else "WRR1"
            tr_slot = "WRR1" if trainee_mens else "MRR1"
            place_training(day, t_slot,  trainee_name, placed, "Training-D1-Trainee")
            place_training(day, tr_slot, trainer_name, placed, "Training-D1-Trainer")
        else:
            # Same eligibility: Trainer → Restroom 1+2, Trainee → MP
            tr_slot = "MRR1" if trainer_mens else "WRR1"
            place_training(day, tr_slot, trainer_name, placed, "Training-D1-Trainer")
            for mp in ["MP1","MP2"]:
                if place_training(day, mp, trainee_name, placed, "Training-D1-Trainee"): break

    elif training_day == 2:
        # Day 2: Zone Training — Trainer in Zone 1/2/3/4, Trainee as MP
        placed_trainer = False
        for z_slot in ["Zone1","Zone2","Zone3","Zone4"]:
            if elig(trainer_rk, SLOT_TO_ELIG[z_slot]):
                if place_training(day, z_slot, trainer_name, placed, "Training-D2-Trainer"):
                    placed_trainer = True; break
        if placed_trainer:
            for mp in ["MP1","MP2"]:
                if place_training(day, mp, trainee_name, placed, "Training-D2-Trainee"): break
        else:
            audit_items.append({"severity":"warning","type":"TRAINING_NO_ELIGIBLE_ZONE",
                "tm_name":trainer_name,"detail":"Day 2: Trainer not eligible for Zone 1-4"})

    elif training_day == 3:
        # Day 3: Smoking Room / Zone 9
        if is_wkend:
            # Weekend: Trainee → Zone9SR, Trainer → Zone9
            place_training(day, "Zone9SR", trainee_name, placed, "Training-D3-Trainee")
            place_training(day, "Zone9",   trainer_name, placed, "Training-D3-Trainer")
        else:
            # Weekday: Trainer → Zone9, Trainee → MP
            place_training(day, "Zone9", trainer_name, placed, "Training-D3-Trainer")
            for mp in ["MP1","MP2"]:
                if place_training(day, mp, trainee_name, placed, "Training-D3-Trainee"): break

    elif training_day == 4:
        # Day 4: Shadow Restroom
        all_rr_slots = ["MRR1","WRR1","MRR6","WRR6","MRR7","WRR7","MRR8","WRR8","MRR10","WRR10"]
        # Weekend only has MRR1/WRR1 available; otherwise skip slot-1 on weekdays
        available_rr = [s for s in all_rr_slots if s in cell_map.get(day, {})]

        trainee_rr_slot = None
        for rs in available_rr:
            elig_col = SLOT_TO_ELIG[rs]
            if elig(trainee_rk, elig_col):
                if place_training(day, rs, trainee_name, placed, "Training-D4-Trainee"):
                    trainee_rr_slot = rs; break

        if trainee_rr_slot:
            t_gender   = "M" if trainee_rr_slot.startswith("M") else "W"
            opp_gender = "W" if t_gender == "M" else "M"
            rr_num     = trainee_rr_slot[3:]  # "1","6","7","8","10"
            opp_slot   = opp_gender + "RR" + rr_num

            opp_elig_col = SLOT_TO_ELIG.get(opp_slot)
            if opp_elig_col and elig(trainer_rk, opp_elig_col) and opp_slot in cell_map.get(day, {}):
                # Opposite eligible: Trainer in opposite-gender same-zone restroom
                place_training(day, opp_slot, trainer_name, placed, "Training-D4-Trainer")
            else:
                # Same eligibility fallback: Trainer in corresponding Zone
                zone_fallback = RESTROOM_TO_ZONE.get(trainee_rr_slot)
                if zone_fallback and zone_fallback in cell_map.get(day, {}):
                    if elig(trainer_rk, SLOT_TO_ELIG.get(zone_fallback, "")):
                        place_training(day, zone_fallback, trainer_name, placed, "Training-D4-Trainer")
                    else:
                        audit_items.append({"severity":"warning","type":"TRAINING_D4_TRAINER_UNPLACEABLE",
                            "tm_name":trainer_name,
                            "detail":f"Not eligible for {zone_fallback} (Day 4 fallback)"})
                else:
                    audit_items.append({"severity":"warning","type":"TRAINING_D4_NO_FALLBACK",
                        "tm_name":trainer_name,"detail":f"No zone fallback for {trainee_rr_slot}"})
        else:
            audit_items.append({"severity":"warning","type":"TRAINING_D4_NO_RESTROOM",
                "tm_name":trainee_name,"detail":"No eligible restroom found for Day 4"})

    elif training_day == 5:
        # Day 5: Shadow Zone — Trainee in Zone 1/2/3/4/6/7, Trainer adjacent
        trainee_z_slot = None
        for z_slot in ["Zone1","Zone2","Zone3","Zone4","Zone6","Zone7"]:
            if elig(trainee_rk, SLOT_TO_ELIG[z_slot]) and z_slot in cell_map.get(day, {}):
                if place_training(day, z_slot, trainee_name, placed, "Training-D5-Trainee"):
                    trainee_z_slot = z_slot; break

        if trainee_z_slot:
            adjacent = ZONE_ADJACENCY.get(trainee_z_slot, [])
            placed_trainer = False
            for adj in adjacent:
                if adj in cell_map.get(day, {}) and elig(trainer_rk, SLOT_TO_ELIG[adj]):
                    if place_training(day, adj, trainer_name, placed, "Training-D5-Trainer"):
                        placed_trainer = True; break
            if not placed_trainer:
                audit_items.append({"severity":"warning","type":"TRAINING_D5_NO_ADJACENT",
                    "tm_name":trainer_name,
                    "detail":f"No adjacent zone available to {trainee_z_slot} for Trainer"})
        else:
            audit_items.append({"severity":"warning","type":"TRAINING_D5_NO_ZONE",
                "tm_name":trainee_name,"detail":"No eligible Zone 1-4/6-7 for Trainee (Day 5)"})

    elif training_day == 6:
        # Day 6: ADP Courses — Trainee as MP, Trainer in Zone 1/2/3/4/6/7
        for mp in ["MP1","MP2"]:
            if place_training(day, mp, trainee_name, placed, "Training-D6-Trainee"): break
        for z_slot in ["Zone1","Zone2","Zone3","Zone4","Zone6","Zone7"]:
            if elig(trainer_rk, SLOT_TO_ELIG[z_slot]) and z_slot in cell_map.get(day, {}):
                if place_training(day, z_slot, trainer_name, placed, "Training-D6-Trainer"): break

# ── Engine override accumulator ───────────────────────────────────────────────
# Collects IDs of engine_overrides rows consumed this run so applied_count
# can be incremented after the fill completes.
_applied_override_ids: list[str] = []

# ── MAIN FILL LOOP ───────────────────────────────────────────────────
for day in DAYS:
    d      = DAY_DATES[day]
    gpool  = list(daily_pools[d]["grave"])
    pmpool = list(daily_pools[d]["pm_ol"])
    ampool = list(daily_pools[d]["am_ol"])
    placed = set()
    filled_slots = set()
    is_weekend = day in WEEKEND

    # ── Phase D: ENGINE OVERRIDES — hard-filter unavailable TMs ──────────────
    # Load supervisor knowledge for this night.  If WEEK_ID is unknown (e.g.
    # stand-alone local run), this block is a no-op.
    # K.4 (5/6/26): soft overrides now bias scoring. Each override_type
    # collected per-TM gets passed to scorecard.set_soft_overrides_for_day()
    # so score_placement() biases per-slot picks. Hard 'unavailable' is
    # still filtered from the pool entirely (does not reach scoring).
    _soft_overrides_by_dn: "dict[str, set[str]]" = {}
    if _WEEK_ID:
        try:
            _day_overrides = get_engine_overrides(_WEEK_ID, d)
        except Exception as _ov_err:
            _day_overrides = {}
            print(f"  [warn] engine_overrides fetch failed for {d}: {_ov_err}")

        if _day_overrides:
            _unavail_dns: set[str] = set()
            for _tm_id_ov, _ov_list in _day_overrides.items():
                for _ov in _ov_list:
                    _ov_type = _ov.get("override_type", "")
                    _ov_dn   = _tmid_to_dn.get(_tm_id_ov, "")
                    if _ov_type == "unavailable":
                        if _ov_dn:
                            _unavail_dns.add(_ov_dn)
                        _applied_override_ids.append(_ov["id"])
                        audit_items.append({
                            "severity": "info",
                            "type":     "OVERRIDE_UNAVAILABLE",
                            "tm_name":  _ov_dn or _tm_id_ov,
                            "detail":   (
                                f"Hard-filtered from {d} pools via engine_override "
                                f"(reason: {(_ov.get('payload') or {}).get('reason', 'n/a')})"
                            ),
                        })
                    # K.4 — soft overrides bias placement via scorecard.
                    # prefer_easier / avoid_high_load / priority_placement
                    # have actual scoring components in scorecard.py;
                    # training_pair / skip_rotation / special_context are
                    # tracked + audited but not yet wired to a scoring rule.
                    elif _ov_type in ("prefer_easier", "avoid_high_load", "priority_placement",
                                      "training_pair", "skip_rotation", "special_context"):
                        if _ov_dn:
                            _soft_overrides_by_dn.setdefault(_ov_dn, set()).add(_ov_type)
                        _applied_override_ids.append(_ov["id"])
                        wired_types = {"prefer_easier", "avoid_high_load", "priority_placement"}
                        audit_items.append({
                            "severity": "info",
                            "type":     f"OVERRIDE_{_ov_type.upper()}",
                            "tm_name":  _ov_dn or _tm_id_ov,
                            "detail":   (
                                f"Soft override applied to scoring for {d}"
                                if _ov_type in wired_types else
                                f"Soft override noted for {d} (scoring integration pending)"
                            ),
                        })

            # Apply hard filter: remove unavailable TMs from all pools for this night.
            # NOTE: gpool/pmpool/ampool contain roster keys (rk), but
            # `_unavail_dns` is a set of display names — the original filter
            # compared rk against dn and never matched, so called-off TMs were
            # silently still placed. Resolve each pool member's display name
            # via the roster map and compare on that.
            if _unavail_dns:
                def _is_unavail(rk: str) -> bool:
                    return roster.get(rk, {}).get("display_name") in _unavail_dns
                removed_g  = [rk for rk in gpool  if _is_unavail(rk)]
                removed_pm = [rk for rk in pmpool if _is_unavail(rk)]
                removed_am = [rk for rk in ampool if _is_unavail(rk)]
                gpool  = [rk for rk in gpool  if not _is_unavail(rk)]
                pmpool = [rk for rk in pmpool if not _is_unavail(rk)]
                ampool = [rk for rk in ampool if not _is_unavail(rk)]
                all_removed_dns = [
                    roster.get(rk, {}).get("display_name", rk)
                    for rk in removed_g + removed_pm + removed_am
                ]
                if all_removed_dns:
                    print(f"  [{day}] engine_overrides: removed {all_removed_dns} from pools (unavailable)")

    # K.4 — refresh scorecard's per-night soft override map even if empty.
    # Empty dict resets the previous night's overrides so they don't leak
    # into subsequent days. Sits OUTSIDE the WEEK_ID guard so a stand-alone
    # local run (no WEEK_ID) still resets cleanly between days.
    try:
        _sc.set_soft_overrides_for_day(_soft_overrides_by_dn)
    except AttributeError:
        # Defensive: older scorecard builds without the helper.
        pass

    # Tell rotation_key what day it's filling so the area-diversity penalty
    # can compare against tm_areas_by_date[yesterday].
    _current_day_iso = str(d)
    _sc.set_current_day(_current_day_iso)

    day_slots = list(cell_map.get(day, {}).keys())
    total_slots += len(day_slots)

    # ── TRAINING PRE-PASS (runs first, overrides everything) ─────────
    training_prepass(day, placed)

    # ─────────────────────────────────────────────────────────────────
    # NEW UNIFORM FILL ORDER (no weekday/weekend split)
    # Brian's spec (4/28/2026):
    #   1. Restrooms (most-constrained pool — gender-locked)
    #   2. Admin    (specialty role — soft-prefer rockstars; rotation lets others in)
    #   3. Z9 SR    (specialist Fri/Sat only)
    #   4. Zones    (skip-priority order — Z9 last so it's first to skip when short)
    #   5. Trash    (overflow)
    #   6. MP       (more overflow / trainees)
    # ─────────────────────────────────────────────────────────────────

    # ── 1. RESTROOMS — all 5 pairs, every day ─────────────────────────
    # All RR slots are double-coverage candidates — skill_priority=True so
    # the strongest available TM gets the slot (handles potential zone cover).
    # RESTROOM_PREFERRED still hard-picks TMs with physical accommodation.
    rr_pref = RESTROOM_PREFERRED if RESTROOM_PREFERRED else None
    elig_mrr = {"MRR1":"Mens 1 + 2","MRR6":"Mens 6","MRR7":"Mens 7","MRR8":"Mens 8","MRR10":"Mens 10"}
    elig_wrr = {"WRR1":"Womens 1 + 2","WRR6":"Womens 6","WRR7":"Womens 7","WRR8":"Womens 8","WRR10":"Womens 10"}
    for s in ("MRR1","MRR6","MRR7","MRR8","MRR10"):
        place(day, s, elig_mrr[s], gpool, placed, priority="Restroom",
              prefer_names=rr_pref, skill_priority=(s in DOUBLE_COVERAGE_SLOTS))
    for s in ("WRR1","WRR6","WRR7","WRR8","WRR10"):
        place(day, s, elig_wrr[s], gpool, placed, priority="Restroom",
              prefer_names=rr_pref, skill_priority=(s in DOUBLE_COVERAGE_SLOTS))

    # ── 2. ADMIN — soft preference for Sheri O / Sherry B / Jamie / Tawnya ──
    # Rockstars favored at the tiebreaker level; Gary / Sam / Kaylee / Cookie
    # naturally rotate in once rockstars have done Admin earlier in the week.
    place(day, "Admin", "Admin", gpool, placed, priority="Admin",
          soft_prefer_names=ADMIN_PREFERRED)

    # ── 3. Z9 SMOKING ROOM — specialist preference Fri/Sat only ──────
    # Joy, Melissa, Mike S, Silvia get hard preference Fri/Sat; other days
    # everyone eligible should rotate through.
    z9sr_pref = ZONE9_WEEKEND_PREFERRED if day in ZONE9_PREF_DAYS else None
    place(day, "Zone9SR", "Zone 9 SR", gpool, placed, priority="Zone-SR",
          prefer_names=z9sr_pref, avoid_names=AVOID_PHYSICAL)

    # ── 4. ZONES — skip-priority order ───────────────────────────────
    # Order produces Brian's intended skip priority on short-staffed nights:
    #   Skip 1st: Z9    Skip 4th: Z2
    #   Skip 2nd: Z6    Skip 5th: Z3
    #   Skip 3rd: Z7    Skip 6th: Z10  (most-protected skippable)
    # Must-fill: Z1, Z4, Z5, Z8 (positions 1-4)
    # Updated 4/29/26 per Brian: Z10 protected last; Z3 drops before Z10
    # since Z5 covers Z3 directly. New full order: Z9 → Z6 → Z7 → Z2 → Z3 → Z10.
    zone_order = [
        ("Zone1",  "Zone 1",  "Womens 1 + 2"),  # must-fill
        ("Zone4",  "Zone 4",  None),            # must-fill
        ("Zone5",  "Zone 5",  None),            # must-fill
        ("Zone8",  "Zone 8",  None),            # must-fill
        ("Zone10", "Zone 10", None),            # skip 6th (most protected)
        ("Zone3",  "Zone 3",  None),            # skip 5th
        ("Zone2",  "Zone 2",  "Mens 1 + 2"),    # skip 4th
        ("Zone7",  "Zone 7",  None),            # skip 3rd
        ("Zone6",  "Zone 6",  None),            # skip 2nd
        ("Zone9",  "Zone 9",  None),            # skip 1st (no specialist pref)
    ]
    for slot, ec, pref in zone_order:
        avoid = AVOID_PHYSICAL if slot == "Zone9" else None
        place(day, slot, ec, gpool, placed, priority="Zone",
              prefer_elig=pref, avoid_names=avoid,
              skill_priority=(slot in DOUBLE_COVERAGE_SLOTS))

    # Audit: warn if BOTH Zone 9 and Z9 SR are unstaffed (smoke room
    # area is then operationally uncovered — manual escalation needed).
    z9_filled    = "Zone9"   in filled_slots
    z9sr_filled  = "Zone9SR" in filled_slots
    if not z9_filled and not z9sr_filled and "Zone9" in cell_map.get(day, {}):
        audit_items.append({"severity":"critical","type":"Z9_AREA_UNCOVERED",
            "tm_name":None,"detail":f"{day} {d}: Zone 9 AND Z9 SR both unstaffed — smoke room area uncovered"})

    # ── TRASH (both day types, before MP) ────────────────────────────
    place(day, "Trash1", "Trash 1", gpool, placed, priority="Trash", avoid_names=AVOID_PHYSICAL)
    place(day, "Trash2", "Trash 2", gpool, placed, priority="Trash", avoid_names=AVOID_PHYSICAL)

    # ── MP (trainees first) ──────────────────────────────────────────
    for mp_slot, mp_elig_col in [("MP1","MP 1"),("MP2","MP 2")]:
        if mp_slot not in cell_map.get(day, {}): continue
        trainee_cands = sorted(
            [tk for tk in gpool
             if roster[tk]["display_name"] in TRAINEE_DISPLAY
             and roster[tk]["display_name"] not in placed
             and elig(tk, mp_elig_col)],
            key=lambda rk: rotation_key(rk, mp_slot))
        rk = trainee_cands[0] if trainee_cands else None
        if not rk:
            others = sorted(
                [tk for tk in gpool
                 if roster[tk]["display_name"] not in placed
                 and roster[tk]["display_name"] not in TRAINEE_DISPLAY
                 and elig(tk, mp_elig_col)],
                key=lambda rk: rotation_key(rk, mp_slot))
            rk = others[0] if others else None
        if rk:
            dn2 = roster[rk]["display_name"]
            write_cell(day, mp_slot, dn2)
            placed.add(dn2)
            _record_placement(day, mp_slot, dn2, "grave", "MP")
        else:
            unresolved.append({"date":str(d),"zone_slot":mp_slot,"priority":"MP"})

    # ── PM OL ────────────────────────────────────────────────────────
    # pm_idx is a *watermark* across slots: each TM gets at most one PMOL
    # placement attempt for the night (a placed TM can't fill another PMOL
    # slot, and a skipped one — placed elsewhere or not PM-OL eligible —
    # doesn't get re-tried). When pm_idx exhausts pmpool, remaining PMOL
    # slots in the loop go unfilled. Previously those silent skips never
    # made it into `unresolved`, so the audit's filled+unresolved didn't
    # account for them. Track per-slot placement and append to unresolved
    # when the slot exits without a fill.
    pm_idx = 0
    for slot in ["PMOL1","PMOL2","PMOL3","PMOL4","PMOL5","PMOL6"]:
        if slot not in cell_map.get(day, {}): continue
        _slot_filled = False
        while pm_idx < len(pmpool):
            rk = pmpool[pm_idx]; pm_idx += 1
            dn2 = roster[rk]["display_name"]
            if dn2 in placed: continue
            if not elig(rk, "PM OL"): continue
            write_cell(day, slot, dn2)
            placed.add(dn2)
            _record_placement(day, slot, dn2, "pm_ol", "PM-OL")
            _slot_filled = True
            break
        if not _slot_filled:
            unresolved.append({"date":str(d),"zone_slot":slot,"priority":"PM-OL"})

    # ── AM OL ────────────────────────────────────────────────────────
    # Same watermark + per-slot tracking pattern as PM OL above.
    am_idx = 0
    for slot in ["AMOL1","AMOL2","AMOL3","AMOL4","AMOL5","AMOL6"]:
        if slot not in cell_map.get(day, {}): continue
        _slot_filled = False
        while am_idx < len(ampool):
            rk = ampool[am_idx]; am_idx += 1
            dn2 = roster[rk]["display_name"]
            if dn2 in placed: continue
            if not elig(rk, "AM OL"): continue
            write_cell(day, slot, dn2)
            placed.add(dn2)
            _record_placement(day, slot, dn2, "am_ol", "AM-OL")
            _slot_filled = True
            break
        if not _slot_filled:
            unresolved.append({"date":str(d),"zone_slot":slot,"priority":"AM-OL"})

    # ── OVERFLOW (5/1/26) ────────────────────────────────────────────
    # Any unplaced grave-pool TMs land in the new conditional slots.
    # Routing:
    #   F/Sa: 1st overflow → Z9 SR Buddy (paired in smoke room),
    #         2nd overflow → Support 3
    #   All other days: 1st overflow → Support 3
    # Trainees (Seth, Trenidee) and no_sweeper holders are NOT preferred for
    # overflow — they ideally land in canonical slots first.
    unplaced = [tk for tk in gpool
                if roster[tk]["display_name"] not in placed
                and roster[tk]["display_name"] not in TRAINEE_DISPLAY]
    if unplaced:
        overflow_targets = ["Z9SRBuddy", "Support3"] if day in BUDDY_DAYS else ["Support3"]
        for target_slot in overflow_targets:
            if not unplaced: break
            if target_slot not in cell_map.get(day, {}): continue
            # 5/3/26 #11: Score-rank overflow candidates instead of popping by rank.
            # Z9 SR Buddy + Support 3 selection now respects fatigue, preferences,
            # pair affinity etc. — same scorecard the rest of the board uses.
            unplaced.sort(key=lambda rk: rotation_key(rk, target_slot, day_name=day))
            rk = unplaced.pop(0)
            dn2 = roster[rk]["display_name"]
            write_cell(day, target_slot, dn2)
            placed.add(dn2)
            _record_placement(day, target_slot, dn2, "grave", "Overflow")
            audit_items.append({"severity":"info","type":"OVERFLOW_PLACED",
                "tm_name":dn2,"detail":f"{day} {d}: {dn2} placed in {target_slot} "
                                       f"(overflow, scored selection)"})

    print(f"  {day[:3]} ({d}): placed {len(placed)} | "
          f"grave={len(gpool)} pm={len(pmpool)} am={len(ampool)}")

owb.save(OUTPUT_WB)
print(f"\n  Board saved.  Filled: {filled_count} | Unresolved: {len(unresolved)}")

# ── 7. AUDIT + ARCHIVE ───────────────────────────────────────────────
print("[7/7] Writing audit and archive...")
warnings_ = [i for i in audit_items if i["severity"]=="warning"]
errors_   = [i for i in audit_items if i["severity"]=="error"]

headcount = []
for day in DAYS:
    d = DAY_DATES[day]
    grave_sched  = len(daily_pools[d]["grave"])
    total_sched  = grave_sched + len(daily_pools[d]["pm_ol"]) + len(daily_pools[d]["am_ol"])
    # Goal comparison uses grave pool only — OL workers are not counted toward floor staffing goal
    grave_placed = sum(1 for p in placements if p["date"]==str(d) and p["pool_type"]=="grave")
    goal         = STAFFING_GOAL.get(day, 18)
    gap          = grave_placed - goal
    gap_str      = f"+{gap}" if gap >= 0 else str(gap)
    flag         = " ⚠" if gap < 0 else ""
    headcount.append(
        f"  {d} {day[:3]}: grave {grave_sched:2d} | total scheduled {total_sched:2d} | "
        f"grave placed {grave_placed:2d} | goal {goal} ({gap_str}){flag}"
    )

unres_detail = "\n".join(f"  {u['date']} {u['zone_slot']} ({u['priority']})" for u in unresolved)

md = f"""# Grave Deployment Audit — Week ending {WEEK_ENDING}

**Mode:** fill  **Run:** {RUN_TS[:19]}  **Engine:** v6

## Summary

| | |
|-|-|
| Total slots | {total_slots} |
| Filled | {filled_count} |
| Unresolved | {len(unresolved)} |
| Errors | {len(errors_)} |
| Warnings | {len(warnings_)} |

## Headcount

```
{chr(10).join(headcount)}
```

## Unresolved Slots ({len(unresolved)})

```
{unres_detail if unresolved else "None — board is clean."}
```

## ACTION REQUIRED

"""
crit = [u for u in unresolved if any(x in u["zone_slot"] for x in ["Admin","MRR","WRR"])]
if crit:
    md += "\n".join(f"- ⚠ {u['zone_slot']} unfilled on {u['date']}" for u in crit)
else:
    md += "None — all Admin and Restroom slots filled.\n"

md += "\n\n## Training Placements\n\n"
training_ps = [p for p in placements if p["pool_type"]=="training"]
if training_ps:
    md += "| Date | Slot | Role | TM |\n|------|------|------|----|\n"
    for p in training_ps:
        role = "Trainer" if "Trainer" in p["priority"] else "Trainee"
        md += f"| {p['date']} | {p['zone_slot']} | {role} | {p['tm_display_name']} |\n"
else:
    md += "_No training pairs active this week._\n"

md += "\n\n## Placement Log\n\n| Date | Slot | Priority | TM | Pool |\n|------|------|----------|----|------|\n"
for p in placements:
    md += f"| {p['date']} | {p['zone_slot']} | {p['priority']} | {p['tm_display_name']} | {p['pool_type']} |\n"

AUDIT_MD.write_text(md)

# 5/3/26 #9: Dedupe audit_items by (type, tm_name, detail) before writing.
# Repetitive warnings (same training rule firing on multiple trainees, etc.)
# would otherwise inflate the audit log without adding signal.
_seen = set()
_deduped_audit = []
for _item in audit_items:
    _key = (_item.get("type"), _item.get("tm_name"), _item.get("detail"))
    if _key in _seen: continue
    _seen.add(_key); _deduped_audit.append(_item)

# ── Severity tagging on unresolved (Brian's expected-unfilled vs. critical) ──
# An unresolved slot is `severity=low` if it's an auxiliary/overflow slot that
# regularly goes unfilled when the day's pool matches Brian's target headcount
# (Trash, MP, Support3, Z9SRBuddy). It's `severity=critical` otherwise — those
# are real shortages worth surfacing on the deployment book.
# This is the "static priority list" version. A future refinement could
# compute severity dynamically off TARGET_GRAVE_BY_DAY (i.e. tag any slot
# beyond rank=N as low when N grave bodies are scheduled), but the static
# version is easier to reason about and correct in the common case.
for _u in unresolved:
    _u["severity"] = "low" if _u.get("zone_slot") in LOW_PRIORITY_SLOTS else "critical"
_unres_critical = sum(1 for _u in unresolved if _u.get("severity") == "critical")
_unres_low      = len(unresolved) - _unres_critical

AUDIT_JSON.write_text(json.dumps({
    "week_ending":WEEK_ENDING,"run_timestamp":RUN_TS,"engine_version":"v8",
    "summary":{"total_slots":total_slots,"filled":filled_count,
               "unfilled":len(unresolved),
               "unfilled_critical":_unres_critical,
               "unfilled_low":_unres_low,
               "errors":len(errors_),"warnings":len(warnings_),
               "audit_items_total":len(audit_items),"audit_items_unique":len(_deduped_audit),
               "applied_overrides_count":len(_applied_override_ids)},
    # 5/3/26 #5: Capture engine config used for this run so the audit answers
    # "which weights produced this fill?" without spelunking the JSON files.
    "config":{
        "scorecard_weights":SCORECARD_WEIGHTS,
        "fatigue_window_days":FATIGUE_WINDOW_DAYS,
        "slot_loads_count":len(SLOT_LOADS),
        "buddy_days":sorted(BUDDY_DAYS),
        "rotation_weeks":ROTATION_WEEKS,
    },
    "training_schedule":TRAINING_SCHEDULE,
    "placements":placements,"audit_items":_deduped_audit,"unresolved_slots":unresolved,
    # Phase D — engine_overrides consumed this run
    "applied_override_ids":_applied_override_ids,
}, indent=2))

# ── Phase D: mark engine_overrides consumed this run as applied ───────────────
if _applied_override_ids:
    try:
        mark_engine_overrides_applied(_applied_override_ids)
        print(f"  engine_overrides: {len(_applied_override_ids)} override(s) marked applied.")
    except Exception as _mark_err:
        print(f"  [warn] Failed to mark overrides applied: {_mark_err}")

# ── ARCHIVE ──────────────────────────────────────────────────────────
(ARCHIVE_PATH.parent).mkdir(parents=True, exist_ok=True)
if ARCHIVE_PATH.exists():
    awb = openpyxl.load_workbook(ARCHIVE_PATH)
else:
    awb = openpyxl.Workbook(); awb.active.title = "Placements"

if "Placements" not in awb.sheetnames:
    ps = awb.create_sheet("Placements", 0)
    ps.append(["Week Ending","Date","Day","TM","Zone Slot","Priority","Pool Type","Run Timestamp"])
else:
    ps = awb["Placements"]
    rows_to_del = [r for r in ps.iter_rows(min_row=2)
                   if r[0].value and str(r[0].value)[:10] == WEEK_ENDING]
    for r in reversed(rows_to_del): ps.delete_rows(r[0].row)

for p in placements:
    day_name = DATE_TO_DAY.get(date.fromisoformat(p["date"]), "")
    ps.append([WEEK_ENDING, p["date"], day_name, p["tm_display_name"],
               p["zone_slot"], p["priority"], p["pool_type"], RUN_TS[:19]])

# TM Zone History sheet
if "TM Zone History" in awb.sheetnames: del awb["TM Zone History"]
hs = awb.create_sheet("TM Zone History")
all_zones = ["Admin","MRR1","MRR6","MRR7","MRR8","MRR10","WRR1","WRR6","WRR7","WRR8","WRR10",
             "Zone1","Zone2","Zone3","Zone4","Zone5","Zone6","Zone7","Zone8","Zone9","Zone10",
             "Zone9SR","Z9SRBuddy","Trash1","Trash2","MP1","MP2","Support3"]
hs.append(["TM"] + all_zones + ["Total Placements"])

from collections import defaultdict
tm_zone_dates = defaultdict(lambda: defaultdict(list))
for row in ps.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]: continue
    tm_n = str(row[3]).strip(); zone = str(row[4]).strip(); d_raw = row[1]
    try:
        d_val = d_raw if isinstance(d_raw, (date, datetime)) else date.fromisoformat(str(d_raw)[:10])
        tm_zone_dates[tm_n][zone].append(d_val)
    except: pass

for tm_n in sorted(tm_zone_dates.keys()):
    row_data = [tm_n]; total = 0
    for z in all_zones:
        dates = tm_zone_dates[tm_n].get(z, [])
        if dates:
            latest = max(dates)
            latest_str = latest.isoformat() if hasattr(latest,"isoformat") else str(latest)
            row_data.append(f"{len(dates)}x (last {latest_str})")
            total += len(dates)
        else: row_data.append("")
    row_data.append(total); hs.append(row_data)

awb.save(ARCHIVE_PATH)

print(f"\n{'='*62}")
print(f"DONE  |  Filled {filled_count}/{total_slots}  |  Unresolved: {len(unresolved)}")
print(f"{'='*62}")
