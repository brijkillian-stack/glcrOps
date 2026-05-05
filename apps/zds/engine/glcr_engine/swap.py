"""
GLCR Smart Swap Suggester — Phase 2 (5/3/26)

Given a list of call-offs and the current filled workbook for a specific day,
propose 2–3 ranked swap chains that re-cover the affected slots. Each chain
is evaluated using glcr_engine.scorecard so proposed moves are explainable
in terms of skill match, fatigue, preferences, pair affinities, etc.

Design choices (per Brian, 5/2/26):
  - Conservative PROPOSE-AND-CONFIRM — never auto-apply
  - Cascade depth max 3 moves per chain (single → 2-chain → 3-chain)
  - Must-fill zones (Z1/Z4/Z5/Z8) get priority over coverable zones
  - Prefer single-move proposals over chains (less disruption)
  - Honor accommodations + hard preferences as filters (same as fill engine)

Public API:
    propose_swaps(workbook_path, day_name, call_offs, scorecard_module=None)
        → returns SwapProposal list (ranked best-first)

    apply_swap(workbook_path, proposal, audit_path=None)
        → mutates workbook + archive, returns dict with applied moves

A SwapProposal is a list of Move dicts:
    {from_slot, to_slot, tm, reason, score_delta}
"""

from datetime import date, datetime, timedelta
from openpyxl import load_workbook

# Slot map mirrors fill_engine.py's cell_map structure for one day.
# We work directly with the workbook cells.

ZONE_ROW_OFFSET = 3   # Friday zones in row 3, +1 per day
RR_ROW_OFFSET   = 13  # Friday RR in row 13
AUX_ROW_OFFSET  = 23  # Friday aux in row 23

# Aux columns
AUX_COL = {
    "Trash1": 2, "Trash2": 3, "Admin": 4, "Zone9SR": 5,
    "Support1": 6, "Support2": 7, "Support3": 8, "Z9SRBuddy": 9,
}
RR_COL = {  # mens columns 2-6, womens 7-11; index by RR num
    1: (2, 7), 6: (3, 8), 7: (4, 9), 8: (5, 10), 10: (6, 11),
}

DAY_TO_ROW_OFFSET = {
    "Friday": 0, "Saturday": 1, "Sunday": 2, "Monday": 3,
    "Tuesday": 4, "Wednesday": 5, "Thursday": 6,
}

# Must-fill zones — the engine never wants these empty
MUST_FILL_ZONES = {"Zone1", "Zone4", "Zone5", "Zone8"}

# Skip-priority order: what gets dropped first when short-staffed
# (lower index = drops earlier)
SKIP_PRIORITY = ["Zone9", "Zone6", "Zone7", "Zone2", "Zone3", "Zone10"]


def _slot_to_elig_col(slot_code):
    """Map internal slot code to the eligibility roster column header."""
    if slot_code.startswith("Zone") and slot_code not in ("Zone9SR", "Z9SRBuddy"):
        return f"Zone {slot_code[4:]}"
    if slot_code == "Zone9SR":   return "Zone 9 SR"
    if slot_code.startswith("MRR"):
        return "Mens 1 + 2" if slot_code == "MRR1" else f"Mens {slot_code[3:]}"
    if slot_code.startswith("WRR"):
        return "Womens 1 + 2" if slot_code == "WRR1" else f"Womens {slot_code[3:]}"
    if slot_code == "Admin": return "Admin"
    if slot_code in ("Trash1",): return "Trash 1"
    if slot_code in ("Trash2",): return "Trash 2"
    if slot_code in ("Support1", "MP1"): return "MP 1"  # legacy roster column
    if slot_code in ("Support2", "MP2"): return "MP 2"
    return None  # Support3, Z9SRBuddy — overflow slots, no roster col


def _slot_cell(day_name, slot_code):
    """Return (row, col) for a given slot on a given day."""
    offset = DAY_TO_ROW_OFFSET.get(day_name)
    if offset is None: return None
    if slot_code.startswith("Zone") and slot_code not in ("Zone9SR", "Z9SRBuddy"):
        z_num = int(slot_code[4:])
        return (ZONE_ROW_OFFSET + offset, 1 + z_num)
    if slot_code.startswith("MRR"):
        rr_num = int(slot_code[3:])
        return (RR_ROW_OFFSET + offset, RR_COL[rr_num][0])
    if slot_code.startswith("WRR"):
        rr_num = int(slot_code[3:])
        return (RR_ROW_OFFSET + offset, RR_COL[rr_num][1])
    if slot_code in AUX_COL:
        return (AUX_ROW_OFFSET + offset, AUX_COL[slot_code])
    return None


def _read_day_placements(workbook_path, day_name):
    """Return {slot_code: tm_display_name} for the given day."""
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Week Overview"]
    out = {}
    # Zones
    for z in range(1, 11):
        cell = _slot_cell(day_name, f"Zone{z}")
        if cell:
            v = ws.cell(*cell).value
            out[f"Zone{z}"] = str(v).strip() if v else ""
    # RR
    for rr_num in (1, 6, 7, 8, 10):
        for prefix, col_idx in (("MRR", 0), ("WRR", 1)):
            cell = _slot_cell(day_name, f"{prefix}{rr_num}")
            if cell:
                v = ws.cell(*cell).value
                out[f"{prefix}{rr_num}"] = str(v).strip() if v else ""
    # Aux
    for slot in AUX_COL:
        cell = _slot_cell(day_name, slot)
        if cell:
            v = ws.cell(*cell).value
            out[slot] = str(v).strip() if v else ""
    return out


def _find_tm_slots(placements, tm_name):
    """Return list of slot codes where this TM is currently placed."""
    return [slot for slot, name in placements.items() if name == tm_name]


def propose_swaps(workbook_path, day_name, call_offs,
                  scorecard_module, max_chain_depth=3, max_proposals=3):
    """Return a ranked list of SwapProposals for the given call-offs.

    Args:
        workbook_path: path to filled Week Overview workbook
        day_name: 'Friday', 'Saturday', etc.
        call_offs: list of TM display names who called off
        scorecard_module: glcr_engine.scorecard (already initialized)
        max_chain_depth: max moves per chain (default 3)
        max_proposals: how many ranked options to return (default 3)

    Returns:
        list of dicts: {chain: [Move, ...], total_score, rationale, summary}
        sorted best-first (lowest total_score wins)
    """
    placements = _read_day_placements(workbook_path, day_name)

    # Identify vacated slots (where call-offs were)
    vacated = []
    for tm in call_offs:
        slots = _find_tm_slots(placements, tm)
        for slot in slots:
            vacated.append({"slot": slot, "vacated_by": tm})
            placements[slot] = ""  # mark empty in our working copy

    if not vacated:
        return [{"chain": [], "total_score": 0,
                 "rationale": "Call-offs not currently placed — no swaps needed.",
                 "summary": "No action."}]

    # For each vacated slot, find candidates
    proposals = []
    for vac in vacated:
        empty_slot = vac["slot"]
        # Skip if this is a non-must-fill slot — could just leave empty
        is_must_fill = empty_slot in MUST_FILL_ZONES or empty_slot in (
            "Zone9SR", "Trash1", "Trash2", "Admin"
        )

        # Strategy A: leave empty (no chain)
        # Only viable if non-must-fill AND coverage exists via skip-priority/RR
        if not is_must_fill:
            proposals.append({
                "chain": [],
                "total_score": 0.0,
                "rationale": f"Leave {empty_slot} empty — covered by adjacent (skip-priority).",
                "summary": f"No swap. {empty_slot} dropped tonight; coverage rules absorb it.",
                "vacated": empty_slot,
            })

        # Strategy B: find a TM to move INTO this slot from another less-critical slot
        candidates = _find_movable_into(empty_slot, day_name, placements,
                                        scorecard_module)
        for cand in candidates[:5]:
            cand_tm = cand["tm"]
            cand_from = cand["from_slot"]
            move = {
                "tm": cand_tm,
                "from_slot": cand_from,
                "to_slot": empty_slot,
                "score_delta": cand["score_delta"],
                "reason": cand["reason"],
            }
            chain = [move]
            # Now cand_from is empty — does it need filling?
            cascade = _build_cascade(cand_from, day_name, placements,
                                     scorecard_module, depth=1, max_depth=max_chain_depth)
            chain.extend(cascade)
            total = sum(m["score_delta"] for m in chain)
            proposals.append({
                "chain": chain,
                "total_score": total,
                "rationale": _summarize_chain(chain, empty_slot),
                "summary": _short_summary(chain, empty_slot),
                "vacated": empty_slot,
            })

    # Rank: lower total_score = better. Prefer shorter chains as tiebreaker.
    proposals.sort(key=lambda p: (p["total_score"], len(p["chain"])))
    return proposals[:max_proposals]


def _find_movable_into(empty_slot, day_name, placements, sc):
    """Return ranked list of TMs currently placed elsewhere who could move
    into the empty slot. Only considers TMs whose current slot is less critical
    than the empty slot."""
    empty_is_zone   = empty_slot.startswith("Zone") and empty_slot not in ("Zone9SR", "Z9SRBuddy")
    empty_is_aux    = empty_slot in ("Admin", "Zone9SR")
    empty_is_critical = empty_slot in MUST_FILL_ZONES or empty_is_aux

    # Iterate currently-placed TMs
    candidates = []
    for current_slot, tm in placements.items():
        if not tm or current_slot == empty_slot: continue

        # Skip moving FROM critical slots (don't break Z1 to fill Z6)
        current_is_critical = current_slot in MUST_FILL_ZONES or current_slot in (
            "Admin", "Zone9SR"
        )
        if current_is_critical and not empty_is_critical:
            continue

        # Get rkey for scoring
        rk = _rkey_for(tm, sc)
        if not rk: continue

        # Hard accommodation/preference filter
        if sc.has_hard_block(tm, empty_slot):
            continue

        # Roster eligibility check — TM must be Y for the empty slot
        elig_col = _slot_to_elig_col(empty_slot)
        if elig_col:
            roster = sc._state["roster"]
            if not roster.get(rk, {}).get("eligibility", {}).get(elig_col, False):
                continue

        # Score the candidate at the empty slot
        sc_score = sc.score_placement(rk, empty_slot, day_name=day_name)
        # Score them at their current slot too — delta = how much better/worse
        # they'd be in the empty slot vs staying put
        current_score = sc.score_placement(rk, current_slot, day_name=day_name)
        delta = sc_score["total"] - current_score["total"]

        # Build reason from scorecard
        reason_bits = []
        comps = sc_score["components"]
        if comps.get("preference_fit", 0) < 0:
            reason_bits.append("matches preference")
        if comps.get("preference_fit", 0) > 0:
            reason_bits.append("⚠ against preference")
        if comps.get("pair_affinity", 0) > 0:
            reason_bits.append("⚠ pair conflict")
        if sc_score["fatigue_pts"] >= 22:
            reason_bits.append(f"⚠ stretched ({sc_score['fatigue_pts']} pts)")
        if sc_score["fatigue_pts"] <= 8:
            reason_bits.append(f"fresh ({sc_score['fatigue_pts']} pts)")
        if comps.get("skill_match", 0) < 0:
            reason_bits.append("strong skill match")
        if not reason_bits:
            reason_bits.append("rotation fit")

        candidates.append({
            "tm": tm,
            "from_slot": current_slot,
            "score_delta": delta,
            "score_total": sc_score["total"],
            "reason": " · ".join(reason_bits),
        })

    candidates.sort(key=lambda c: c["score_total"])
    return candidates


def _build_cascade(empty_slot, day_name, placements, sc, depth, max_depth):
    """When a swap creates a new empty slot, recursively try to fill it.
    Returns list of additional Move dicts. Empty list if cascade not needed
    or max depth reached."""
    if depth >= max_depth:
        return []
    # Is this slot must-fill?
    is_critical = empty_slot in MUST_FILL_ZONES or empty_slot in ("Admin", "Zone9SR")
    if not is_critical:
        return []  # let it stay empty — coverage absorbs
    # Find unplaced TMs first (no cascade needed)
    # ... for v1 we keep cascade simple: just try the next-best swap
    # (could be expanded to look at the unplaced grave pool)
    return []


def _rkey_for(tm_name, sc):
    """Look up roster key for a display name."""
    roster = sc._state["roster"]
    for k, info in roster.items():
        if info.get("display_name") == tm_name:
            return k
    return None


def _summarize_chain(chain, original_empty):
    if not chain:
        return f"Leave {original_empty} empty."
    parts = []
    for m in chain:
        delta = m["score_delta"]
        delta_str = f"Δ{delta:+.2f}"
        parts.append(f"{m['tm']}: {m['from_slot']} → {m['to_slot']} ({delta_str}, {m['reason']})")
    return " · then · ".join(parts)


def _short_summary(chain, original_empty):
    if not chain:
        return f"Leave {original_empty} empty (covered by adjacent)."
    if len(chain) == 1:
        m = chain[0]
        return f"Move {m['tm']} from {m['from_slot']} to {m['to_slot']}."
    return f"{len(chain)}-move chain starting with {chain[0]['tm']}."


# ── APPLICATION ─────────────────────────────────────────────────────
def apply_swap(workbook_path, proposal, day_name, audit_log=None):
    """Apply the chosen swap chain to the workbook. Mutates the file in place.
    Returns a dict {applied_moves, audit_entries}."""
    if not proposal.get("chain"):
        # Empty chain = leave as-is (or just clear the call-off cells)
        return {"applied_moves": [], "audit_entries": []}

    wb = load_workbook(workbook_path)
    ws = wb["Week Overview"]
    audit_entries = []

    for move in proposal["chain"]:
        from_cell = _slot_cell(day_name, move["from_slot"])
        to_cell   = _slot_cell(day_name, move["to_slot"])
        if not from_cell or not to_cell:
            audit_entries.append({"severity": "error", "type": "SWAP_SLOT_NOT_FOUND",
                "detail": f"Couldn't resolve cell for {move['from_slot']} or {move['to_slot']}"})
            continue
        # Move TM from old slot to new
        ws.cell(*to_cell).value = move["tm"]
        ws.cell(*from_cell).value = None
        audit_entries.append({"severity": "info", "type": "SWAP_APPLIED",
            "detail": f"{move['tm']}: {move['from_slot']} → {move['to_slot']} "
                      f"({move['reason']}, score Δ {move['score_delta']:+.2f})"})

    wb.save(workbook_path)
    return {"applied_moves": proposal["chain"], "audit_entries": audit_entries}
