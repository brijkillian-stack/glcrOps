"""
seed_week.py — Populate Supabase from a Week Overview Excel workbook.

Actual xlsx structure (3 vertically-stacked sections):

  Zone section:  row 1  = headers (ZONE 1–10, PM OL 1–6)
                 row 2  = DATE label row
                 rows 3–9 = night data (Thu→Wed, day_num 1–7)
                   col 2–11  = zone_1 through zone_10
                   col 14–19 = PM overlap TMs 1–6

  RR section:    row 11 = headers (MENS 1+2, MENS 6, ... WOMENS 10, AM OL 1–6)
                 row 12 = DATE label row
                 rows 13–19 = night data (Thu→Wed)
                   col 2–6   = rr_1_2/rr_6/rr_7/rr_8/rr_10  (mens)
                   col 7–11  = rr_1_2/rr_6/rr_7/rr_8/rr_10  (womens)
                   col 14–19 = AM overlap TMs 1–6

  Aux section:   row 21 = headers (TRASH, TRASH, ADMIN, Z9 SR, SUPPORT 1, SUPPORT 2, SUPPORT 3)
                 row 22 = DATE label row
                 rows 23–29 = night data (Thu→Wed)
                   col 2 = trash_1, col 3 = trash_2, col 4 = admin, col 5 = z9_sr
                   col 6 = support_1, col 7 = support_2, col 8 = support_3

Usage:
    python3 seed_week.py --xlsx "/path/to/Week Overview - Filled - 2026-05-07.xlsx" \\
                         --week-ending 2026-05-07

    # Preview without writing to Supabase:
    python3 seed_week.py --xlsx "..." --week-ending 2026-05-07 --dry-run
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date, timedelta

from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

# ── Supabase client ───────────────────────────────────────────────────────────
from supabase import create_client

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_SERVICE_KEY = os.environ["SUPABASE_SERVICE_KEY"]
db = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)


# ── Section layout constants (1-indexed xlsx columns) ─────────────────────────

# Zone section — rows 3-9
ZONE_DATA_START_ROW = 3   # Thu = row 3, Fri = 4, ..., Wed = 9
ZONE_COLS = {             # col index → (slot_key, rr_side, sort_order)
    2:  ("zone_1",  None, 1),
    3:  ("zone_2",  None, 2),
    4:  ("zone_3",  None, 3),
    5:  ("zone_4",  None, 4),
    6:  ("zone_5",  None, 5),
    7:  ("zone_6",  None, 6),
    8:  ("zone_7",  None, 7),
    9:  ("zone_8",  None, 8),
    10: ("zone_9",  None, 9),
    11: ("zone_10", None, 10),
}
PM_OL_COLS = [14, 15, 16, 17, 18, 19]   # PM overlap TM columns

# RR section — rows 13-19
RR_DATA_START_ROW = 13
RR_COLS = {               # col index → (slot_key, rr_side, sort_order)
    2:  ("rr_1_2", "mens",   11),
    3:  ("rr_6",   "mens",   13),
    4:  ("rr_7",   "mens",   15),
    5:  ("rr_8",   "mens",   17),
    6:  ("rr_10",  "mens",   19),
    7:  ("rr_1_2", "womens", 12),
    8:  ("rr_6",   "womens", 14),
    9:  ("rr_7",   "womens", 16),
    10: ("rr_8",   "womens", 18),
    11: ("rr_10",  "womens", 20),
}
AM_OL_COLS = [14, 15, 16, 17, 18, 19]   # AM overlap TM columns

# Aux section — rows 23-29
AUX_DATA_START_ROW = 23
AUX_COLS = {              # col index → (slot_key, sort_order)
    2: ("trash_1",   23),
    3: ("trash_2",   24),
    4: ("admin",     22),
    5: ("z9_sr",     21),
    6: ("support_1", 25),
    7: ("support_2", 26),
    8: ("support_3", 27),
}

# Day sequence (day_num 1–7 = Thu–Wed)
DAY_SEQUENCE = [
    (1, "Thursday"),
    (2, "Friday"),
    (3, "Saturday"),
    (4, "Sunday"),
    (5, "Monday"),
    (6, "Tuesday"),
    (7, "Wednesday"),
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _clean(val) -> str:
    """Return stripped string or empty string for None / whitespace-only."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("none", "nan", "unfilled") else s


def _build_tm_map() -> dict[str, str]:
    """Return {display_name: entity_id} for all active TMs (case-insensitive lookup)."""
    res = db.table("entities").select("id, display_name").eq("status", "active").execute()
    return {row["display_name"]: row["id"] for row in (res.data or [])}


def _tm_id(name: str, tm_map: dict[str, str]) -> str | None:
    """Resolve display_name → entity id; warn if not found."""
    name = _clean(name)
    if not name:
        return None
    # Exact match first
    if name in tm_map:
        return tm_map[name]
    # Case-insensitive fallback
    lower = name.lower()
    for k, v in tm_map.items():
        if k.lower() == lower:
            return v
    print(f"  ⚠ TM not found in entities: {name!r}", file=sys.stderr)
    return None


def _cell(ws, row: int, col: int):
    """Return cell value (or None) at 1-indexed row, col."""
    return ws.cell(row=row, column=col).value


# ── xlsx parser ───────────────────────────────────────────────────────────────

class NightData:
    """All parsed assignments for one night."""
    def __init__(self, day_num: int, day_name: str, night_date: str):
        self.day_num    = day_num
        self.day_name   = day_name
        self.night_date = night_date
        self.zones: list[tuple]    = []  # (slot_key, rr_side, sort_order, tm_name)
        self.rrs:   list[tuple]    = []
        self.auxs:  list[tuple]    = []
        self.pm_ol: list[str]      = []  # TM names, up to 6
        self.am_ol: list[str]      = []


def _parse_xlsx(xlsx_path: str, week_ending: str) -> list[NightData]:
    """Parse the three sections and return one NightData per night (Thu→Wed)."""
    wb = load_workbook(xlsx_path, data_only=True)
    # Try named sheet, fall back to first
    ws = wb["Week Overview"] if "Week Overview" in wb.sheetnames else wb.active

    we  = date.fromisoformat(week_ending)   # Wednesday
    thu = we - timedelta(days=6)            # Thursday of same week

    nights: list[NightData] = []
    for offset, (day_num, day_name) in enumerate(DAY_SEQUENCE):
        night_date = (thu + timedelta(days=offset)).isoformat()
        nd = NightData(day_num, day_name, night_date)

        zone_row = ZONE_DATA_START_ROW + offset
        rr_row   = RR_DATA_START_ROW   + offset
        aux_row  = AUX_DATA_START_ROW  + offset

        # ── Zone assignments ──────────────────────────────────────────────────
        for col, (slot_key, rr_side, sort_order) in ZONE_COLS.items():
            nd.zones.append((slot_key, rr_side, sort_order, _clean(_cell(ws, zone_row, col))))

        # ── RR assignments ────────────────────────────────────────────────────
        for col, (slot_key, rr_side, sort_order) in RR_COLS.items():
            nd.rrs.append((slot_key, rr_side, sort_order, _clean(_cell(ws, rr_row, col))))

        # ── Aux assignments ───────────────────────────────────────────────────
        for col, (slot_key, sort_order) in AUX_COLS.items():
            nd.auxs.append((slot_key, sort_order, _clean(_cell(ws, aux_row, col))))

        # ── Overlaps ──────────────────────────────────────────────────────────
        nd.pm_ol = [_clean(_cell(ws, zone_row, c)) for c in PM_OL_COLS]
        nd.am_ol = [_clean(_cell(ws, rr_row,   c)) for c in AM_OL_COLS]

        nights.append(nd)

    return nights


# ── Supabase writer ───────────────────────────────────────────────────────────

def _upsert_zone_slot(night_id: str, slot_type: str, slot_key: str,
                      rr_side, sort_order: int, tm_id, dry_run: bool):
    if dry_run:
        return
    db.table("zone_assignments").upsert({
        "night_id":   night_id,
        "slot_type":  slot_type,
        "slot_key":   slot_key,
        "rr_side":    rr_side,
        "sort_order": sort_order,
        "tm_id":      tm_id,
        "is_filled":  tm_id is not None,
        "is_empty":   tm_id is None,
    }, on_conflict="night_id,slot_type,slot_key,rr_side").execute()


def _upsert_overlap(night_id: str, tm_id, overlap_window: str,
                    position: int, dry_run: bool):
    if dry_run or tm_id is None:
        return
    db.table("overlap_assignments").upsert({
        "night_id":        night_id,
        "tm_id":           tm_id,
        "overlap_window":  overlap_window,
        "position":        position,
        "task":            None,   # task column filled separately if needed
    }, on_conflict="night_id,overlap_window,position").execute()


# ── Main seed function ────────────────────────────────────────────────────────

def seed(xlsx_path: str, week_ending: str, dry_run: bool = False) -> None:
    tm_map = _build_tm_map()
    print(f"✓ Loaded {len(tm_map)} TMs from Supabase")

    nights = _parse_xlsx(xlsx_path, week_ending)
    print(f"✓ Parsed {len(nights)} nights from {xlsx_path}")

    # ── Upsert week ───────────────────────────────────────────────────────────
    label = f"Week of {week_ending}"
    if not dry_run:
        res = (
            db.table("weeks")
            .upsert({"week_ending": week_ending, "label": label, "status": "draft"},
                    on_conflict="week_ending")
            .execute()
        )
        week_id = res.data[0]["id"]
        print(f"✓ Week upserted: {week_id}")
    else:
        week_id = "DRY-RUN"
        print(f"[dry-run] Would upsert week {week_ending}")

    # ── Iterate nights ────────────────────────────────────────────────────────
    for nd in nights:
        page_num = nd.day_num * 2 - 1   # deployment pages 1,3,5,7,9,11,13

        # Count filled zone slots as in_rotation proxy
        filled_zones = sum(1 for (_, _, _, tm_name) in nd.zones if tm_name)

        print(f"\n  Night: {nd.day_name} {nd.night_date} | zones_filled={filled_zones}"
              f" | pm_ol={sum(1 for x in nd.pm_ol if x)}"
              f" | am_ol={sum(1 for x in nd.am_ol if x)}")

        if dry_run:
            for slot_key, rr_side, sort_order, tm_name in nd.zones:
                if tm_name:
                    tid = _tm_id(tm_name, tm_map)
                    status = "✓" if tid else "⚠ not found"
                    print(f"    [zone] {slot_key:8s}  {tm_name:<16s} {status}")
            for slot_key, rr_side, sort_order, tm_name in nd.rrs:
                if tm_name:
                    tid = _tm_id(tm_name, tm_map)
                    status = "✓" if tid else "⚠ not found"
                    print(f"    [rr  ] {slot_key:8s} ({rr_side:6s}) {tm_name:<16s} {status}")
            for slot_key, sort_order, tm_name in nd.auxs:
                if tm_name:
                    tid = _tm_id(tm_name, tm_map)
                    status = "✓" if tid else "⚠ not found"
                    print(f"    [aux ] {slot_key:10s} {tm_name:<16s} {status}")
            for i, tm_name in enumerate(nd.pm_ol, 1):
                if tm_name:
                    tid = _tm_id(tm_name, tm_map)
                    status = "✓" if tid else "⚠ not found"
                    print(f"    [PM OL {i}]          {tm_name:<16s} {status}")
            for i, tm_name in enumerate(nd.am_ol, 1):
                if tm_name:
                    tid = _tm_id(tm_name, tm_map)
                    status = "✓" if tid else "⚠ not found"
                    print(f"    [AM OL {i}]          {tm_name:<16s} {status}")
            continue

        # Upsert night row
        res = (
            db.table("nights")
            .upsert({
                "week_id":     week_id,
                "night_date":  nd.night_date,
                "day_name":    nd.day_name,
                "day_num":     nd.day_num,
                "page_num":    page_num,
                "in_rotation": filled_zones,
                "breaks_5":    0,
                "breaks_9":    0,
                "breaks_4":    0,
            }, on_conflict="week_id,night_date")
            .execute()
        )
        night_id = res.data[0]["id"]

        # Zone assignments
        for slot_key, rr_side, sort_order, tm_name in nd.zones:
            tm_id = _tm_id(tm_name, tm_map)
            _upsert_zone_slot(night_id, "zone", slot_key, rr_side, sort_order, tm_id, dry_run)

        # RR assignments
        for slot_key, rr_side, sort_order, tm_name in nd.rrs:
            tm_id = _tm_id(tm_name, tm_map)
            _upsert_zone_slot(night_id, "rr", slot_key, rr_side, sort_order, tm_id, dry_run)

        # Aux assignments
        for slot_key, sort_order, tm_name in nd.auxs:
            tm_id = _tm_id(tm_name, tm_map)
            _upsert_zone_slot(night_id, "aux", slot_key, None, sort_order, tm_id, dry_run)

        # PM overlaps
        for i, tm_name in enumerate(nd.pm_ol, 1):
            tm_id = _tm_id(tm_name, tm_map)
            _upsert_overlap(night_id, tm_id, "pm", i, dry_run)

        # AM overlaps
        for i, tm_name in enumerate(nd.am_ol, 1):
            tm_id = _tm_id(tm_name, tm_map)
            _upsert_overlap(night_id, tm_id, "am", i, dry_run)

        total_slots = len(nd.zones) + len(nd.rrs) + len(nd.auxs)
        print(f"    ✓ {total_slots} slots + overlaps upserted for {nd.day_name}")

    print("\n✅ Seed complete.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Seed Supabase from Week Overview XLSX")
    parser.add_argument("--xlsx",        required=True, help="Path to Week Overview Filled .xlsx")
    parser.add_argument("--week-ending", required=True, help="Week-ending date YYYY-MM-DD (Wednesday)")
    parser.add_argument("--dry-run",     action="store_true",
                        help="Parse and print without writing to Supabase")
    args = parser.parse_args()

    seed(args.xlsx, args.week_ending, dry_run=args.dry_run)
